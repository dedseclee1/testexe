# -*- coding: utf-8 -*-
# 建议在文件开头添加这个编码声明

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl.styles import Alignment, Font, PatternFill # 导入样式
from openpyxl.utils import get_column_letter # Excel列操作工具
import os
import datetime
import time
import traceback # 用于错误输出

from pathlib import Path # 用于获取桌面路径

# --- 检查并导入必要的库 ---
libraries_ok = True
try:
    import pandas as pd
except ImportError:
    messagebox.showerror("依赖缺失", "需要安装 'pandas' 库。\n请在命令行运行: pip install pandas")
    libraries_ok = False
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    messagebox.showerror("依赖缺失", "需要安装 'openpyxl' 库。\n请在命令行运行: pip install openpyxl")
    libraries_ok = False
try:
    import pyodbc
except ImportError:
    messagebox.showerror("依赖缺失", "需要安装 'pyodbc' 库。\n请在命令行运行: pip install pyodbc")
    libraries_ok = False
# <<< 新增 xlrd 检查 >>>
try:
    import xlrd
except ImportError:
    messagebox.showerror("依赖缺失", "需要安装 'xlrd' 库 (用于读取 .xls 文件)。\n请在命令行运行: pip install xlrd")
    libraries_ok = False
# <<< 检查结束 >>>

# --- 数据库连接字符串 ---
# --- !!! 请根据你的环境修改下面的 pyodbc 连接字符串 !!! ---
PYODBC_CONN_STRING = "DRIVER={ODBC Driver 17 for SQL Server};SERVER=192.168.0.117;DATABASE=FQD;UID=test;PWD=Forcome123;"


# --- “维护数据”功能的核心处理逻辑 ---
# <<< MODIFIED: 函数已修改以添加产量比较筛选 >>>
def process_assembly_data_preserve_macros(source_file, target_xlsm_file, assembly_filter_value, header_row_index):
    """
    读取源Excel文件，根据选择的车间筛选数据 (包括深加工车间)，
    然后筛选掉“预计产量”等于“已完工数量”的行，
    计算所需日期，并将结果更新到目标XLSM文件的活动工作表中，
    同时保留该XLSM文件中的宏。(修改版：增加产量比较筛选)
    """
    start_time = time.time()
    print(f"开始 'process_assembly_data_preserve_macros'...")
    print(f"  源文件: {source_file}")
    print(f"  目标文件: {target_xlsm_file}")
    print(f"  车间筛选: {assembly_filter_value}")
    print(f"  表头行索引: {header_row_index}")

    # --- 1. 读取和处理源数据 ---
    df_target = pd.DataFrame()
    try:
        print(f"  信息：正在从第 {header_row_index + 1} 行读取表头...")
        read_start = time.time()

        # <<< MODIFIED: 维护功能需要增加 '已完工数量' 列用于比较 >>>
        required_source_cols = ['车间', '工单单号', '单别', '品号', '品名', '物料描述',
                                '预计产量', '已完工数量', # <<< 新增：读取已完工数量 >>>
                                '开工日期']
        print(f"  信息：[维护功能] 将尝试读取以下列：{required_source_cols}")
        # <<< MODIFICATION END >>>

        df_source = None
        # 尝试用 pandas 默认（底层可能是 openpyxl 或其他）
        try:
            print(f"  信息：尝试使用 pandas 默认引擎读取...")
            df_source = pd.read_excel(
                source_file, sheet_name=0, header=header_row_index, engine=None,
                usecols=lambda x: x in required_source_cols
            )
            missing_cols_check = [col for col in required_source_cols if col not in df_source.columns]
            if missing_cols_check:
                print(f"  警告：读取后，以下维护必需列缺失：{missing_cols_check}")
                # <<< MODIFIED: 如果缺少任何必需列，则报错并退出，因为数量比较也需要这些列 >>>
                messagebox.showerror("列缺失错误 (维护)", f"源文件中找不到以下维护所需的列：\n{', '.join(missing_cols_check)}\n请检查源文件和表头行号。")
                return False # 直接返回失败
                # df_source = None # 不再设置为None，直接返回
                # <<< MODIFICATION END >>>
            else: print(f"  信息：使用 pandas 默认引擎读取成功！")
        except Exception as e: print(f"  警告：使用 pandas 默认引擎读取失败：{e}")

        # --- 如果默认引擎失败，尝试 openpyxl (逻辑保持，但增加了错误处理) ---
        if df_source is None: # 仅在默认引擎失败时尝试
            try:
                print(f"  信息：尝试显式使用 openpyxl 引擎读取...")
                df_source = pd.read_excel(
                    source_file, sheet_name=0, header=header_row_index, engine='openpyxl',
                    usecols=lambda x: x in required_source_cols
                )
                missing_cols_check = [col for col in required_source_cols if col not in df_source.columns]
                if missing_cols_check:
                    print(f"  警告：使用 openpyxl 引擎读取后，以下维护列缺失：{missing_cols_check}")
                    messagebox.showerror("列缺失错误 (维护)", f"使用 openpyxl 引擎读取时，源文件中找不到以下维护所需的列：\n{', '.join(missing_cols_check)}\n请检查源文件和表头行号。")
                    return False # 直接返回失败
                else: print(f"  信息：使用 openpyxl 引擎读取成功！")
            except Exception as e_openpyxl:
                print(f"  警告：显式使用 openpyxl 引擎读取也失败：{e_openpyxl}")
                messagebox.showerror("读取错误", f"尝试使用两种引擎都无法读取源文件维护所需列。\n请检查文件、表头行号({header_row_index + 1})及列名({required_source_cols})是否存在。\n错误详情: {e_openpyxl}")
                return False # 读取失败，返回

        # --- 列存在性最终检查 (再次确认) ---
        # Check again, although the previous checks should have caught issues.
        final_missing_cols = [col for col in required_source_cols if col not in df_source.columns]
        if final_missing_cols:
            messagebox.showerror("列缺失错误 (维护)", f"最终确认时，源文件中仍缺少以下维护所需的列：\n{', '.join(final_missing_cols)}")
            return False
        # if '车间' not in df_source.columns: # 之前的检查已经包含这个了
        #     messagebox.showerror("列缺失错误", "源文件中找不到 '车间' 列，无法进行筛选。")
        #     return False

        read_duration = time.time() - read_start
        print(f"  信息：读取源文件耗时：{read_duration:.2f} 秒，共 {len(df_source)} 行数据。")

        # --- 步骤 1b: 按车间筛选 (已根据 '生产的' 归属更新逻辑) --- # <<< MODIFIED COMMENT >>>
        filter_start = time.time()
        df_source['车间'] = df_source['车间'].astype(str).str.strip() # 确保比较前是字符串并去除首尾空格

        if assembly_filter_value == "深加工车间":
            print(f"  信息：[维护功能] 筛选 深加工车间 (排除 组装一, 组装二, 生产)...") # <<< MODIFIED LOG >>>
            # 排除列表增加 '生产的'
            exclusion_list = ["组装一", "组装二", "生产"] # <<< MODIFIED >>>
            df_filtered = df_source[~df_source['车间'].isin(exclusion_list)].copy()
        elif assembly_filter_value == "组装二":
            print(f"  信息：[维护功能] 筛选 组装二 (包括 组装二, 生产)...") # <<< MODIFIED LOG >>>
            # 包含列表定义为 '组装二' 和 '生产的'
            inclusion_list_zj2 = ["组装二", "生产"] # <<< MODIFIED >>>
            df_filtered = df_source[df_source['车间'].isin(inclusion_list_zj2)].copy() # <<< MODIFIED >>>
        elif assembly_filter_value == "组装一":
             # 组装一保持精确匹配逻辑
             print(f"  信息：[维护功能] 筛选 组装一 (精确匹配)...") # <<< ADDED/MODIFIED LOG >>>
             df_filtered = df_source[df_source['车间'] == assembly_filter_value].copy() # <<< Logic unchanged but now explicit >>>
        else:
             # 处理未知的车间选项
             messagebox.showerror("错误", f"未知的车间筛选选项: {assembly_filter_value}")
             print(f"  错误: 未知的车间筛选选项: {assembly_filter_value}")
             return False # 发生错误则返回失败

        print(f"  信息：按车间筛选完成，找到 {len(df_filtered)} 条记录，耗时：{time.time() - filter_start:.2f} 秒。")

        # --- 步骤 1c: <<< 新增筛选：预计产量 不等于 已完工数量 >>> ---
        filter_qty_start = time.time()
        print("  信息：应用额外筛选：预计产量不等于已完工数量...")
        # 检查必需的列是否存在于筛选后的 DataFrame 中
        if '预计产量' in df_filtered.columns and '已完工数量' in df_filtered.columns:
            try:
                # 将比较列转换为数值类型，无法转换的变为 NaN，然后用 0 填充 NaN
                # 这样做是为了确保可以进行数学比较，并处理 Excel 中可能的文本或空值
                qty_produced_numeric = pd.to_numeric(df_filtered['预计产量'], errors='coerce').fillna(0)
                qty_completed_numeric = pd.to_numeric(df_filtered['已完工数量'], errors='coerce').fillna(0)

                # 创建筛选条件：预计产量 不等于 已完工数量
                non_equal_mask = qty_produced_numeric != qty_completed_numeric

                original_count = len(df_filtered)
                # 应用筛选条件
                df_filtered = df_filtered[non_equal_mask]
                new_count = len(df_filtered)
                removed_count = original_count - new_count
                if removed_count > 0:
                    print(f"  信息：数量筛选完成，移除了 {removed_count} 行 (预计产量 == 已完工数量)。")
                else:
                    print(f"  信息：数量筛选完成，没有行因产量相等被移除。")
                print(f"  信息：数量筛选后剩余 {new_count} 条记录，耗时：{time.time() - filter_qty_start:.2f} 秒。")

            except Exception as qty_err:
                 # 如果数值转换或比较时出错
                 messagebox.showerror("数量筛选错误", f"在比较'预计产量'和'已完工数量'时发生错误:\n{qty_err}\n\n将跳过此筛选步骤。")
                 print(f"  错误：数量比较筛选失败: {qty_err}")
                 traceback.print_exc()
                 # 决定是否继续，这里选择继续，但跳过筛选
        else:
            # 如果需要的列在 df_filtered 中缺失（理论上不应发生，因为已在读取时检查）
            missing_qty_cols = [col for col in ['预计产量', '已完工数量'] if col not in df_filtered.columns]
            messagebox.showwarning("列缺失警告 (筛选)", f"无法执行'预计产量 != 已完工数量'筛选，因为以下列在筛选数据中缺失: {missing_qty_cols}\n可能是读取或初步筛选步骤有问题。将跳过此筛选。")
            print(f"  警告：跳过数量比较筛选，因为列 {missing_qty_cols} 在 df_filtered 中缺失。")
        # <<< 新增筛选结束 >>>

        # --- 步骤 1d: 准备目标 DataFrame (基于最终筛选结果) ---
        print("  信息：准备目标 DataFrame...")
        prep_start = time.time()
        if not df_filtered.empty:
             # 注意：这里的 column_mapping 不包含 '已完工数量'，因为它不需要写入目标文件
             column_mapping = {
                '工单单号': '工单单号', '单别': '工单单别', '品号': '产品品号',
                '品名': '品名', '物料描述': '规格', '预计产量': '预计产量',
                '开工日期': '开工日期' # 保留开工日期用于计算
            }
             # 只选取需要的列进行重命名和后续处理
             # <<< MODIFIED: 确保从最终的 df_filtered 中选取 >>>
             df_target = df_filtered[list(column_mapping.keys())].rename(columns=column_mapping).copy()

             # --- 日期计算逻辑 (保持不变) ---
             if '开工日期' in df_target.columns:
                # 先尝试转换为日期时间，无效的转为NaT
                df_target['开工日期'] = pd.to_datetime(df_target['开工日期'], errors='coerce')

                # 创建新日期列，初始化为NaT
                df_target['备料完成时间'] = pd.NaT
                df_target['物料到齐时间'] = pd.NaT

                # 仅对有效的开工日期进行计算
                valid_dates_mask = df_target['开工日期'].notna()
                df_target.loc[valid_dates_mask, '备料完成时间'] = df_target.loc[valid_dates_mask, '开工日期'] - pd.Timedelta(days=1)

                # 仅对有效的备料完成时间进行计算
                valid_prep_dates_mask = df_target['备料完成时间'].notna()
                df_target.loc[valid_prep_dates_mask, '物料到齐时间'] = df_target.loc[valid_prep_dates_mask, '备料完成时间'] - pd.Timedelta(days=2)

                # 格式化日期列为 yyyy-mm-dd 字符串或None (保持不变)
                date_cols_to_format = ['开工日期', '备料完成时间', '物料到齐时间']
                for col in date_cols_to_format:
                    if col in df_target.columns:
                         # 先转日期对象，再格式化为字符串，NaT转为None
                         df_target[col] = pd.to_datetime(df_target[col], errors='coerce').dt.date.astype(object).where(pd.notna(df_target[col]), None)

             else:
                 print("  警告：源数据中缺少 '开工日期' 列，无法计算相关日期。")
                 # 确保列存在，即使是None (保持不变)
                 if '开工日期' not in df_target.columns: df_target['开工日期'] = None
                 if '备料完成时间' not in df_target.columns: df_target['备料完成时间'] = None
                 if '物料到齐时间' not in df_target.columns: df_target['物料到齐时间'] = None
        else:
             # <<< MODIFIED: 更新提示信息，可能因为车间或产量筛选后为空 >>>
             print("  信息：经过车间和产量筛选后，无数据需要维护。")
        print(f"  信息：准备目标 DataFrame 耗时：{time.time() - prep_start:.2f} 秒")

    except Exception as e:
        # 源数据处理的通用错误处理 (保持不变)
        traceback_str = traceback.format_exc()
        messagebox.showerror("源数据处理错误 (维护)", f"处理源文件时出错：\n{e}\n\n详细信息:\n{traceback_str}")
        return False

    # --- 2. 更新目标 XLSM 文件 (这部分逻辑保持不变) ---
    #    (加载XLSM, 读取目标表头, 调整DataFrame列序, 清空旧数据, 写入新数据, 设置格式, 保存)
    workbook = None
    try:
        print(f"  信息：加载目标 XLSM 文件 '{os.path.basename(target_xlsm_file)}' (保留VBA)...")
        load_start = time.time()
        workbook = openpyxl.load_workbook(target_xlsm_file, read_only=False, keep_vba=True)
        try:
            sheet = workbook.active
            if sheet is None: raise ValueError("无法获取活动工作表")
            print(f"  信息：目标工作表 '{sheet.title}' 已找到。")
        except Exception as e_sheet:
             messagebox.showerror("工作表错误", f"无法在目标文件中找到或访问活动工作表：{e_sheet}")
             if workbook: workbook.close()
             return False
        print(f"  信息：加载目标 XLSM 文件耗时：{time.time() - load_start:.2f} 秒")

        print("  信息：读取目标文件表头...")
        read_header_start = time.time()
        try:
            # 读取目标文件的第一行作为表头
            actual_target_headers = [cell.value for cell in sheet[1] if cell.value is not None]
            if not actual_target_headers:
                 messagebox.showerror("目标文件错误", f"无法读取目标文件 '{os.path.basename(target_xlsm_file)}' 的表头（第1行为空或无效）。")
                 workbook.close()
                 return False
            print(f"  信息：读取到目标表头 ({len(actual_target_headers)} 列): {actual_target_headers}")
        except Exception as e:
            messagebox.showerror("目标文件错误", f"读取目标文件表头时出错：\n{e}")
            workbook.close()
            return False
        print(f"  信息：读取目标表头耗时：{time.time() - read_header_start:.2f} 秒")

        print("  信息：根据目标表头调整 DataFrame...")
        adjust_df_start = time.time()
        if df_target.empty: # df_target 现在是经过两轮筛选后的结果
            # 如果源数据为空，创建一个空的DataFrame，列与目标文件一致，用于清空操作
            df_write = pd.DataFrame(columns=actual_target_headers)
            print("  信息：筛选后无数据，准备清空目标区域。")
        else:
            # 确保df_target包含目标文件所需的所有列，没有的填充None
            for header_col in actual_target_headers:
                if header_col not in df_target.columns:
                    df_target[header_col] = None # 或 pd.NA
            # 按照目标表头的顺序重新排列df_target的列
            try:
                # 选择并排序最终写入的DataFrame
                df_write = df_target[actual_target_headers]
            except KeyError as e:
                 missing_key = str(e).strip("'")
                 messagebox.showerror("列匹配错误", f"尝试按目标表头排序时出错: 内部错误，列 '{missing_key}' 未能正确处理。")
                 print(f"错误：最终数据准备缺少列 {missing_key}")
                 workbook.close()
                 return False
        print(f"  信息：调整 DataFrame 耗时：{time.time() - adjust_df_start:.2f} 秒")
        print(f"  信息：最终准备写入的列顺序：{df_write.columns.tolist()}")


        print("  信息：清空目标工作表数据行 (从第2行开始)...")
        clear_start = time.time()
        if sheet.max_row > 1:
            try:
                 sheet.delete_rows(2, sheet.max_row - 1)
                 print(f"  信息：清除工作表数据完成，耗时：{time.time() - clear_start:.2f} 秒")
            except Exception as e:
                 messagebox.showerror("清除错误", f"清除目标工作表数据时出错：\n{e}")
                 if workbook: workbook.close()
                 return False
        else:
            print("  信息：工作表除了表头没有其他数据，无需清除。")

        print(f"  信息：开始写入 {len(df_write)} 行数据到目标工作表...")
        write_start = time.time()
        if not df_write.empty:
            # 逐行写入，处理 None 值
            for row_tuple in df_write.itertuples(index=False, name=None):
                 processed_row = [v if pd.notna(v) else None for v in row_tuple]
                 sheet.append(processed_row)
            print(f"  信息：写入数据完成，耗时：{time.time() - write_start:.2f} 秒")

            print("  信息：设置日期列格式...")
            format_start = time.time()
            date_format = 'yyyy/mm/dd' # Excel 日期格式
            header_list = df_write.columns.tolist()
            start_row_for_format = 2 # 数据从第二行开始
            end_row_for_format = sheet.max_row # 格式化到最后一行

            # 找到日期列的索引
            date_col_indices = {
                header: i for i, header in enumerate(header_list)
                if header in ['开工日期', '备料完成时间', '物料到齐时间']
            }

            if date_col_indices:
                 for row_idx in range(start_row_for_format, end_row_for_format + 1):
                     for header, col_idx_zero in date_col_indices.items():
                         col_letter = get_column_letter(col_idx_zero + 1)
                         cell = sheet[f"{col_letter}{row_idx}"]
                         if cell.value is not None:
                             cell.number_format = date_format
            print(f"  信息：设置日期格式完成，耗时：{time.time() - format_start:.2f} 秒")
        else:
             # <<< MODIFIED: 更新提示信息 >>>
             print(f"  信息：没有数据需要写入 '{sheet.title}' (可能因筛选条件为空)。")

        print("  信息：正在保存目标 XLSM 文件...")
        save_start = time.time()
        workbook.save(target_xlsm_file)
        workbook.close()
        print(f"  信息：保存文件完成，耗时：{time.time() - save_start:.2f} 秒")

        total_time = time.time() - start_time
        print(f"--- '维护数据' 总处理时间: {total_time:.2f} 秒 ---")
        return True # 函数成功结束

    except PermissionError:
         # 文件占用错误处理 (保持不变)
         messagebox.showerror("文件占用", f"无法写入目标文件，请确保该文件未在 Excel 或其他程序中打开：\n{target_xlsm_file}")
         if workbook: workbook.close()
         return False
    except Exception as e:
        # XLSM 更新的通用错误处理 (保持不变)
        traceback_str = traceback.format_exc()
        messagebox.showerror("XLSM 更新错误", f"更新目标 XLSM 文件时发生未预料的错误：\n{e}\n\n详细信息:\n{traceback_str}")
        if workbook: workbook.close()
        return False

# <<< 新增或修改的常量定义 >>>
# (这些常量应放在函数内部或作为全局配置，如果它们是固定的Excel表头)
# WORKSHOP_COL_EXCEL = '车间' # E列，根据您的最新说明，此列不直接用于“组装一”等筛选了
WO_NUM_COL_EXCEL = '工单单号'     # D列
PARENT_PRODUCT_ID_COL_EXCEL = '品号' # C列
WORK_ORDER_TYPE_COL_EXCEL = '单别' # F列，这是新的主要筛选列和分组键之一

# --- GUI 文件选择函数 ---
# <<< 新增辅助函数：从Excel读取T+1的父项生产计划 >>>
def read_t1_excel_plan_data(source_excel_path,
                            workshop_name_for_filter,  # 例如 "组装一"
                            t1_date_obj,
                            header_row_index,  # 0-based
                            report_options_config):  # 全局的 report_options 字典
    print("-" * 20 + " DEBUG INFO in read_t1_excel_plan_data " + "-" * 20) # 调试信息开始标记
    print(f"  传入的 workshop_name_for_filter: '{workshop_name_for_filter}' (类型: {type(workshop_name_for_filter)})")
    print(f"  传入的 report_options_config: {report_options_config}")
    """
    (修改版)
    从指定的源Excel文件读取特定“单别”在T+1日期的生产计划。
    筛选逻辑基于 report_options_config 将 workshop_name_for_filter 映射到“单别”代码列表。
    """
    print(f"  [Excel读取T+1计划 - 新逻辑] 开始读取: {os.path.basename(source_excel_path)}")
    print(
        f"    筛选显示名称(用于查找单别): '{workshop_name_for_filter}', T+1日期: {t1_date_obj.strftime('%Y-%m-%d')}, 表头行(0-based): {header_row_index}")

    df_t1_plan_final = pd.DataFrame()

    try:
        # 1. 确定需要从Excel读取的核心列名
        # 主要需要 '单别', '工单单号', '品号' 和T+1数量列。
        required_excel_header_cols = [
            WORK_ORDER_TYPE_COL_EXCEL,
            WO_NUM_COL_EXCEL,
            PARENT_PRODUCT_ID_COL_EXCEL,
            # 如果Excel的 E列 '车间' 仍有其他用途，也可以加入这里读取，但它不用于主要的分类筛选了
            # WORKSHOP_COL_EXCEL
        ]
        print(f"    1. 预期从Excel主表头读取的列: {required_excel_header_cols}")

        # 2. 读取主表头行，获取上述核心列的实际索引
        df_main_header = pd.read_excel(source_excel_path, sheet_name=0, header=None, skiprows=header_row_index, nrows=1)
        if df_main_header.empty:
            messagebox.showerror("Excel读取错误(T+1计划)", f"无法读取主表头行 (应在Excel第 {header_row_index + 1} 行)。")
            return df_t1_plan_final
        main_header_list = [str(h).strip() if pd.notna(h) else "" for h in df_main_header.iloc[0].tolist()]

        col_indices_map = {}
        for req_col_name in required_excel_header_cols:
            try:
                col_indices_map[req_col_name] = main_header_list.index(req_col_name)
            except ValueError:
                messagebox.showerror("Excel列缺失(T+1计划)",
                                     f"主表头行 (Excel第 {header_row_index + 1} 行) 缺少必需列: '{req_col_name}'")
                return df_t1_plan_final

        # 3. 读取日期表头行，找到T+1日期对应的列索引 (这部分逻辑与您原代码类似，保持不变)
        date_header_row_excel_num = header_row_index + 1 + 1
        df_date_header = pd.read_excel(source_excel_path, sheet_name=0, header=None, skiprows=header_row_index + 1,
                                       nrows=1)
        if df_date_header.empty:
            messagebox.showerror("Excel读取错误(T+1计划)",
                                 f"无法读取日期表头行 (应在Excel第 {date_header_row_excel_num} 行)。")
            return df_t1_plan_final

        t1_date_col_idx = -1
        date_header_values = df_date_header.iloc[0]
        for idx, header_val in enumerate(date_header_values):
            if pd.isna(header_val): continue
            try:
                parsed_dt = None
                if isinstance(header_val, (datetime.datetime, datetime.date)):
                    parsed_dt = header_val.date() if hasattr(header_val, 'date') else header_val
                elif isinstance(header_val, (float, int)):  # Excel numeric date
                    parsed_dt = pd.to_datetime(header_val, unit='D', origin='1899-12-30').date()
                else:  # String date
                    parsed_dt = pd.to_datetime(str(header_val).split(" ")[0], errors='coerce').date()

                if parsed_dt == t1_date_obj:
                    t1_date_col_idx = idx
                    print(
                        f"       找到T+1日期列: 表头值='{header_val}', 解析日期='{parsed_dt}', 列索引(0-based)={t1_date_col_idx}")
                    break
            except Exception:
                continue

        if t1_date_col_idx == -1:
            messagebox.showerror("Excel日期列缺失(T+1计划)",
                                 f"在日期表头行未找到与T+1日期 {t1_date_obj.strftime('%Y-%m-%d')} 匹配的列。\n请检查Excel文件日期格式和表头设置。")
            return df_t1_plan_final

        # 4. 确定最终要读取的所有列的索引 (使用整数索引列表)
        usecols_indices_sorted = sorted(list(set(
            [col_indices_map[col_name] for col_name in required_excel_header_cols] + [t1_date_col_idx]
        )))
        print(f"    2. 最终将读取的Excel列索引(0-based, 排序后): {usecols_indices_sorted}")

        # 5. 读取Excel数据区域 (使用 header=header_row_index 和 usecols=usecols_indices_sorted)
        df_data_raw = pd.read_excel(
            source_excel_path,
            sheet_name=0,
            header=header_row_index,  # 指定表头行，Pandas会用这行的值作为列名
            usecols=usecols_indices_sorted,
            dtype=str  # 先全部按字符串读取，后续再转换
        )
        print(f"    3. 成功从Excel数据区读取 {len(df_data_raw)} 行 (仅含所需列)。")

        # 6. 重命名列为内部标准名称
        #    df_data_raw.columns 现在是Excel表头行中对应 usecols_indices_sorted 的那些列名
        #    我们需要将它们映射到内部一致的名称，如 WorkOrderType_Excel_Plan 等

        #   构建一个从原始Excel表头名到内部标准名的映射
        #   这里需要小心，因为 df_data_raw.columns 的顺序是 usecols_indices_sorted 决定的
        #   而 main_header_list 是完整的原始表头顺序

        #   一个更稳健的方法是，直接用原始Excel列名进行操作，或者在读取后基于原始列名重命名
        #   为了简化，我们假设读取后df_data_raw的列名就是我们在main_header_list中找到的那些
        #   例如，如果 WORK_ORDER_TYPE_COL_EXCEL = '单别', 那么 df_data_raw 就有一列叫 '单别'

        #   进行必要的清理
        for col_excel_name in required_excel_header_cols:
            if col_excel_name in df_data_raw.columns:
                df_data_raw[col_excel_name] = df_data_raw[col_excel_name].astype(str).str.replace(r'\.0$', '',
                                                                                                  regex=True).str.strip()

        # 获取T+1数量列的实际列名 (它可能是一个日期字符串或Excel给的通用名)
        # 我们需要通过其索引 t1_date_col_idx 在 df_data_raw.columns 中找到它
        # df_data_raw.columns 的顺序是基于 usecols_indices_sorted 的
        # 我们需要知道 t1_date_col_idx 在 usecols_indices_sorted 中的位置，才能得到它在 df_data_raw.columns 中的列名
        try:
            actual_t1_qty_col_name_in_df = df_data_raw.columns[usecols_indices_sorted.index(t1_date_col_idx)]
        except (ValueError, IndexError) as e_colfind:
            messagebox.showerror("内部错误(T+1计划)", f"无法确定T+1数量列在读取后DataFrame中的名称: {e_colfind}")
            return pd.DataFrame()

        # 重命名为标准内部列名
        rename_map = {
            WORK_ORDER_TYPE_COL_EXCEL: 'WorkOrderType_Excel_Plan',
            WO_NUM_COL_EXCEL: 'WorkOrderNum_Excel',
            PARENT_PRODUCT_ID_COL_EXCEL: 'ParentProductID_Excel',
            actual_t1_qty_col_name_in_df: 'ParentPlannedQty_T1_Excel'
            # 如果还读取了 E列 '车间', 例如 WORKSHOP_COL_EXCEL (其值为'车间')
            # WORKSHOP_COL_EXCEL: 'WorkshopColumnData_Excel' # 给它一个不冲突的名字
        }
        # 只重命名实际存在的列
        valid_rename_map = {k: v for k, v in rename_map.items() if k in df_data_raw.columns}
        df_data_raw.rename(columns=valid_rename_map, inplace=True)

        # 7. 根据 workshop_name_for_filter (界面选择的车间显示名) 对应的“单别”代码列表进行筛选
        print(f"    4. 根据界面选择 '{workshop_name_for_filter}' 查找对应'单别'代码并筛选...")
        target_danbie_codes_for_filter = []
        found_option_in_config = False
        for _gui_name, (danbie_list, name_part_in_config) in report_options_config.items():
            print(
                f"  循环检查: _gui_name='{_gui_name}', name_part_in_config='{name_part_in_config}' (类型: {type(name_part_in_config)})")  # 打印每次循环的检查对象
            if name_part_in_config == workshop_name_for_filter:
                print(f"    匹配成功! _gui_name='{_gui_name}'")
                print(f"    匹配到的 danbie_list: {danbie_list} (类型: {type(danbie_list)})")
                target_danbie_codes_for_filter = danbie_list
                print(f"    赋值后 target_danbie_codes_for_filter: {target_danbie_codes_for_filter}")
                found_option_in_config = True
                break

        if not found_option_in_config:
            messagebox.showerror("配置错误(T+1计划)",
                                 f"在 report_options 中未找到与 '{workshop_name_for_filter}' 相关的配置。")
            return pd.DataFrame()
        print(f"  循环结束后, found_option_in_config: {found_option_in_config}")
        print(f"  循环结束后, target_danbie_codes_for_filter: {target_danbie_codes_for_filter}")
        if not target_danbie_codes_for_filter:  # 如果找到了配置但代码列表为空
            messagebox.showwarning("配置警告(T+1计划)",
                                   f"车间 '{workshop_name_for_filter}' 的'单别'代码列表为空，无法筛选。")
            return pd.DataFrame()  # 或者返回所有数据，取决于业务需求

        # 确保 WorkOrderType_Excel_Plan 列存在 (这是 '单别' 列重命名后的名字)
        if 'WorkOrderType_Excel_Plan' not in df_data_raw.columns:
            messagebox.showerror("列缺失(T+1计划)",
                                 f"DataFrame中缺少 'WorkOrderType_Excel_Plan' 列 (应来自Excel的 '{WORK_ORDER_TYPE_COL_EXCEL}' 列)。")
            return pd.DataFrame()

        # 执行筛选 (确保数据类型一致，例如都转为字符串进行比较)
        df_data_raw['WorkOrderType_Excel_Plan'] = df_data_raw['WorkOrderType_Excel_Plan'].astype(str).str.strip()

        df_filtered_by_danbie = df_data_raw[
            df_data_raw['WorkOrderType_Excel_Plan'].isin(target_danbie_codes_for_filter)].copy()
        print(f"       筛选条件: Excel的'{WORK_ORDER_TYPE_COL_EXCEL}'列 IN {target_danbie_codes_for_filter}")
        print(f"       按'单别'筛选后剩 {len(df_filtered_by_danbie)} 行。")

        if df_filtered_by_danbie.empty:
            return pd.DataFrame()  # 如果筛选后为空，直接返回空表

        # 8. 筛选T+1计划数量 > 0 并且关键字段非空的行
        df_filtered_by_danbie['ParentPlannedQty_T1_Excel'] = pd.to_numeric(
            df_filtered_by_danbie['ParentPlannedQty_T1_Excel'], errors='coerce').fillna(0)

        key_cols_for_final_plan = ['WorkOrderType_Excel_Plan', 'WorkOrderNum_Excel', 'ParentProductID_Excel',
                                   'ParentPlannedQty_T1_Excel']
        # 检查这些列是否存在
        missing_key_cols = [col for col in key_cols_for_final_plan if col not in df_filtered_by_danbie.columns]
        if missing_key_cols:
            messagebox.showerror("列缺失(T+1计划)", f"筛选后DataFrame缺少核心列: {', '.join(missing_key_cols)}")
            return pd.DataFrame()

        df_filtered_by_danbie.dropna(subset=['WorkOrderNum_Excel', 'ParentProductID_Excel'], inplace=True)  # 工单号和品号不能为空
        df_t1_plan_final = df_filtered_by_danbie[df_filtered_by_danbie['ParentPlannedQty_T1_Excel'] > 0].copy()
        print(f"       筛选有效计划数量 (>0) 后剩 {len(df_t1_plan_final)} 行。")

        if not df_t1_plan_final.empty:
            # 只保留最终需要的列
            df_t1_plan_final = df_t1_plan_final[key_cols_for_final_plan]

        print(f"  [Excel读取T+1计划 - 新逻辑] 成功结束, 返回 {len(df_t1_plan_final)} 条计划。")
        return df_t1_plan_final

    except FileNotFoundError:
        messagebox.showerror("文件错误(T+1计划)", f"找不到源数据文件：\n{source_excel_path}")
    except ValueError as ve:  # 例如列名找不到
        messagebox.showerror("Excel内容错误(T+1计划)",
                             f"处理Excel时遇到值错误 (可能列名不匹配等): {ve}\n{traceback.format_exc()}")
    except ImportError as ie:  # 例如缺少 xlrd
        if 'xlrd' in str(ie).lower():
            messagebox.showerror("依赖缺失(T+1计划)",
                                 f"读取 .xls 需要 'xlrd' 库。\n请运行: pip install xlrd\n错误: {ie}")
        else:
            messagebox.showerror("导入错误(T+1计划)", f"导入错误: {ie}")
    except Exception as e:
        messagebox.showerror("Excel读取异常(T+1计划)",
                             f"读取T+1 Excel计划时发生意外错误：\n{type(e).__name__}: {e}\n{traceback.format_exc()}")

    return pd.DataFrame()  # 发生任何错误则返回空DataFrame

# <<< --- 辅助函数添加结束 --- >>>
def select_file(entry_var):
    filepath = filedialog.askopenfilename(
        title="选择 Excel 文件",
        filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")]
    )
    if filepath:
        entry_var.set(filepath)

# <<< 新增: "总计划" Excel 文件选择函数 >>>
def select_master_plan_file(entry_var):
    filepath = filedialog.askopenfilename(
        title="选择包含'总计划'工作表的 Excel 文件",
        filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")] # 允许各种 Excel 格式
    )
    if filepath:
        entry_var.set(filepath)

def select_xlsm_file(entry_var):
     filepath = filedialog.askopenfilename(
        title="选择 XLSM 目标文件 (用于'维护数据')",
        filetypes=[("Macro-Enabled Excel", "*.xlsm"), ("All files", "*.*")]
    )
     if filepath:
        if not filepath.lower().endswith('.xlsm'):
            messagebox.showwarning("文件类型", "维护数据功能的目标文件应为 .xlsm 格式。")
        entry_var.set(filepath)

# --- “维护数据”按钮的触发函数 ---
def run_single_process():
    """
    当“维护数据到目标 XLSM”按钮被点击时执行。
    收集GUI输入，并调用核心处理函数来更新目标XLSM文件。
    """
    source_file = source_file_var.get()
    target_file = target_file_var.get()
    assembly_filter_value = workshop_var.get() # 现在包含 "深加工车间" 选项

    if not source_file: messagebox.showerror("错误", "请先选择源数据文件！"); return
    if not target_file: messagebox.showerror("错误", "请先选择目标 XLSM 文件！"); return
    if not assembly_filter_value: messagebox.showerror("输入错误", "请选择要处理的车间(维护用)。"); return
    if not os.path.exists(source_file): messagebox.showerror("错误", f"源数据文件不存在：\n{source_file}"); return
    if not os.path.exists(target_file): messagebox.showerror("错误", f"目标 XLSM 文件不存在：\n{target_file}"); return
    if not target_file.lower().endswith('.xlsm'): messagebox.showerror("文件类型错误", f"目标文件必须是 .xlsm 格式：\n{target_file}"); return

    try:
        header_row_excel = header_row_var.get()
        header_row_index = header_row_excel - 1
        if header_row_index < 0: messagebox.showerror("输入错误", "表头行号必须大于 0。"); return
    except tk.TclError: messagebox.showerror("输入错误", "请输入有效的表头行号（数字）。"); return
    except ValueError: messagebox.showerror("输入错误", "请输入有效的表头行号（数字）。"); return

    disable_buttons()
    status_var.set(f"正在处理 [{assembly_filter_value}] 数据到 {os.path.basename(target_file)}...") # 更新状态文本
    root.update_idletasks()

    success = process_assembly_data_preserve_macros(
        source_file, target_file, assembly_filter_value, header_row_index
    )

    enable_buttons()

    if success:
        status_var.set(f"[{assembly_filter_value}] 数据维护完成！目标文件已更新。")
        messagebox.showinfo("完成", f"[{assembly_filter_value}] 数据已成功更新到目标文件 (宏已保留)：\n{target_file}")
    else:
        status_var.set(f"[{assembly_filter_value}] 数据维护失败。请查看错误信息。")

# <<< 新增函数：从 Excel 读取指定日期的工单 >>>
def get_specific_date_work_orders_from_excel(source_excel_path, workshop_name, target_date):
    """
    (基于 get_work_orders_from_plan_excel 修改) 读取源Excel文件，仅加载必要列。
    处理日期表头在主数据表头下一行的情况。
    筛选条件：车间匹配 (深加工特殊处理)，并且在 target_date 对应的日期列中，数值为非零。
    返回满足条件的工单单号列表。

    Args:
        source_excel_path (str): "源数据文件(维护用)" 的路径 (.xls 或 .xlsx)。
        workshop_name (str): 用于筛选的车间名称 ("组装一", "组装二", "深加工车间")。
        target_date (datetime.date): 要筛选的目标日期 (例如, 明天 T+1)。

    Returns:
        list: 唯一的工单单号 (str) 列表。错误则返回 None，无数据则返回空列表。
    """
    # --- 配置 ---
    PLAN_SHEET_NAME = 0
    WORKSHOP_COLUMN_NAME = '车间'
    WORK_ORDER_COLUMN_NAME = '工单单号'
    # --- 配置结束 ---

    read_start_time = time.time()
    print("-" * 30)
    print(f"[指定日期工单筛选] 开始处理: {os.path.basename(source_excel_path)}")

    # --- 获取主数据和日期表头行号 (从GUI) ---
    try:
        header_row_excel_main = header_row_var.get()
        main_header_row_index = header_row_excel_main - 1 # 0-based index
        if main_header_row_index < 0:
            messagebox.showerror("配置错误 (T+1 Excel读取)", "主数据表头行号设置无效 (< 1)。")
            return None
        date_header_row_index = main_header_row_index + 1 # 0-based index
    except tk.TclError:
        messagebox.showerror("配置错误 (T+1 Excel读取)", "无法读取主数据表头行号。")
        return None

    print(f"  主数据表头行(1-based): {main_header_row_index + 1}, 日期表头行(1-based): {date_header_row_index + 1}")
    print(f"  筛选车间标识: '{workshop_name}', 目标日期: {target_date.strftime('%Y-%m-%d')}")

    cols_to_read_indices = [] # 存储需要读取的列的 0-based 索引
    workshop_col_idx = -1
    work_order_col_idx = -1
    target_date_col_idx = -1 # <<< 修改：只找目标日期的索引 >>>

    try:
        # --- 步骤 1: 快速读取主表头行 (同 T+2) ---
        print(f"  1. 快速读取主表头行 (索引 {main_header_row_index})...")
        df_main_header = pd.read_excel(source_excel_path, sheet_name=PLAN_SHEET_NAME,
                                       header=None, skiprows=main_header_row_index, nrows=1)
        if df_main_header.empty: messagebox.showerror("源文件错误 (T+1 Excel读取)", f"无法读取主表头行 (应在第 {main_header_row_index + 1} 行)。"); return None
        main_header_list = df_main_header.iloc[0].tolist()
        try:
            workshop_col_idx = main_header_list.index(WORKSHOP_COLUMN_NAME)
            work_order_col_idx = main_header_list.index(WORK_ORDER_COLUMN_NAME)
            cols_to_read_indices.extend([workshop_col_idx, work_order_col_idx])
            print(f"     '{WORKSHOP_COLUMN_NAME}' 在索引 {workshop_col_idx}, '{WORK_ORDER_COLUMN_NAME}' 在索引 {work_order_col_idx}")
        except ValueError:
            missing = [col for col in [WORKSHOP_COLUMN_NAME, WORK_ORDER_COLUMN_NAME] if col not in main_header_list]
            messagebox.showerror("源文件错误 (T+1 Excel读取)", f"主表头行 (第 {main_header_row_index + 1} 行) 缺少必需列：{missing}"); return None

        # --- 步骤 2: 快速读取日期表头行，获取 T+1 日期列的索引 ---
        print(f"  2. 快速读取日期表头行 (索引 {date_header_row_index})...")
        df_date_header = pd.read_excel(source_excel_path, sheet_name=PLAN_SHEET_NAME,
                                      header=None, skiprows=date_header_row_index, nrows=1)
        if df_date_header.empty: messagebox.showerror("源文件错误 (T+1 Excel读取)", f"未能读取到日期表头行 (应在第 {date_header_row_index + 1} 行)。"); return None

        date_header_row = df_date_header.iloc[0]
        print(f"     开始在日期表头查找目标日期 '{target_date.strftime('%Y-%m-%d')}'...")
        found_target_date_col = False
        for idx, header_val in enumerate(date_header_row):
            if pd.isna(header_val): continue
            try:
                # 尝试解析日期 (与 T+2 逻辑相同)
                if isinstance(header_val, (float, int)):
                    try: parsed_datetime = pd.to_datetime(header_val, unit='D', origin='1899-12-30')
                    except ValueError: parsed_datetime = pd.NaT
                else:
                    parsed_datetime = pd.to_datetime(header_val, errors='coerce')

                if pd.notna(parsed_datetime):
                    parsed_date = parsed_datetime.date()
                    # <<< 修改：精确匹配 target_date >>>
                    if parsed_date == target_date:
                        target_date_col_idx = idx
                        found_target_date_col = True
                        print(f"       找到目标日期列: 值='{header_val}', 解析日期='{parsed_date}', 索引={target_date_col_idx}")
                        break # 找到就停止查找
            except Exception as parse_err:
                 print(f"       警告: 解析日期表头值 '{header_val}' (索引 {idx}) 时出错: {parse_err}")
                 continue

        if not found_target_date_col:
            messagebox.showerror("未找到日期列 (T+1 Excel读取)", f"在日期表头行 (第 {date_header_row_index + 1} 行) 未找到与目标日期 {target_date.strftime('%Y-%m-%d')} 匹配的有效日期列。\n请检查 Excel 文件内容和格式。")
            return None

        # 将目标日期列索引加入待读取列表
        cols_to_read_indices.append(target_date_col_idx)
        cols_to_read_indices = sorted(list(set(cols_to_read_indices)))
        print(f"     最终需要读取的列索引: {cols_to_read_indices}")

        # --- 步骤 3: 使用 usecols 读取优化后的数据 (同 T+2) ---
        print(f"  3. 使用 'usecols' 读取指定列的数据 (header={main_header_row_index})...")
        read_main_data_start = time.time()
        df_data = pd.read_excel(
            source_excel_path, sheet_name=PLAN_SHEET_NAME, header=main_header_row_index,
            usecols=cols_to_read_indices, dtype={WORK_ORDER_COLUMN_NAME: str}
        )
        read_main_data_duration = time.time() - read_main_data_start
        print(f"     读取指定列数据耗时: {read_main_data_duration:.2f} 秒，共 {len(df_data)} 行。")

        # 列存在性检查 (同 T+2)
        if WORKSHOP_COLUMN_NAME not in df_data.columns or WORK_ORDER_COLUMN_NAME not in df_data.columns:
             missing_after_read = [col for col in [WORKSHOP_COLUMN_NAME, WORK_ORDER_COLUMN_NAME] if col not in df_data.columns]
             messagebox.showerror("内部错误 (T+1 Excel读取)", f"使用 usecols 读取后，必需列 {missing_after_read} 未找到。")
             return None

        # --- 步骤 4: 执行筛选 ---
        print("  4. 在加载的数据上执行筛选...")
        # 4a. 筛选车间 (同 T+2)
        df_data[WORKSHOP_COLUMN_NAME] = df_data[WORKSHOP_COLUMN_NAME].astype(str).str.strip()
        workshop_name_str = str(workshop_name)
        if workshop_name_str == "深加工车间":
            exclusion_list = ["组装一", "组装二"]
            df_filtered_workshop = df_data[~df_data[WORKSHOP_COLUMN_NAME].isin(exclusion_list)].copy()
            print(f"     按车间筛选 (排除 {exclusion_list}) 后，剩余 {len(df_filtered_workshop)} 行。")
        elif workshop_name_str in ["组装一", "组装二"]:
            df_filtered_workshop = df_data[df_data[WORKSHOP_COLUMN_NAME] == workshop_name_str].copy()
            print(f"     按车间筛选 (精确匹配 '{workshop_name_str}') 后，剩余 {len(df_filtered_workshop)} 行。")
        else:
            messagebox.showerror("内部错误 (T+1 Excel筛选)", f"接收到未知的车间名称 '{workshop_name_str}'。")
            return None

        if df_filtered_workshop.empty:
             print(f"[指定日期工单筛选] 按车间筛选后无数据 (车间: {workshop_name})。")
             print(f"总耗时: {time.time() - read_start_time:.2f} 秒"); print("-" * 30)
             return []

        # 4b. 找出 DataFrame 中对应 target_date_col_idx 的实际列名 (简化版)
        print(f"     查找 DataFrame 中对应原始目标日期索引 {target_date_col_idx} 的实际列名...")
        # 再次读取完整表头用于映射
        df_full_header = pd.read_excel(source_excel_path, sheet_name=PLAN_SHEET_NAME,
                                       header=main_header_row_index, nrows=0)
        original_header_list = df_full_header.columns.tolist()
        try:
            target_date_column_name_in_df = original_header_list[target_date_col_idx]
            print(f"     原始目标日期索引 {target_date_col_idx} 对应实际列名: '{target_date_column_name_in_df}'")
            # 确认这个列名确实在我们读取的数据中
            if target_date_column_name_in_df not in df_filtered_workshop.columns:
                raise KeyError(f"映射得到的日期列 '{target_date_column_name_in_df}' 不在读取的 DataFrame 列中。")
        except IndexError:
            messagebox.showerror("内部错误 (T+1 Excel读取)", f"无法从原始表头获取索引 {target_date_col_idx} 的列名。")
            return None
        except KeyError as e:
             messagebox.showerror("内部错误 (T+1 Excel读取)", f"映射日期列时出错: {e}")
             return None

        # 4c. 检查目标日期列的非零值
        print(f"     使用实际目标日期列名 '{target_date_column_name_in_df}' 进行筛选...")
        try:
            target_data_column = df_filtered_workshop[target_date_column_name_in_df]
        except KeyError as e:
            messagebox.showerror("内部错误 (T+1 Excel读取)", f"尝试选择目标日期列数据时出错: {e}")
            return None

        # 尝试转为数值，判断非零
        target_data_numeric = pd.to_numeric(target_data_column, errors='coerce')
        non_zero_mask = target_data_numeric.notna() & (target_data_numeric != 0)

        # 4d. 应用日期筛选条件
        df_final_filtered = df_filtered_workshop[non_zero_mask]
        print(f"     在列 '{target_date_column_name_in_df}' 中筛选非零值后，剩余 {len(df_final_filtered)} 行。")

        # --- 处理 df_final_filtered 为空的情况 ---
        if df_final_filtered.empty:
            print(f"[指定日期工单筛选] 在目标日期 {target_date.strftime('%Y-%m-%d')} 列筛选后无数据。")
            print(f"总耗时: {time.time() - read_start_time:.2f} 秒"); print("-" * 30)
            return []

        # --- 步骤 5: 提取工单号 (同 T+2) ---
        print(f"  5. 提取工单号...")
        if WORK_ORDER_COLUMN_NAME not in df_final_filtered.columns: messagebox.showerror("内部错误 (T+1 Excel读取)", f"最终结果缺少工单号列。"); return None
        valid_work_orders = df_final_filtered[WORK_ORDER_COLUMN_NAME].dropna()
        cleaned_work_orders = valid_work_orders.astype(str).str.replace(r'\.0$', '', regex=True).unique().tolist()
        print(f"     提取并清理到 {len(cleaned_work_orders)} 个唯一工单: {cleaned_work_orders[:20]}{'...' if len(cleaned_work_orders)>20 else ''}")

        print(f"[指定日期工单筛选] 成功结束。")
        print(f"总耗时: {time.time() - read_start_time:.2f} 秒"); print("-" * 30)
        return cleaned_work_orders

    # --- 异常处理 (与 T+2 函数保持一致) ---
    except FileNotFoundError: messagebox.showerror("文件错误 (T+1 Excel读取)", f"找不到源数据文件：\n{source_excel_path}"); return None
    except ValueError as ve: messagebox.showerror("源文件内容错误 (T+1 Excel读取)", f"处理源文件时遇到值错误: {ve}"); traceback.print_exc(); return None
    except KeyError as ke: messagebox.showerror("源文件列错误 (T+1 Excel读取)", f"找不到预期的列名: {ke}"); traceback.print_exc(); return None
    except ImportError as ie:
         if 'xlrd' in str(ie).lower(): messagebox.showerror("依赖缺失 (T+1 Excel读取)", f"读取 .xls 需要 'xlrd' 库。\n请运行: pip install xlrd\n错误: {ie}")
         else: messagebox.showerror("导入错误 (T+1 Excel读取)", f"导入错误: {ie}")
         traceback.print_exc(); return None
    except Exception as e: messagebox.showerror("源文件处理错误 (T+1 Excel读取)", f"处理源文件时发生意外错误：\n{e}"); traceback.print_exc(); return None
# <<<--- 函数添加结束 --- >>>

# <<< 新增/修改: "T+2" 从源Excel筛选工单的优化版 (修正缩进, 增加深加工处理逻辑) >>>
def get_work_orders_from_plan_excel(source_excel_path, workshop_name, start_date, end_date):
    """
    (优化版 - 修正缩进 + 深加工逻辑 + 修正 .xls 读取问题) 读取源Excel文件("源数据文件(维护用)")，仅加载必要列。
    处理日期表头在主数据表头下一行的情况。
    筛选条件：车间匹配 (深加工特殊处理)，并且在 start_date 到 end_date (含)
    对应的任一日期列中，数值为非零。返回满足条件的工单单号列表。

    Args:
        source_excel_path (str): "源数据文件(维护用)" 的路径 (.xls 或 .xlsx)。
        workshop_name (str): 用于筛选的车间名称 ("组装一", "组装二", "深加工车间")。
        start_date (datetime.date): 筛选范围的开始日期 (例如, 今天 T+0)。
        end_date (datetime.date): 筛选范围的结束日期 (例如, 今天 T+3)。

    Returns:
        list: 唯一的工单单号 (str) 列表。错误则返回 None，无数据则返回空列表。
    """
    # --- 配置 ---
    PLAN_SHEET_NAME = 0
    WORKSHOP_COLUMN_NAME = '车间'
    WORK_ORDER_COLUMN_NAME = '工单单号'
    # --- 配置结束 ---

    read_start_time = time.time()
    print("-" * 30)
    print(f"[优化版 T+0~T+3 筛选] 开始处理: {os.path.basename(source_excel_path)}")

    # --- 获取主数据和日期表头行号 (从GUI) ---
    try:
        header_row_excel_main = header_row_var.get()
        main_header_row_index = header_row_excel_main - 1 # 0-based index
        if main_header_row_index < 0:
            messagebox.showerror("配置错误 (T+2)", "主数据表头行号设置无效 (< 1)。")
            return None
        date_header_row_index = main_header_row_index + 1 # 0-based index
    except tk.TclError:
        messagebox.showerror("配置错误 (T+2)", "无法读取主数据表头行号。")
        return None

    print(f"  主数据表头行(1-based): {main_header_row_index + 1}, 日期表头行(1-based): {date_header_row_index + 1}")
    print(f"  筛选车间标识: '{workshop_name}', 日期范围: {start_date.strftime('%Y-%m-%d')} 到 {end_date.strftime('%Y-%m-%d')}")

    cols_to_read_indices = [] # 存储需要读取的列的 0-based 索引
    workshop_col_idx = -1
    work_order_col_idx = -1
    date_col_indices_in_range = [] # 存储范围内日期列的索引

    try:
        # --- 步骤 1: 快速读取主表头行，获取 '车间' 和 '工单单号' 的列索引 ---
        print(f"  1. 快速读取主表头行 (索引 {main_header_row_index})...")
        # <<< 修改: 移除 engine='openpyxl' >>>
        df_main_header = pd.read_excel(source_excel_path, sheet_name=PLAN_SHEET_NAME,
                                       header=None, skiprows=main_header_row_index, nrows=1)
                                       # engine='openpyxl') # Removed engine specifier
        if df_main_header.empty:
            messagebox.showerror("源文件错误 (T+2)", f"无法读取主表头行 (应在第 {main_header_row_index + 1} 行)。")
            return None
        main_header_list = df_main_header.iloc[0].tolist()
        try:
            workshop_col_idx = main_header_list.index(WORKSHOP_COLUMN_NAME)
            work_order_col_idx = main_header_list.index(WORK_ORDER_COLUMN_NAME)
            cols_to_read_indices.extend([workshop_col_idx, work_order_col_idx])
            print(f"     '{WORKSHOP_COLUMN_NAME}' 在索引 {workshop_col_idx}, '{WORK_ORDER_COLUMN_NAME}' 在索引 {work_order_col_idx}")
        except ValueError:
            missing = [col for col in [WORKSHOP_COLUMN_NAME, WORK_ORDER_COLUMN_NAME] if col not in main_header_list]
            messagebox.showerror("源文件错误 (T+2)", f"主表头行 (第 {main_header_row_index + 1} 行) 缺少必需列：{missing}")
            return None

        # --- 步骤 2: 快速读取日期表头行，获取 T+0~T+3 日期列的索引 ---
        print(f"  2. 快速读取日期表头行 (索引 {date_header_row_index})...")
        # <<< 修改: 移除 engine='openpyxl' >>>
        df_date_header = pd.read_excel(source_excel_path, sheet_name=PLAN_SHEET_NAME,
                                      header=None, skiprows=date_header_row_index, nrows=1)
                                      # engine='openpyxl') # Removed engine specifier
        if df_date_header.empty:
             messagebox.showerror("源文件错误 (T+2)", f"未能读取到日期表头行 (应在第 {date_header_row_index + 1} 行)。")
             return None

        date_header_row = df_date_header.iloc[0]
        print(f"     开始在日期表头查找范围内的日期列...")
        for idx, header_val in enumerate(date_header_row):
            if pd.isna(header_val): continue
            try:
                # 尝试将各种格式的日期/时间戳转换为日期对象
                # <<< 注意：xlrd 读取日期可能直接是 float 或 datetime，需要健壮处理 >>>
                if isinstance(header_val, (float, int)): # xlrd might read dates as floats
                    # Attempt conversion from Excel serial date number
                    # This requires knowing the date system (1900 or 1904)
                    # For simplicity, let pd.to_datetime try its best first
                    try:
                        parsed_datetime = pd.to_datetime(header_val, unit='D', origin='1899-12-30') # Common origin for Windows Excel
                    except ValueError:
                        parsed_datetime = pd.NaT # Failed conversion
                else:
                    parsed_datetime = pd.to_datetime(header_val, errors='coerce')

                if pd.notna(parsed_datetime):
                    parsed_date = parsed_datetime.date()
                    if start_date <= parsed_date <= end_date:
                        date_col_indices_in_range.append(idx)
                        print(f"       找到符合范围的日期列: 值='{header_val}', 解析日期='{parsed_date}', 索引={idx}")
            except Exception as parse_err:
                 # 打印解析错误，但继续查找
                 print(f"       警告: 解析日期表头值 '{header_val}' (索引 {idx}) 时出错: {parse_err}")
                 continue # 跳过无法解析的列

        if not date_col_indices_in_range:
            messagebox.showerror("未找到日期列 (T+2)", f"在日期表头行 (第 {date_header_row_index + 1} 行) 未找到任何在 {start_date.strftime('%Y-%m-%d')} 到 {end_date.strftime('%Y-%m-%d')} 范围内的有效日期列。\n请检查 Excel 文件内容和格式。")
            return None

        print(f"     找到 {len(date_col_indices_in_range)} 个在范围内的日期列，索引: {date_col_indices_in_range}")
        cols_to_read_indices.extend(date_col_indices_in_range)
        cols_to_read_indices = sorted(list(set(cols_to_read_indices)))
        print(f"     最终需要读取的列索引: {cols_to_read_indices}")

        # --- 步骤 3: 使用 usecols 读取优化后的数据 ---
        print(f"  3. 使用 'usecols' 读取指定列的数据 (header={main_header_row_index})...")
        read_main_data_start = time.time()
        # <<< 修改: 移除 engine='openpyxl' >>>
        df_data = pd.read_excel(
            source_excel_path,
            sheet_name=PLAN_SHEET_NAME,
            header=main_header_row_index,
            usecols=cols_to_read_indices,
            # engine='openpyxl', # Removed engine specifier
            dtype={WORK_ORDER_COLUMN_NAME: str} # 确保工单号读为字符串
        )
        read_main_data_duration = time.time() - read_main_data_start
        print(f"     读取指定列数据耗时: {read_main_data_duration:.2f} 秒，共 {len(df_data)} 行。")

        # 检查读取后列是否存在（usecols 应该已确保，但再次检查）
        if WORKSHOP_COLUMN_NAME not in df_data.columns or WORK_ORDER_COLUMN_NAME not in df_data.columns:
             missing_after_read = [col for col in [WORKSHOP_COLUMN_NAME, WORK_ORDER_COLUMN_NAME] if col not in df_data.columns]
             messagebox.showerror("内部错误 (T+2)", f"使用 usecols 读取后，必需列 {missing_after_read} 未找到。\n实际列: {df_data.columns.tolist()}")
             return None

        # --- 步骤 4: 执行筛选 (在已加载的少量数据上进行) ---
        print("  4. 在加载的数据上执行筛选...")

        # 4a. 筛选车间 (<<< 修改点: 处理深加工逻辑 >>>)
        df_data[WORKSHOP_COLUMN_NAME] = df_data[WORKSHOP_COLUMN_NAME].astype(str).str.strip() # 确保是字符串并去除首尾空格
        workshop_name_str = str(workshop_name) # 获取传入的车间名标识

        if workshop_name_str == "深加工车间":
            print(f"     [T+2 深加工 筛选逻辑] 筛选条件: '{WORKSHOP_COLUMN_NAME}' 列不等于 '组装一' 且不等于 '组装二'")
            exclusion_list = ["组装一", "组装二"]
            df_filtered_workshop = df_data[~df_data[WORKSHOP_COLUMN_NAME].isin(exclusion_list)].copy()
            print(f"     按车间筛选 (排除 {exclusion_list}) 后，剩余 {len(df_filtered_workshop)} 行。")
        elif workshop_name_str in ["组装一", "组装二"]:
            print(f"     [T+2 {workshop_name_str} 筛选逻辑] 筛选条件: '{WORKSHOP_COLUMN_NAME}' 列精确匹配 '{workshop_name_str}'")
            df_filtered_workshop = df_data[df_data[WORKSHOP_COLUMN_NAME] == workshop_name_str].copy()
            print(f"     按车间筛选 (精确匹配 '{workshop_name_str}') 后，剩余 {len(df_filtered_workshop)} 行。")
        else:
            # 处理未预期的 workshop_name
            messagebox.showerror("内部错误 (T+2 筛选)", f"接收到未知的车间名称 '{workshop_name_str}' 进行源文件筛选。")
            print(f"     错误: 未知的 workshop_name '{workshop_name_str}' 用于 T+2 筛选。")
            return None # 返回错误，避免继续处理

        if df_filtered_workshop.empty:
             print(f"[优化版 T+0~T+3 筛选] 按车间筛选后无数据 (车间: {workshop_name})。")
             print(f"总耗时: {time.time() - read_start_time:.2f} 秒")
             print("-" * 30)
             return [] # 返回空列表

        # 4b. 找出 DataFrame 中对应 date_col_indices_in_range 的实际列名 (修正映射逻辑)
        print(f"     查找 DataFrame 中对应原始日期索引 {date_col_indices_in_range} 的实际列名...")
        actual_columns_in_df = df_filtered_workshop.columns.tolist()
        # 重新读取一次完整的表头（仅表头行）来做映射可能更可靠
        # <<< 修改: 移除 engine='openpyxl' >>>
        df_full_header = pd.read_excel(source_excel_path, sheet_name=PLAN_SHEET_NAME,
                                       header=main_header_row_index, nrows=0) # 只读表头
                                       # engine='openpyxl') # Removed engine specifier
        original_header_list = df_full_header.columns.tolist()
        print(f"     原始完整表头列表 (用于映射): {original_header_list}")

        # 构建索引到原始列名的映射 (基于完整表头)
        original_index_to_name_map = {i: name for i, name in enumerate(original_header_list)}
        print(f"     原始索引到完整列名的映射: {original_index_to_name_map}")

        # 根据读取的列索引 (cols_to_read_indices)，找出它们在 df_filtered_workshop 中的实际名称
        actual_name_map_for_read_cols = {}
        for idx in cols_to_read_indices:
            original_name = original_index_to_name_map.get(idx)
            if original_name and original_name in actual_columns_in_df:
                actual_name_map_for_read_cols[idx] = original_name
            elif original_name:
                 print(f"     警告: 原始索引 {idx} 对应的列名 '{original_name}' 不在读取后的 DataFrame 列中 ({actual_columns_in_df})。这不应该发生。")
                 # 可能需要错误处理
            else:
                 print(f"     警告: 无法从原始完整表头获取索引 {idx} 的列名。")
                 # 可能需要错误处理

        print(f"     读取的原始索引到实际列名的映射: {actual_name_map_for_read_cols}")

        # 从这个映射中找出范围内日期列的实际名称
        date_column_names_in_df = []
        for original_idx in date_col_indices_in_range:
            actual_name = actual_name_map_for_read_cols.get(original_idx)
            if actual_name:
                date_column_names_in_df.append(actual_name)
                print(f"     原始日期索引 {original_idx} 对应实际列名: '{actual_name}'")
            else:
                # 如果映射失败，这是一个错误
                messagebox.showerror("内部错误 (T+2)", f"无法映射原始日期索引 {original_idx} 到实际读取的列名。映射: {actual_name_map_for_read_cols}")
                print(f"     错误: 无法在映射 {actual_name_map_for_read_cols} 中找到原始日期索引 {original_idx} 对应的实际列名。")
                return None

        if not date_column_names_in_df:
            messagebox.showerror("错误 (T+2)", "无法在读取的数据中定位任何有效的日期列进行筛选 (映射后)。")
            print(f"     未能从映射 {actual_name_map_for_read_cols} 中找到与日期索引 {date_col_indices_in_range} 对应的列名。")
            return None

        # 4c. 选择这些日期列的数据，并检查非零值
        print(f"     使用实际日期列名进行筛选: {date_column_names_in_df}")
        try:
            # 确保这些列真的存在于 df_filtered_workshop 中
            missing_date_cols = [col for col in date_column_names_in_df if col not in df_filtered_workshop.columns]
            if missing_date_cols:
                raise KeyError(f"映射得到的日期列 {missing_date_cols} 在 DataFrame 中缺失。")
            target_data_columns = df_filtered_workshop[date_column_names_in_df]
        except KeyError as e:
            messagebox.showerror("内部错误 (T+2)", f"尝试使用映射得到的列名选择数据时出错。\n错误: {e}\n列名列表: {date_column_names_in_df}\n可用列: {df_filtered_workshop.columns.tolist()}")
            print(f"     错误: KeyError - {e}")
            return None

        # 尝试将这些列转为数值，无法转换的视为 NaN
        target_data_numeric = target_data_columns.apply(pd.to_numeric, errors='coerce')
        # 判断是否有任何一个日期列的值是有效的数字且不为 0
        # .any(axis=1) 会检查每一行是否至少有一个 True
        non_zero_mask = (target_data_numeric.notna() & (target_data_numeric != 0)).any(axis=1)

        # 4d. 应用日期筛选条件
        df_final_filtered = df_filtered_workshop[non_zero_mask]
        print(f"     在列 {date_column_names_in_df} 中筛选任意非零值后，剩余 {len(df_final_filtered)} 行。")

        # --- 移到 try 块内部，但在筛选之后 ---
        if df_final_filtered.empty:
            print(f"[优化版 T+0~T+3 筛选] 日期值筛选后无数据 (车间: {workshop_name}, 日期范围: {start_date.strftime('%Y-%m-%d')}~{end_date.strftime('%Y-%m-%d')})。")
            print(f"总耗时: {time.time() - read_start_time:.2f} 秒")
            print("-" * 30)
            return [] # 如果筛选后为空，直接返回空列表

        # --- 步骤 5: 提取工单号 (仅当 df_final_filtered 不为空时执行) ---
        print(f"  5. 提取工单号...")
        # 确保工单号列存在
        if WORK_ORDER_COLUMN_NAME not in df_final_filtered.columns:
            messagebox.showerror("内部错误 (T+2)", f"最终筛选结果中缺少工单号列 '{WORK_ORDER_COLUMN_NAME}'。")
            print(f"     错误: 最终结果列: {df_final_filtered.columns.tolist()}")
            return None

        # 提取非空的工单号
        valid_work_orders = df_final_filtered[WORK_ORDER_COLUMN_NAME].dropna()
        # 清理格式 (去掉可能存在的 .0) 并获取唯一值
        cleaned_work_orders = valid_work_orders.astype(str).str.replace(r'\.0$', '', regex=True).unique().tolist()
        print(f"     提取并清理到 {len(cleaned_work_orders)} 个唯一工单: {cleaned_work_orders[:20]}{'...' if len(cleaned_work_orders)>20 else ''}")

        print(f"[优化版 T+0~T+3 筛选] 成功结束。")
        print(f"总耗时: {time.time() - read_start_time:.2f} 秒")
        print("-" * 30)
        return cleaned_work_orders
        # --- 修正缩进结束 ---

    except FileNotFoundError:
        messagebox.showerror("文件错误 (T+2)", f"找不到指定的源数据文件：\n{source_excel_path}")
        return None
    except ValueError as ve:
        # 可能发生在列名查找 (index)、日期解析或 usecols 内部
        messagebox.showerror("源文件内容错误 (T+2)", f"处理源数据文件时遇到值错误，可能列名不匹配或日期格式问题: {ve}")
        traceback.print_exc()
        return None
    except KeyError as ke:
         # 可能发生在 DataFrame 列选择时
         messagebox.showerror("源文件列错误 (T+2)", f"处理源数据时找不到预期的列名: {ke}")
         traceback.print_exc()
         return None
    except ImportError as ie:
         # Handle case where required engine (e.g., xlrd) is missing
         if 'xlrd' in str(ie).lower():
              messagebox.showerror("依赖缺失 (T+2)", f"读取 .xls 文件需要 'xlrd' 库，但未找到。\n请运行: pip install xlrd\n错误: {ie}")
         else:
              messagebox.showerror("导入错误 (T+2)", f"处理源数据文件时发生导入错误: {ie}")
         traceback.print_exc()
         return None
    except Exception as e:
        # Catch other potential errors during reading (.xls specific errors)
        messagebox.showerror("源文件处理错误 (T+2)", f"处理源数据文件以进行 T+2 筛选时发生意外错误：\n{e}")
        traceback.print_exc()
        return None

# <<< 新增函数：生成 T+1 缺料报告 (格式同 T+7) >>>
def generate_t1_shortage_report(workshop_codes, report_name_part, specific_work_orders):
    """
    生成明天 (T+1) 的缺料报告，格式与 T+7/T+15 报告一致。
    数据准备逻辑类似于“全部缺料”，最终过滤依赖于 specific_work_orders。

    Args:
        workshop_codes (list): 用于筛选工单单别 (TB001) 的列表。
        report_name_part (str): 用于文件名和状态信息的报告名称部分。
        specific_work_orders (list): 从 T+1 Excel 计划中筛选出的工单单号列表。
    """
    # --- 硬编码 Offset 仅用于标识和文件名 ---
    offset_for_naming = 1
    print(f"--- 开始生成 [{report_name_part}] T+1 缺料报告 (格式同 T+7) ---")
    print(f"  筛选车间单别: {workshop_codes}")
    print(f"  基于Excel T+1计划筛选出的工单数: {len(specific_work_orders)}")

    # --- 输入验证 ---
    # specific_work_orders 必须是列表，且不能为空（因为调用前应该检查过）
    if not specific_work_orders or not isinstance(specific_work_orders, list):
         messagebox.showerror("内部错误 (T+1 报告)", "未提供有效的 specific_work_orders 列表。")
         print("错误: 调用 generate_t1_shortage_report 时 specific_work_orders 无效。")
         # 理论上不应发生，因为 trigger 函数会检查
         return False
    # workshop_codes 也应有效
    if not workshop_codes or not isinstance(workshop_codes, list):
         messagebox.showerror("内部错误 (T+1 报告)", "未提供有效的 workshop_codes 列表。")
         print("错误: 调用 generate_t1_shortage_report 时 workshop_codes 无效。")
         return False

    disable_buttons()
    status_var.set(f"正在为 [{report_name_part}] 生成 T+1 缺料报告 (Python)...")
    root.update_idletasks()

    conn = None
    output_filepath_original = None
    final_filepath = None

    try:
        # --- 1. 定义目标日期 (使用遥远未来，同“全部缺料”) ---
        target_date_yyyymmdd = '99991231'
        print(f"数据准备 SQL 使用目标日期: {target_date_yyyymmdd} (以包含所有供应)")
        today_date = datetime.date.today()

        # --- 2. 连接数据库 (同 T+7) ---
        print(f"连接数据库...")
        conn = pyodbc.connect(PYODBC_CONN_STRING, autocommit=False)
        cursor = conn.cursor()
        print("数据库连接成功。")

        # --- 3. 执行数据准备 SQL (使用“全部缺料”逻辑: TEMPMNFL 不过滤 UDF01) ---
        print(f"执行数据准备 SQL (全部缺料逻辑)...")
        tables_to_clear = ["TEMPMNFL", "TEMPINVMC", "TEMPMOCTA", "TEMPPURTD"]
        for table in tables_to_clear:
            cursor.execute(f"DELETE FROM {table}")

        # TEMPMNFL 插入 (无 UDF01 过滤)
        sql_insert_tempmfl = """
        INSERT INTO TEMPMNFL(TB001,TB002,TA006,TA034,TA035,TB003,MB002,MB003,MB004,TBYJYL,MOCTAUDF01,TC015,MOCTAUDF02,MA002,MD002)
        SELECT TB.TB001, TB.TB002, TA.TA006, TA.TA034, TA.TA035, TB.TB003, MB.MB002, MB.MB003, MB.MB004, (TB.TB004-TB.TB005) AS TBYJYL, TA.UDF01 AS MOCTAUDF01, TC.TC015, TA.UDF02 AS MOCTAUDF02, MA.MA002, MD.MD002
        FROM MOCTB AS TB INNER JOIN MOCTA AS TA ON TA.TA001 = TB.TB001 AND TA.TA002 = TB.TB002 LEFT JOIN COPTD AS TD ON TD.TD001 = TA.TA026 AND TD.TD002 = TA.TA027 AND TD.TD003 = TA.TA028 LEFT JOIN COPTC AS TC ON TC.TC001 = TD.TD001 AND TC.TC002 = TD.TD002 LEFT JOIN INVMB AS MB ON MB.MB001 = TB.TB003 LEFT JOIN PURMA AS MA ON MA.MA001 = MB.MB032 LEFT JOIN CMSMD AS MD ON MD.MD001 = TA.TA021
        WHERE TA.TA011 IN ('1','2','3') AND (TB.TB009 IN ('700', '710')) AND TA.TA013 = 'Y' AND TB.TB004-TB.TB005 > 0 AND TB.TB011 IN ('1','2') ORDER BY TB.TB003, TA.UDF01
        """
        cursor.execute(sql_insert_tempmfl)
        print(f"   {cursor.rowcount} 行插入 TEMPMNFL (全部缺料逻辑)。")

        # 其他 TEMP 表插入 (同 T+7, 使用 99991231)
        sql_insert_tempinvmc = """
        INSERT INTO TEMPINVMC(INVMB001,SUMINVMC007) SELECT MB.MB001,SUM(ISNULL(INV.MC007,0)) AS SUMINVMC007 FROM INVMB AS MB LEFT JOIN INVMC AS INV ON INV.MC001=MB.MB001 LEFT JOIN CMSMC AS CMS ON CMS.MC001=INV.MC002 WHERE CMS.MC005='Y' AND (INV.MC002 IN ('700', '710')) AND MB.MB001 IN (SELECT TB003 FROM TEMPMNFL) GROUP BY MB.MB001
        """
        cursor.execute(sql_insert_tempinvmc)
        sql_insert_tempmocta = f"""
        INSERT INTO TEMPMOCTA(MOCTA006,MOCMOUNT) SELECT TA.TA006, ISNULL(SUM(TA.TA015),0) - ISNULL(SUM(TA.TA017),0) AS MOUNT FROM MOCTA AS TA LEFT JOIN CMSMC AS CMS ON TA.TA020 = CMS.MC001 WHERE CMS.MC005 = 'Y' AND TA.TA013 = 'Y' AND (TA.TA011 NOT IN ('Y', 'y')) AND (TA.TA020 IN ('700', '710')) AND TA.TA006 IN (SELECT TB003 FROM TEMPMNFL) AND TA.TA010 <= ? GROUP BY TA.TA006
        """
        cursor.execute(sql_insert_tempmocta, target_date_yyyymmdd)
        sql_insert_temppurtd = f"""
        INSERT INTO TEMPPURTD(PURTD004,PURMOUNT) SELECT TD.TD004, ISNULL(SUM(TD.TD008),0) - ISNULL(SUM(TD.TD015),0) AS MOUNT FROM PURTD TD INNER JOIN PURTC TC ON TC.TC001 = TD.TD001 AND TC.TC002 = TD.TD002 INNER JOIN CMSMC CMS ON TD.TD007 = CMS.MC001 LEFT JOIN INVMB MB ON TD.TD004 = MB.MB001 WHERE CMS.MC005 = 'Y' AND (TD.TD007 IN ('700', '710')) AND TD.TD018 = 'Y' AND TD.TD016 = 'N' AND TD.TD004 IN (SELECT TB003 FROM TEMPMNFL) AND TD.TD012 <= ? GROUP BY TD.TD004 HAVING (ISNULL(SUM(TD.TD008),0) - ISNULL(SUM(TD.TD015),0)) > 0
        """
        cursor.execute(sql_insert_temppurtd, target_date_yyyymmdd)

        conn.commit()
        print("数据准备 SQL 执行完毕并已提交。")

        # --- 4. 执行主查询 SQL (过滤条件类似 T+2) ---
        print("执行主查询 SQL...")
        sql_main_query_base = """
        SELECT TMF.TB001, TMF.TB002, TMF.TA006, TMF.TA034, TMF.TA035, TMF.TB003, TMF.MB002, TMF.MB003, TMF.MB004, TMF.TBYJYL,
               TMF.MOCTAUDF01, TMF.TC015, TMF.MOCTAUDF02, TMF.MA002, TMF.MD002,
               ISNULL(TI.SUMINVMC007,0) AS SUMINV,
               ISNULL(TMOC.MOCMOUNT,0) AS SUMMOC,
               ISNULL(TPUR.PURMOUNT,0) AS SUMPUR
        FROM TEMPMNFL AS TMF
        LEFT JOIN TEMPINVMC AS TI ON TI.INVMB001 = TMF.TB003
        LEFT JOIN TEMPMOCTA AS TMOC ON TMOC.MOCTA006 = TMF.TB003
        LEFT JOIN TEMPPURTD AS TPUR ON TPUR.PURTD004 = TMF.TB003
        """
        where_clauses = []
        params = []

        # 1. 添加基于特定工单号 (TB002) 的筛选
        # specific_work_orders 在函数入口已验证非空
        cleaned_specific_wo = [str(wo).replace('.0', '') for wo in specific_work_orders]
        wo_placeholders = ', '.join('?' * len(cleaned_specific_wo))
        where_clauses.append(f"TMF.TB002 IN ({wo_placeholders})")
        params.extend(cleaned_specific_wo)
        print(f"  筛选条件1: 特定工单号 (TMF.TB002 IN (...))")

        # 2. 添加基于车间代码 (TB001) 的筛选
        # workshop_codes 在函数入口已验证非空
        placeholders = ', '.join('?' * len(workshop_codes))
        where_clauses.append(f"TMF.TB001 IN ({placeholders})") # 添加 TB001 筛选
        params.extend(workshop_codes) # 添加 workshop_codes 到参数列表
        print(f"  筛选条件2: 车间代码 (TMF.TB001 IN ({', '.join(workshop_codes)}))")

        # 组合最终 SQL
        sql_final_query = sql_main_query_base
        if where_clauses: sql_final_query += " WHERE " + " AND ".join(where_clauses)
        sql_final_query += " ORDER BY TMF.MA002, TMF.TB003, TMF.MOCTAUDF01, TMF.TB001, TMF.TB002" # 同 T+7 排序

        print(f"  执行最终查询 (参数化)...")
        print(f"  参数: {params}")
        df = pd.read_sql(sql_final_query, conn, params=params)
        print(f"查询到 {len(df)} 条记录。")

        # --- 5. 在 Pandas DataFrame 中完成数据处理、筛选、调整 (同 T+7/T+15) ---
        #    (包括列重命名、计算结余/入库、筛选结余<0、调整日期)
        print("在 Pandas 中处理数据...")
        # (此处逻辑与 generate_simulation_report 函数中第 5 步完全相同)
        # --- 定义最终列顺序 (从 generate_simulation_report 复制过来) ---
        # --- 定义新的、统一的列顺序 ---
        final_columns_to_use = [
            '工单单别', '工单编号', '产品品号', '产品品名',
            '预计领用日',
            '主供应商名称',
            '料件品号', '料件品名', '料件规格', '单位', '现有库存',
            '预计用量', '库存结余', '预计入库',
            '回复到料时间', '最晚到料时间', 'PO#',
            '工作中心'
        ]

        # 假设这部分处理完后得到 df_final
        if not df.empty:
            # 5a. 列重命名和类型转换 (同 T+7)
            column_mapping = {
                 'TB001': '工单单别', 'TB002': '工单编号', 'TA006': '产品品号', 'TA034': '产品品名', 'TA035': '产品规格',
                 'TB003': '料件品号', 'MB002': '料件品名', 'MB003': '料件规格', 'MB004': '单位',
                 'SUMINV': '现有库存', 'TBYJYL': '预计用量', 'MOCTAUDF01': '最晚到料时间_raw', 'TC015': 'PO#',
                 'MOCTAUDF02': '预计领用日_raw', 'MA002': '主供应商名称', 'MD002': '工作中心',
                 'SUMMOC': 'SUMMOC', 'SUMPUR': 'SUMPUR'
            }
            df.rename(columns=column_mapping, inplace=True)
            numeric_cols = ['现有库存', '预计用量', 'SUMMOC', 'SUMPUR']
            for col in numeric_cols: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

            # 5b. 计算列 (同 T+7)
            df['库存结余'] = df['现有库存'] - df.groupby('料件品号')['预计用量'].cumsum()
            df['预计入库'] = df['SUMMOC'] + df['SUMPUR']

            # 5c. 格式化日期列 (同 T+7)
            if 'format_yyyymmdd_str' not in globals(): # 确保函数存在
                def format_yyyymmdd_str(date_input):
                   # ... (省略函数体, 与 generate_simulation_report 中的一致) ...
                   if pd.isna(date_input) or date_input == '': return None
                   date_str = str(date_input).strip().replace('.0', '') # Clean up
                   if len(date_str) == 8 and date_str.isdigit():
                       try: return datetime.datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
                       except ValueError: return date_str
                   if len(date_str) == 10 and (date_str[4] == '-' or date_str[4] == '/'):
                       try: return pd.to_datetime(date_str).strftime('%Y-%m-%d')
                       except ValueError: return date_str
                   return date_str
            df['最晚到料时间'] = df['最晚到料时间_raw'].apply(format_yyyymmdd_str)
            df['预计领用日'] = df['预计领用日_raw'].apply(format_yyyymmdd_str)

            # 5d. 筛选数据 (库存结余 < 0) - 同 T+7
            print(" - 在 DataFrame 中筛选库存结余 < 0 的行...")
            df_filtered = df[df['库存结余'] < 0].copy()
            print(f"   筛选后剩余 {len(df_filtered)} 行。")

            # --- 处理 df_filtered 为空的情况 ---
            if df_filtered.empty:
                print("信息：筛选后无缺料数据。将生成空报告。")
                # --- 使用新的统一列顺序定义空表结构 ---
                print(f"  空表结构：应用统一列顺序 ({len(final_columns_to_use)} 列)。")
                df_final = pd.DataFrame(columns=final_columns_to_use)  # 使用新的统一列表
            else:
                # 5e. 调整日期 (最晚到料时间 < 今天 -> 今天) - 同 T+7
                print(" - 在 DataFrame 中调整超期 '最晚到料时间'...")
                today_dt_str = today_date.strftime('%Y-%m-%d')
                today_ts = pd.Timestamp(today_date)
                has_date_str_mask = df_filtered['最晚到料时间'].notna()
                df_filtered['最晚到料时间_dt'] = pd.NaT
                if has_date_str_mask.any():
                     df_filtered.loc[has_date_str_mask, '最晚到料时间_dt'] = pd.to_datetime(
                         df_filtered.loc[has_date_str_mask, '最晚到料时间'], errors='coerce'
                     )
                     final_update_mask = has_date_str_mask & df_filtered['最晚到料时间_dt'].notna() & (df_filtered['最晚到料时间_dt'] < today_ts)
                     if final_update_mask.any():
                         update_count = final_update_mask.sum()
                         df_filtered.loc[final_update_mask, '最晚到料时间'] = today_dt_str
                         print(f"   {update_count} 行的超期 '最晚到料时间' 已调整为今天。")
                if '最晚到料时间_dt' in df_filtered.columns: df_filtered.drop(columns=['最晚到料时间_dt'], inplace=True)

                # 5f. 插入 '回复到料时间' 列 (同 T+7)
                print(" - 在 DataFrame 中插入 '回复到料时间' 列...")
                if '回复到料时间' not in df_filtered.columns:
                    df_filtered['回复到料时间'] = None

                    # 5g. 选择最终列并使用新的统一顺序
                    # --- （移除旧的基于 report_name_part 选择列顺序的 if/else 逻辑） ---
                    print(f" - 应用统一的列顺序 ({len(final_columns_to_use)} 列)...")
                    df_final = pd.DataFrame()
                    for col in final_columns_to_use:  # 使用新的统一列表
                        if col in df_filtered.columns:
                            df_final[col] = df_filtered[col]
                        else:
                            print(f"警告：最终列 '{col}' 缺失");
                            df_final[col] = None
        else:  # 原始查询为空
            print("信息：未查询到符合条件的基础数据。将生成空报告。")
            # --- 使用新的统一列顺序定义空表结构 ---
            print(f"  空表结构：应用统一列顺序 ({len(final_columns_to_use)} 列)。")
            df_final = pd.DataFrame(columns=final_columns_to_use)  # 使用新的统一列表
        # --- 6. 生成 Excel 文件 (同 T+7) ---
        #    (文件名使用 T+1 标识)
        # ... [略去重复的文件路径、目录检查代码] ...
        # --- 修改：硬编码输出路径 ---
        desktop_path = Path(r"E:\Users\Desktop")
        # --- 修改结束 ---
        try:
            if not desktop_path.exists(): desktop_path.mkdir(parents=True, exist_ok=True)
            elif not os.access(desktop_path, os.W_OK): print(f"警告：无写入权限 {desktop_path}")
        except Exception as e: messagebox.showerror("目录错误 (T+1)", f"无法访问目录: {e}"); return False # 简化错误处理

        timestamp_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')
        safe_report_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in report_name_part)
        # <<< 修改文件名中的 T+ 偏移标识 >>>
        output_filename_original = f"TEMP_{safe_report_name}_T+{offset_for_naming}缺料_{timestamp_str}.xlsx"
        output_filepath_original = desktop_path / output_filename_original
        print(f"将处理后的数据写入临时文件: {output_filepath_original}")
        df_final.to_excel(output_filepath_original, index=False, engine='openpyxl')
        print("临时 Excel 文件已生成。")

        # --- 7. 使用 openpyxl 进行格式化 ---
        print("应用 Excel 格式化...")
        if not output_filepath_original.exists():
            raise FileNotFoundError(f"未能找到生成的临时文件: {output_filepath_original}")

        wb_process = openpyxl.load_workbook(output_filepath_original)
        ws_process = wb_process.active
        if ws_process is None:
            raise ValueError("无法加载临时 Excel 文件的工作表。")

        # 定义样式
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # 黄色填充
        header_font = Font(name='微软雅黑', size=9, bold=True)  # 表头字体
        data_font = Font(name='微软雅黑', size=8)  # 数据区域字体
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中对齐
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 左对齐
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)  # 右对齐
        date_format = 'yyyy-mm-dd'  # 日期格式

        print(" - 格式化表头...")
        headers_final = df_final.columns.tolist()  # 获取最终的列名列表
        for col_idx, header in enumerate(headers_final, 1):  # 从1开始计数列
            cell = ws_process.cell(row=1, column=col_idx)  # 获取表头单元格
            cell.font = header_font  # 应用表头字体
            cell.alignment = center_align  # 应用居中对齐
            # 特殊列应用黄色背景 (保持不变)
            if header in ['回复到料时间', '主供应商名称']:
                cell.fill = yellow_fill

        print(" - 设置数据区域字体、格式和列宽...")
        # 定义需要右对齐和设置为日期格式的列名 (保持不变)
        right_align_headers = {'现有库存', '预计用量', '库存结余', '预计入库'}
        date_headers = {'最晚到料时间', '预计领用日', '回复到料时间'}

        # 定义固定列宽映射
        fixed_column_widths = {
            '工单单别': 5,
            '工单编号': 7.5,
            '产品品号': 10,
            '主供应商名称': 10,
            '料件品号': 7.5,
            '料件品名': 25,
            '料件规格': 15,
            '现有库存': 7.5,
            '预计用量': 7.5,
            '库存结余': 7.5,
            '预计入库': 7.5,
            '工作中心': 7
        }

        # 初始化列宽字典（对于未指定固定宽度的列）
        col_widths = {}
        for col_idx, header in enumerate(headers_final, 1):
            if header in fixed_column_widths:
                col_widths[col_idx] = fixed_column_widths[header]
            else:
                # 对于未指定固定宽度的列，计算表头字符宽度
                header_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in header)
                col_widths[col_idx] = max(header_len + 1, 6)  # 加一点边距，最小宽度为6

        # 遍历数据行进行格式化和计算非固定列的宽度
        if ws_process.max_row > 1:  # 检查是否有数据行
            for row_idx in range(2, ws_process.max_row + 1):  # 从第2行开始
                for col_idx, header in enumerate(headers_final, 1):  # 遍历列
                    cell = ws_process.cell(row=row_idx, column=col_idx)
                    cell.font = data_font  # 设置数据字体

                    # 设置默认对齐和格式
                    alignment_to_apply = left_align  # 所有单元格都设置为自动换行
                    number_format_to_apply = '@'  # 默认为文本格式

                    # 根据列名应用特定对齐和格式 (逻辑不变)
                    if header in right_align_headers:
                        alignment_to_apply = right_align
                        number_format_to_apply = '#,##0'  # 数字格式
                    elif header in date_headers:
                        number_format_to_apply = date_format  # 日期格式
                        alignment_to_apply = center_align  # 日期居中

                    cell.alignment = alignment_to_apply
                    cell.number_format = number_format_to_apply

                    # 仅对未固定宽度的列计算单元格内容宽度
                    if header not in fixed_column_widths and cell.value is not None:
                        # 将日期对象转为字符串计算宽度
                        if isinstance(cell.value, (datetime.datetime, datetime.date)):
                            cell_str = cell.value.strftime(date_format)  # 使用目标格式计算
                        else:
                            cell_str = str(cell.value)
                        # 计算字符宽度 (中文算2，英文算1)
                        cell_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_str)
                        col_widths[col_idx] = max(col_widths.get(col_idx, 0), cell_len + 1)  # 更新最大宽度

        # 应用列宽
        max_allowed_width = 60  # 非固定宽度列的最大宽度限制
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)  # 获取列字母
            if headers_final[col_idx - 1] in fixed_column_widths:
                # 使用固定宽度
                ws_process.column_dimensions[col_letter].width = width
            else:
                # 非固定宽度列，使用计算宽度但有最大限制
                adjusted_width = min(width, max_allowed_width)
                ws_process.column_dimensions[col_letter].width = adjusted_width

        print(" - 冻结首行...")
        ws_process.freeze_panes = ws_process['A2']  # 冻结第一行

        # --- 8. 重命名文件 (同 T+7, 但使用 T+1 标识) ---
        today_mmdd = datetime.date.today().strftime('%m%d')
        # <<< 修改文件名中的 T+ 偏移标识 >>>
        final_filename = f"{safe_report_name}_T+{offset_for_naming}缺料_{today_mmdd}.xlsx"
        final_filepath = desktop_path / final_filename
        print(f"准备将文件重命名/保存为: {final_filepath}")

        # 保存格式化后的工作簿
        wb_process.save(final_filepath)
        print(f"文件已保存到: {final_filepath}")
        wb_process.close()

        # 删除临时文件
        try:
            if output_filepath_original.exists(): os.remove(output_filepath_original)
            print(f"已删除临时文件: {output_filepath_original}")
        except Exception as remove_err: print(f"警告：删除临时文件时出错: {remove_err}")

        # --- 9. 结束处理 (同 T+7) ---
        status_var.set(f"[{report_name_part}] T+1 缺料报告已成功生成！")
        messagebox.showinfo("完成", f"T+1 缺料报告已生成并保存到桌面：\n{final_filepath}")
        return True

    # --- 异常处理部分 (同 T+7) ---
    # ... [略去重复的 try...except...finally 代码块，与 generate_simulation_report 的一致] ...
    except pyodbc.Error as db_err:
        sqlstate = db_err.args[0]; message = str(db_err.args[1]); print(f"DB错误: {message}"); traceback.print_exc()
        messagebox.showerror("数据库错误 (T+1)", f"SQLSTATE: {sqlstate}\n错误: {message}")
        status_var.set(f"[{report_name_part}] T+1 报告失败 (数据库错误)。")
        if conn: conn.rollback(); return False
    except PermissionError as pe:
         target_path = final_filepath or output_filepath_original; print(f"文件权限错误: {pe}"); traceback.print_exc()
         messagebox.showerror("文件错误 (T+1)", f"文件权限错误: {pe}\n文件: {target_path}")
         status_var.set(f"[{report_name_part}] T+1 报告失败 (文件权限错误)。")
         if conn: conn.rollback(); return False
    except FileNotFoundError as fnf_err:
         print(f"文件未找到错误 (T+1): {fnf_err}"); traceback.print_exc()
         messagebox.showerror("文件错误 (T+1)", f"文件未找到: {fnf_err}")
         status_var.set(f"[{report_name_part}] T+1 报告失败 (文件未找到)。")
         if conn: conn.rollback(); return False
    except Exception as e:
        print(f"生成 T+1 报告时错误: {e}"); traceback.print_exc()
        messagebox.showerror("执行错误 (T+1)", f"意外错误: {type(e).__name__}\n详情: {e}")
        status_var.set(f"[{report_name_part}] T+1 报告失败 ({type(e).__name__})。")
        if conn: conn.rollback(); return False
    finally:
        if conn:
            try: conn.close(); print("数据库连接已关闭。")
            except Exception as close_err: print(f"关闭DB连接出错: {close_err}")
        enable_buttons()
        root.update_idletasks()
# <<<--- 函数添加结束 --- >>>

# --- 模拟发料报告生成 (修改版：增加 specific_work_orders 参数 + 列顺序调整) ---
def generate_simulation_report(offset, workshop_codes, report_name_part, specific_work_orders=None):
    """
    纯 Python 实现模拟发料报告生成及后处理(修改版)。
    根据传入的车间代码列表和报告名称部分进行筛选和命名。
    如果提供了 specific_work_orders (用于T+2)，则额外按此列表筛选工单编号，
    并且 T+2 模式下也会根据 workshop_codes 筛选工单单别。
    *** T+7 和 T+15 的目标日期计算逻辑已修改，以匹配特定的 VBA 逻辑。***
    *** 列顺序已调整: 主供应商名称 在 料件品号 之前 ***
    """
    base_date_str = base_date_var.get() # 从全局变量获取基准日期字符串

    # --- 输入验证 ---
    if not base_date_str: messagebox.showerror("错误", "请输入模拟发料的基准日期！"); return False # 返回失败
    try: base_date = datetime.datetime.strptime(base_date_str, '%Y-%m-%d').date()
    except ValueError: messagebox.showerror("错误", "基准日期格式不正确，请输入YYYY-MM-DD 格式！"); return False

    # T+2 模式下 workshop_codes 和 specific_work_orders 都需要验证 (specific 可能为空列表)
    # T+7/T+15 模式下 workshop_codes 需要验证
    if specific_work_orders is None: # T+7/T+15
        if not workshop_codes or not isinstance(workshop_codes, list) or not workshop_codes:
             messagebox.showerror("内部错误", f"T+{offset} 模式下未提供有效的车间代码列表给报告生成函数。")
             return False
    else: # T+2
        if not workshop_codes or not isinstance(workshop_codes, list) or not workshop_codes:
            # T+2 也需要 workshop_codes 用于单别筛选
            messagebox.showerror("内部错误", f"T+2 模式下未提供有效的车间代码列表 (用于单别筛选)。")
            return False
        if not isinstance(specific_work_orders, list): # 允许空列表
             messagebox.showerror("内部错误", f"T+2 模式下 specific_work_orders 不是有效列表。")
             return False

    if not report_name_part:
         messagebox.showerror("内部错误", "未提供有效的报告名称部分给报告生成函数。")
         return False

    disable_buttons() # 调用全局函数禁用按钮
    # 更新状态栏信息
    status_message = f"正在为 [{report_name_part}] 生成 T+{offset} 缺料报告 (Python)..."
    if specific_work_orders is not None: # 明确检查是否提供了列表 (即使是空列表)
         status_message = f"正在为 [{report_name_part}] 生成 T+{offset} 缺料报告 (基于源文件筛选 + 车间代码)..." # 更新提示
    status_var.set(status_message) # 更新全局状态变量
    root.update_idletasks() # 更新 GUI

    conn = None
    output_filepath_original = None
    final_filepath = None

    try:
        # --- 1. 计算目标日期 (用于数据库日期过滤) ---
        # <<< --- 修改开始: T+7 和 T+15 日期计算逻辑 --- >>>
        # 使用 Python 的 isoweekday() 来匹配 VBA 的 Weekday(..., vbMonday) (1=周一, 7=周日)
        weekday_monday_start = base_date.isoweekday()

        if offset == 7:
            # 匹配 VBA 逻辑: inputDate + (7 - Weekday(inputDate, vbMonday)) + 7
            offset_days_vba_t7 = (7 - weekday_monday_start) + 7
            target_date_obj = base_date + datetime.timedelta(days=offset_days_vba_t7)
            print(f"基准日期: {base_date_str}, 偏移: {offset}, 计算数据库过滤目标日期 (T+7, VBA逻辑): {target_date_obj.strftime('%Y%m%d')}")
        elif offset == 15:
            # 匹配 VBA 逻辑: inputDate + (7 - Weekday(inputDate, vbMonday)) + 14
            offset_days_vba_t15 = (7 - weekday_monday_start) + 14
            target_date_obj = base_date + datetime.timedelta(days=offset_days_vba_t15)
            print(f"基准日期: {base_date_str}, 偏移: {offset}, 计算数据库过滤目标日期 (T+15, VBA逻辑): {target_date_obj.strftime('%Y%m%d')}")
        # <<< --- 修改结束 --- >>>
        elif offset == 2: # T+2 逻辑保持不变
            target_date_obj = base_date + datetime.timedelta(days=3) # T+2报告使用T+3日期
            print(f"基准日期: {base_date_str}, 偏移: {offset}, 计算数据库过滤目标日期 (T+2报告使用T+3): {target_date_obj.strftime('%Y%m%d')}")
        else:
            messagebox.showerror("配置错误", f"模拟发料报告不支持 T+{offset} 的偏移设置。")
            print(f"错误: 不支持的偏移天数: {offset}")
            enable_buttons()
            status_var.set(f"报告生成失败 (无效偏移 T+{offset})。")
            root.update_idletasks()
            return False # 返回失败

        target_date_yyyymmdd = target_date_obj.strftime('%Y%m%d')
        today_date = datetime.date.today() # 获取当前日期

        # --- 2. 连接数据库 ---
        print(f"连接数据库...")
        conn = pyodbc.connect(PYODBC_CONN_STRING, autocommit=False) # 关闭自动提交
        cursor = conn.cursor()
        print("数据库连接成功。")

        # --- 3. 执行数据准备 SQL (放在事务中) ---
        print(f"执行数据准备 SQL (基于日期 {target_date_yyyymmdd})...")
        tables_to_clear = ["TEMPMNFL", "TEMPINVMC", "TEMPMOCTA", "TEMPPURTD"]
        for table in tables_to_clear:
            print(f" - 清空 {table}...")
            cursor.execute(f"DELETE FROM {table}")

        print(" - 插入 TEMPMNFL...")
        sql_insert_tempmfl = f"""
        INSERT INTO TEMPMNFL(TB001,TB002,TA006,TA034,TA035,TB003,MB002,MB003,MB004,TBYJYL,MOCTAUDF01,TC015,MOCTAUDF02,MA002,MD002)
        SELECT TB.TB001, TB.TB002, TA.TA006, TA.TA034, TA.TA035, TB.TB003, MB.MB002, MB.MB003, MB.MB004, (TB.TB004-TB.TB005) AS TBYJYL, TA.UDF01 AS MOCTAUDF01, TC.TC015, TA.UDF02 AS MOCTAUDF02, MA.MA002, MD.MD002
        FROM MOCTB AS TB
        INNER JOIN MOCTA AS TA ON TA.TA001 = TB.TB001 AND TA.TA002 = TB.TB002
        LEFT JOIN COPTD AS TD ON TD.TD001 = TA.TA026 AND TD.TD002 = TA.TA027 AND TD.TD003 = TA.TA028
        LEFT JOIN COPTC AS TC ON TC.TC001 = TD.TD001 AND TC.TC002 = TD.TD002
        LEFT JOIN INVMB AS MB ON MB.MB001 = TB.TB003
        LEFT JOIN PURMA AS MA ON MA.MA001 = MB.MB032
        LEFT JOIN CMSMD AS MD ON MD.MD001 = TA.TA021
        WHERE TA.TA011 IN ('1','2','3') /*工单状态*/
          AND (TB.TB009 IN ('700', '710')) /*用料库别*/
          AND TA.TA013 = 'Y' /*审核*/
          AND TB.TB004-TB.TB005 > 0 /*预计用量>0*/
          AND TA.UDF01 IS NOT NULL AND TA.UDF01 <> '' /*最晚到料时间非空*/
          AND TB.TB011 IN ('1','2') /*用料确认码*/
          AND TA.UDF01 <= ? /*最晚到料时间在目标日期前*/
        ORDER BY TB.TB003, TA.UDF01
        """
        cursor.execute(sql_insert_tempmfl, target_date_yyyymmdd)
        print(f"   {cursor.rowcount} 行插入 TEMPMNFL。")


        print(" - 插入 TEMPINVMC...")
        sql_insert_tempinvmc = """
        INSERT INTO TEMPINVMC(INVMB001,SUMINVMC007)
        SELECT MB.MB001,SUM(ISNULL(INV.MC007,0)) AS SUMINVMC007
        FROM INVMB AS MB
        LEFT JOIN INVMC AS INV ON INV.MC001=MB.MB001
        LEFT JOIN CMSMC AS CMS ON CMS.MC001=INV.MC002
        WHERE CMS.MC005='Y' /*仓库纳入库存*/
          AND (INV.MC002 IN ('700', '710')) /*库存库别*/
          AND MB.MB001 IN (SELECT TB003 FROM TEMPMNFL) /*只统计TEMPMNFL中需要的料件*/
        GROUP BY MB.MB001
        """
        cursor.execute(sql_insert_tempinvmc)
        print(f"   {cursor.rowcount} 行插入 TEMPINVMC。")

        print(" - 插入 TEMPMOCTA...")
        sql_insert_tempmocta = f"""
        INSERT INTO TEMPMOCTA(MOCTA006,MOCMOUNT)
        SELECT TA.TA006, ISNULL(SUM(TA.TA015),0) - ISNULL(SUM(TA.TA017),0) AS MOUNT /*预计产量-已入库*/
        FROM MOCTA AS TA
        LEFT JOIN CMSMC AS CMS ON TA.TA020 = CMS.MC001
        WHERE CMS.MC005 = 'Y' /*仓库纳入库存*/
          AND TA.TA013 = 'Y' /*审核*/
          AND (TA.TA011 NOT IN ('Y', 'y')) /*未完工结案*/
          AND (TA.TA020 IN ('700', '710')) /*工单库别*/
          AND TA.TA006 IN (SELECT TB003 FROM TEMPMNFL) /*只统计TEMPMNFL中缺料的料件（作为产品）的未入库工单*/
          AND TA.TA010 <= ? /*工单预计完工日期在目标日期前*/
        GROUP BY TA.TA006
        """
        cursor.execute(sql_insert_tempmocta, target_date_yyyymmdd)
        print(f"   {cursor.rowcount} 行插入 TEMPMOCTA。")

        print(" - 插入 TEMPPURTD...")
        # <<< --- 修改开始: 使用 HAVING 确保正确过滤未交量 > 0 的采购数据 --- >>>
        sql_insert_temppurtd = f"""
        INSERT INTO TEMPPURTD(PURTD004,PURMOUNT)
        SELECT TD.TD004, ISNULL(SUM(TD.TD008),0) - ISNULL(SUM(TD.TD015),0) AS MOUNT /*采购量-已交量*/
        FROM PURTD TD
        INNER JOIN PURTC TC ON TC.TC001 = TD.TD001 AND TC.TC002 = TD.TD002
        INNER JOIN CMSMC CMS ON TD.TD007 = CMS.MC001
        LEFT JOIN INVMB MB ON TD.TD004 = MB.MB001
        WHERE CMS.MC005 = 'Y' /*仓库纳入库存*/
          AND (TD.TD007 IN ('700', '710')) /*采购库别*/
          AND TD.TD018 = 'Y' /*审核*/
          AND TD.TD016 = 'N' /*未结案*/
          /* --- 聚合条件从 WHERE 移除 --- */
          /* AND (ISNULL(TD.TD008,0) - ISNULL(SUM(TD.TD015),0)) > 0 */
          AND TD.TD004 IN (SELECT TB003 FROM TEMPMNFL) /*只统计TEMPMNFL中缺料的料件的未交采购单*/
        GROUP BY TD.TD004
        /* --- 使用 HAVING 对聚合结果进行筛选 --- */
        HAVING (ISNULL(SUM(TD.TD008),0) - ISNULL(SUM(TD.TD015),0)) > 0 /*筛选未交量>0的料号*/
        """
        # <<< --- 修改结束 --- >>>
        cursor.execute(sql_insert_temppurtd)
        print(f"   {cursor.rowcount} 行插入 TEMPPURTD。")

        conn.commit()
        print("数据准备 SQL 执行完毕并已提交。")


        # --- 4. 执行主查询 SQL ---
        print("执行主查询 SQL...")
        sql_main_query_base = """
        SELECT TMF.TB001, TMF.TB002, TMF.TA006, TMF.TA034, TMF.TA035, TMF.TB003, TMF.MB002, TMF.MB003, TMF.MB004, TMF.TBYJYL,
               TMF.MOCTAUDF01, TMF.TC015, TMF.MOCTAUDF02, TMF.MA002, TMF.MD002,
               ISNULL(TI.SUMINVMC007,0) AS SUMINV,
               ISNULL(TMOC.MOCMOUNT,0) AS SUMMOC,
               ISNULL(TPUR.PURMOUNT,0) AS SUMPUR
        FROM TEMPMNFL AS TMF
        LEFT JOIN TEMPINVMC AS TI ON TI.INVMB001 = TMF.TB003
        LEFT JOIN TEMPMOCTA AS TMOC ON TMOC.MOCTA006 = TMF.TB003
        LEFT JOIN TEMPPURTD AS TPUR ON TPUR.PURTD004 = TMF.TB003
        """

        where_clauses = []
        params = []

        # T+2 模式下，同时根据 specific_work_orders 和 workshop_codes 筛选
        if specific_work_orders is not None:  # T+2 模式
            print("  模式: T+2 (基于源文件筛选结果 + 车间代码 双重筛选)")

            # 1. 添加基于特定工单号 (TB002) 的筛选
            if isinstance(specific_work_orders, list) and len(specific_work_orders) > 0:
                cleaned_specific_wo = [str(wo).replace('.0', '') for wo in specific_work_orders]
                wo_placeholders = ', '.join('?' * len(cleaned_specific_wo))
                where_clauses.append(f"TMF.TB002 IN ({wo_placeholders})")
                params.extend(cleaned_specific_wo)
                print(f"  筛选条件1: 特定工单号 (TMF.TB002 IN ({', '.join(cleaned_specific_wo[:5])}{'...' if len(cleaned_specific_wo) > 5 else ''}))")
            else:
                # 如果 specific_work_orders 为空，则查询无结果
                print("  信息: specific_work_orders 为空列表 (源文件筛选无结果)，主查询将不返回任何数据。")
                where_clauses.append("1 = 0") # 添加永假条件，确保无结果

            # 2. 添加基于车间代码 (TB001) 的筛选 (仅当上面未确定无结果时)
            if not where_clauses or where_clauses[-1] != "1 = 0": # 检查是否已经确定无结果
                if workshop_codes and isinstance(workshop_codes, list): # workshop_codes 在函数入口已验证非空
                    placeholders = ', '.join('?' * len(workshop_codes))
                    where_clauses.append(f"TMF.TB001 IN ({placeholders})") # 添加 TB001 筛选
                    params.extend(workshop_codes) # 添加 workshop_codes 到参数列表
                    print(f"  筛选条件2: 车间代码 (TMF.TB001 IN ({', '.join(workshop_codes)}))")
                else:
                    # 此处理论上不应发生，因为入口已检查 workshop_codes
                    print("  错误: T+2 模式下内部错误，workshop_codes 无效。")
                    # 为安全起见，添加永假条件
                    if not where_clauses or where_clauses[-1] != "1 = 0":
                         where_clauses.append("1 = 0")


        else:  # T+7 / T+15 模式: 只按车间代码 (TB001) 筛选
            print(f"  模式: T+{offset} (仅按车间代码筛选)")
            # workshop_codes 在函数入口已验证非空
            placeholders = ', '.join('?' * len(workshop_codes))
            where_clauses.append(f"TMF.TB001 IN ({placeholders})")
            params.extend(workshop_codes)
            print(f"  筛选条件: 车间代码 (TMF.TB001 IN ({', '.join(workshop_codes)}))")

        # --- 组合最终的SQL查询语句 ---
        sql_final_query = sql_main_query_base
        if where_clauses:
            sql_final_query += " WHERE " + " AND ".join(where_clauses)

        # 添加最终排序 (调整排序，将主供应商放到料号前)
        sql_final_query += " ORDER BY TMF.MA002, TMF.TB003, TMF.MOCTAUDF01, TMF.TB001, TMF.TB002" # <<< MODIFIED >>> 排序调整

        print(f"  执行最终查询 (参数化)...")
        print(f"  参数: {params}")
        df = pd.read_sql(sql_final_query, conn, params=params)
        print(f"查询到 {len(df)} 条记录。")

        # --- 定义新的、统一的列顺序 ---
        final_columns_to_use = [
            '工单单别', '工单编号', '产品品号', '产品品名',
            '预计领用日',
            '主供应商名称',
            '料件品号', '料件品名', '料件规格', '单位', '现有库存',
            '预计用量', '库存结余', '预计入库',
            '回复到料时间', '最晚到料时间', 'PO#',
            '工作中心'
        ]
        # --- 5. 在 Pandas DataFrame 中完成数据处理、筛选、调整 ---
        print("在 Pandas 中处理数据...")
        if not df.empty:
            # 5a. 列重命名和类型转换
            column_mapping = {
                 'TB001': '工单单别', 'TB002': '工单编号', 'TA006': '产品品号',
                 'TA034': '产品品名', 'TA035': '产品规格', 'TB003': '料件品号',
                 'MB002': '料件品名', 'MB003': '料件规格', 'MB004': '单位',
                 'SUMINV': '现有库存', 'TBYJYL': '预计用量',
                 'MOCTAUDF01': '最晚到料时间_raw', # 保留原始值用于排序
                 'TC015': 'PO#',
                 'MOCTAUDF02': '预计领用日_raw',   # 保留原始值用于排序
                 'MA002': '主供应商名称', 'MD002': '工作中心',
                 'SUMMOC': 'SUMMOC', 'SUMPUR': 'SUMPUR' # 中间计算列
            }
            df.rename(columns=column_mapping, inplace=True)
            numeric_cols = ['现有库存', '预计用量', 'SUMMOC', 'SUMPUR']
            for col in numeric_cols: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

            # 5b. 计算列
            # 先按要求排序，确保 cumsum 计算正确 (排序已在 SQL 中调整)
            # df.sort_values(by=['料件品号', '最晚到料时间_raw', '工单单别', '工单编号'], inplace=True, na_position='last') # <<< REMOVED/CHANGED >>>
            # 计算库存结余 (按料件分组累加预计用量)
            df['库存结余'] = df['现有库存'] - df.groupby('料件品号')['预计用量'].cumsum()
            # 计算预计入库
            df['预计入库'] = df['SUMMOC'] + df['SUMPUR']

            # 5c. 格式化日期列 (定义辅助函数)
            def format_yyyymmdd_str(date_input):
                if pd.isna(date_input) or date_input == '': return None
                date_str = str(date_input).strip().replace('.0', '') # 清理可能的小数和空格
                if len(date_str) == 8 and date_str.isdigit():
                    try:
                        # 尝试从 YYYYMMDD 格式解析并格式化为 YYYY-MM-DD
                        return datetime.datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
                    except ValueError:
                        print(f"警告: 格式化日期 '{date_str}' 失败 (非YYYYMMDD 格式)。")
                        return date_str # 返回原始字符串
                # 添加对 YYYY-MM-DD 或 YYYY/MM/DD 格式的兼容
                if len(date_str) == 10 and (date_str[4] == '-' or date_str[4] == '/'):
                     try:
                         # 尝试直接解析并格式化为 YYYY-MM-DD
                         return pd.to_datetime(date_str).strftime('%Y-%m-%d')
                     except ValueError:
                         print(f"警告: 格式化日期 '{date_str}' 失败 (无法解析为标准日期)。")
                         return date_str
                # 如果不是已知格式，警告并返回原始值
                print(f"警告: 日期值 '{date_str}' 格式未知，将按原样返回。")
                return date_str

            df['最晚到料时间'] = df['最晚到料时间_raw'].apply(format_yyyymmdd_str)
            df['预计领用日'] = df['预计领用日_raw'].apply(format_yyyymmdd_str)

            # 5d. 筛选数据 (库存结余 < 0) - 这是 Python 版与 VBA 版的核心区别之一
            print(" - 在 DataFrame 中筛选库存结余 < 0 的行...")
            df_filtered = df[df['库存结余'] < 0].copy()
            print(f"   筛选后剩余 {len(df_filtered)} 行。")



            # --- 处理 df_filtered 为空的情况 ---
            if df_filtered.empty:
                print("信息：筛选后无缺料数据。将生成空报告。")
                # <<< MODIFIED >>> 使用调整后的列顺序定义空表结构
                final_columns_structure = final_columns_ordered_zj2 if report_name_part == "组装二" else final_columns_ordered_standard
                if report_name_part == "组装二":
                    print("  空表结构：应用'组装二'列顺序。")
                # 创建空的 DataFrame
                df_final = pd.DataFrame(columns=final_columns_structure)

            # --- 如果有缺料数据 ---
            else:
                # 5e. 调整日期 (最晚到料时间 < 今天 -> 今天)
                print(" - 在 DataFrame 中调整超期 '最晚到料时间'...")
                today_dt_str = today_date.strftime('%Y-%m-%d')
                # 先将格式化后的日期转回 datetime 对象，以便比较
                df_filtered['最晚到料时间_dt'] = pd.to_datetime(df_filtered['最晚到料时间'], errors='coerce')
                # 找到非空且小于今天的日期
                update_mask = df_filtered['最晚到料时间_dt'].notna() & (df_filtered['最晚到料时间_dt'] < pd.Timestamp(today_date))
                # 将这些日期的字符串值更新为今天
                df_filtered.loc[update_mask, '最晚到料时间'] = today_dt_str
                # 删除临时的 datetime 列
                df_filtered.drop(columns=['最晚到料时间_dt'], inplace=True)
                print(f"   日期调整完成。")

                # 5f. 插入 '回复到料时间' 列 (空列，待用户填写)
                print(" - 在 DataFrame 中插入 '回复到料时间' 列...")
                # <<< MODIFIED >>> '回复到料时间' 应该在 '库存结余' 之后, '最晚到料时间' 之前
                # 我们将在步骤 5g 中通过选择列来确保正确的位置
                if '回复到料时间' not in df_filtered.columns:
                    df_filtered['回复到料时间'] = None # 先添加列，后面会按顺序选择

                    # 5g. 选择最终列并按新的统一顺序排列
                    # --- （移除旧的基于 report_name_part 选择列顺序的 if/else 逻辑） ---
                    print(f" - 应用统一的列顺序 ({len(final_columns_to_use)} 列)...")

                    # 创建最终 DataFrame，仅包含所需列并按指定顺序排列
                    df_final = pd.DataFrame()
                    for col in final_columns_to_use:  # 使用新的统一列表
                        if col in df_filtered.columns:
                            df_final[col] = df_filtered[col]
                        else:
                            # 如果某列因意外缺失（如'回复到料时间'添加失败），添加为空列并警告
                            print(f"警告：最终列 '{col}' 在筛选结果中不存在，将添加为空列。")
                            df_final[col] = None  # 添加空列以保持顺序

                # --- 如果原始查询为空 (df.empty is True) ---
                else:  # 对应 if not df.empty:
                    print("信息：未查询到模拟发料基础数据。将生成空报告。")
                    # --- 使用新的统一列顺序定义空表结构 ---
                    print(f"  空表结构：应用统一列顺序 ({len(final_columns_to_use)} 列)。")
                    # 创建空的 DataFrame
                    df_final = pd.DataFrame(columns=final_columns_to_use)  # 使用新的统一列表

        # --- 6. 生成 Excel 文件 (Pandas) ---
        # --- 修改：硬编码输出路径 ---
        desktop_path = Path(r"E:\Users\Desktop")
        print(f"  信息：报告将保存到指定目录: {desktop_path}")
        # --- 修改结束 ---

        # 可选但推荐：检查目录是否存在，如果不存在则尝试创建
        try:
            if not desktop_path.exists():
                print(f"  警告：目标目录 '{desktop_path}' 不存在。正在尝试创建...")
                # parents=True 会创建所有必需的父目录
                # exist_ok=True 如果目录已存在，则不会引发错误
                desktop_path.mkdir(parents=True, exist_ok=True)
                print(f"  信息：已创建目录 '{desktop_path}'。")
            elif not os.access(desktop_path, os.W_OK):
                # 如果目录存在，检查写入权限
                print(f"  警告：没有写入权限到目录 '{desktop_path}'。文件保存可能会失败。")
                # 如果需要，可以在这里显示一个消息框警告
                # messagebox.showwarning("权限警告", f"脚本可能没有写入权限到目录:\n{desktop_path}\n\n文件保存可能会失败。")

        except OSError as os_err:
            # 处理目录创建过程中的错误（例如，上层目录权限不足）
            messagebox.showerror("目录错误",
                                 f"无法创建或访问目标目录:\n{desktop_path}\n\n错误: {os_err}\n\n请检查路径E:是否存在并且您有权限创建文件夹。")
            # 需要返回 False 或引发错误以干净地停止函数
            # 确保必要的清理（如关闭数据库连接）和启用按钮能够执行
            if conn: conn.rollback()  # 假设 'conn' 在此作用域可用
            enable_buttons()  # 假设 'enable_buttons' 可用
            status_var.set(f"[{report_name_part}] 报告生成失败 (目录错误)。")  # 假设这些变量可用
            root.update_idletasks()
            return False  # 停止函数执行
        except Exception as e:  # 捕获其他潜在错误
            messagebox.showerror("目录检查错误", f"检查或创建目录时发生意外错误:\n{desktop_path}\n\n错误: {e}")
            if conn: conn.rollback()
            enable_buttons()
            status_var.set(f"[{report_name_part}] 报告生成失败 (目录检查错误)。")
            root.update_idletasks()
            return False
        # 使用时间戳生成临时文件名，避免冲突
        timestamp_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')
        output_filename_original = f"TEMP_{report_name_part}_T+{offset}缺料_{timestamp_str}.xlsx"
        output_filepath_original = desktop_path / output_filename_original
        print(f"将处理后的数据写入临时文件: {output_filepath_original}")
        # 将最终的 DataFrame 写入 Excel，不包含索引
        df_final.to_excel(output_filepath_original, index=False, engine='openpyxl')
        print("临时 Excel 文件已生成。")

        # --- 7. 使用 openpyxl 进行格式化 ---
        print("应用 Excel 格式化...")
        if not output_filepath_original.exists():
            raise FileNotFoundError(f"未能找到生成的临时文件: {output_filepath_original}")

        wb_process = openpyxl.load_workbook(output_filepath_original)
        ws_process = wb_process.active
        if ws_process is None:
            raise ValueError("无法加载临时 Excel 文件的工作表。")

        # 定义样式
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # 黄色填充
        header_font = Font(name='微软雅黑', size=9, bold=True)  # 表头字体
        data_font = Font(name='微软雅黑', size=8)  # 数据区域字体
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中对齐
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 左对齐
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)  # 右对齐
        date_format = 'yyyy-mm-dd'  # 日期格式

        print(" - 格式化表头...")
        headers_final = df_final.columns.tolist()  # 获取最终的列名列表
        for col_idx, header in enumerate(headers_final, 1):  # 从1开始计数列
            cell = ws_process.cell(row=1, column=col_idx)  # 获取表头单元格
            cell.font = header_font  # 应用表头字体
            cell.alignment = center_align  # 应用居中对齐
            # 特殊列应用黄色背景 (保持不变)
            if header in ['回复到料时间', '主供应商名称']:
                cell.fill = yellow_fill

        print(" - 设置数据区域字体、格式和列宽...")
        # 定义需要右对齐和设置为日期格式的列名 (保持不变)
        right_align_headers = {'现有库存', '预计用量', '库存结余', '预计入库'}
        date_headers = {'最晚到料时间', '预计领用日', '回复到料时间'}

        # 定义固定列宽映射
        fixed_column_widths = {
            '工单单别': 5,
            '工单编号': 7.5,
            '产品品号': 10,
            '主供应商名称': 10,
            '料件品号': 7.5,
            '料件品名': 25,
            '料件规格': 15,
            '现有库存': 7.5,
            '预计用量': 7.5,
            '库存结余': 7.5,
            '预计入库': 7.5,
            '工作中心': 7
        }

        # 初始化列宽字典（对于未指定固定宽度的列）
        col_widths = {}
        for col_idx, header in enumerate(headers_final, 1):
            if header in fixed_column_widths:
                col_widths[col_idx] = fixed_column_widths[header]
            else:
                # 对于未指定固定宽度的列，计算表头字符宽度
                header_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in header)
                col_widths[col_idx] = max(header_len + 1, 6)  # 加一点边距，最小宽度为6

        # 遍历数据行进行格式化和计算非固定列的宽度
        if ws_process.max_row > 1:  # 检查是否有数据行
            for row_idx in range(2, ws_process.max_row + 1):  # 从第2行开始
                for col_idx, header in enumerate(headers_final, 1):  # 遍历列
                    cell = ws_process.cell(row=row_idx, column=col_idx)
                    cell.font = data_font  # 设置数据字体

                    # 设置默认对齐和格式
                    alignment_to_apply = left_align  # 所有单元格都设置为自动换行
                    number_format_to_apply = '@'  # 默认为文本格式

                    # 根据列名应用特定对齐和格式 (逻辑不变)
                    if header in right_align_headers:
                        alignment_to_apply = right_align
                        number_format_to_apply = '#,##0'  # 数字格式
                    elif header in date_headers:
                        number_format_to_apply = date_format  # 日期格式
                        alignment_to_apply = center_align  # 日期居中

                    cell.alignment = alignment_to_apply
                    cell.number_format = number_format_to_apply

                    # 仅对未固定宽度的列计算单元格内容宽度
                    if header not in fixed_column_widths and cell.value is not None:
                        # 将日期对象转为字符串计算宽度
                        if isinstance(cell.value, (datetime.datetime, datetime.date)):
                            cell_str = cell.value.strftime(date_format)  # 使用目标格式计算
                        else:
                            cell_str = str(cell.value)
                        # 计算字符宽度 (中文算2，英文算1)
                        cell_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_str)
                        col_widths[col_idx] = max(col_widths.get(col_idx, 0), cell_len + 1)  # 更新最大宽度

        # 应用列宽
        max_allowed_width = 60  # 非固定宽度列的最大宽度限制
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)  # 获取列字母
            if headers_final[col_idx - 1] in fixed_column_widths:
                # 使用固定宽度
                ws_process.column_dimensions[col_letter].width = width
            else:
                # 非固定宽度列，使用计算宽度但有最大限制
                adjusted_width = min(width, max_allowed_width)
                ws_process.column_dimensions[col_letter].width = adjusted_width

        print(" - 冻结首行...")
        ws_process.freeze_panes = ws_process['A2']  # 冻结第一行

        # --- 8. 重命名文件 ---
        today_mmdd = datetime.date.today().strftime('%m%d') # 获取当天月日
        # 清理报告名称部分，使其适合文件名
        safe_report_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in report_name_part)
        # 构建最终文件名
        final_filename = f"{safe_report_name}_T+{offset}缺料_{today_mmdd}.xlsx"
        final_filepath = desktop_path / final_filename # 最终文件路径
        print(f"准备将文件重命名/保存为: {final_filepath}")

        # 保存格式化后的工作簿到最终路径
        wb_process.save(final_filepath)
        print(f"文件已保存到: {final_filepath}")
        wb_process.close() # 关闭工作簿对象

        # 删除临时文件
        try:
            if output_filepath_original.exists():
                os.remove(output_filepath_original)
                print(f"已删除临时文件: {output_filepath_original}")
        except Exception as remove_err:
            print(f"警告：删除临时文件 {output_filepath_original} 时出错: {remove_err}")


        status_var.set(f"[{report_name_part}] T+{offset} 缺料报告已成功生成！")
        messagebox.showinfo("完成", f"缺料报告已生成并保存到桌面：\n{final_filepath}")
        return True # 表示成功

    # --- 异常处理部分 ---
    except pyodbc.Error as db_err:
        sqlstate = db_err.args[0]; message = str(db_err.args[1])
        print(f"数据库错误 SQLSTATE: {sqlstate}\n消息: {message}"); traceback.print_exc()
        messagebox.showerror("数据库错误", f"处理数据时发生数据库错误。\nSQLSTATE: {sqlstate}\n错误: {message}")
        status_var.set(f"报告生成失败 (数据库错误)。")
        if conn: conn.rollback() # 回滚事务
        return False # 表示失败
    except PermissionError as pe:
         target_path = final_filepath or output_filepath_original # 确定出错时操作的文件
         print(f"文件权限错误: {pe}"); traceback.print_exc()
         messagebox.showerror("文件错误", f"文件访问权限错误，可能是文件被占用或无权限。\n错误: {pe}\n文件: {target_path}")
         status_var.set(f"报告生成失败 (文件权限错误)。")
         if conn: conn.rollback() # 回滚事务
         return False
    except FileNotFoundError as fnf_err:
         print(f"文件未找到错误: {fnf_err}"); traceback.print_exc()
         messagebox.showerror("文件错误", f"未能找到所需文件。\n错误: {fnf_err}")
         status_var.set(f"报告生成失败 (文件未找到)。")
         if conn: conn.rollback() # 回滚事务
         return False
    except Exception as e:
        print(f"生成报告时发生错误: {e}"); traceback.print_exc()
        messagebox.showerror("执行错误", f"生成报告时发生意外错误。\n错误类型: {type(e).__name__}\n错误详情: {e}")
        status_var.set(f"报告生成失败 ({type(e).__name__})。")
        if conn: conn.rollback() # 回滚事务
        return False # 表示失败
    finally:
        # 确保数据库连接关闭
        if conn:
            try: conn.close(); print("数据库连接已关闭。")
            except Exception as close_err: print(f"关闭数据库连接时出错: {close_err}")
        # 确保按钮重新启用
        enable_buttons() # 调用全局函数启用按钮
        root.update_idletasks() # 更新 GUI

# -*- coding: utf-8 -*-
import pandas as pd
import pyodbc
import traceback
import os
from tkinter import messagebox
from pathlib import Path # Keep other imports as they are

# Assume PYODBC_CONN_STRING is defined elsewhere correctly

# <<<--- 更新总计划的核心逻辑函数 (再次修正版) --- >>>
def update_mocta_from_excel(excel_filepath):
    """
    读取指定的Excel文件中的“总计划”工作表，
    并根据其中的数据更新数据库 MOCTA 表的 UDF01 和 UDF02 字段。
    <<< 修改: 根据截图修正了ta001和ta002的Excel列名映射 >>>
    """
    sheet_name = "总计划" # 需要更新的工作表名称

    # <<< 修改: 更正 COLUMN_MAP 以匹配截图 --- >>>
    # 格式: '数据库逻辑字段': 'Excel中的列名'
    COLUMN_MAP = {
        'ta001': '工单单别',     # 数据库 TA001 对应 Excel 的 "工单单别" (Column B)
        'ta002': '工单单号',     # 数据库 TA002 对应 Excel 的 "工单单号" (Column A)
        'udf01': '物料到齐时间', # 数据库 UDF01 对应 Excel 的 "物料到齐时间" (Column L)
        'udf02': '开工日期'       # 数据库 UDF02 对应 Excel 的 "开工日期" (Column N)
    }
    # <<<--- 修改结束 --- >>>
    required_excel_cols = list(COLUMN_MAP.values()) # 需要从Excel读取的列名列表

    print(f"开始从 '{os.path.basename(excel_filepath)}' 的 '{sheet_name}' 工作表更新 MOCTA...")
    print(f"  将使用以下Excel列:")
    print(f"    '{COLUMN_MAP['ta001']}' (Excel) -> TA001 (DB 主键)") # 更正日志
    print(f"    '{COLUMN_MAP['ta002']}' (Excel) -> TA002 (DB 主键)") # 更正日志
    print(f"    '{COLUMN_MAP['udf01']}' (Excel) -> UDF01 (DB 更新字段)")
    print(f"    '{COLUMN_MAP['udf02']}' (Excel) -> UDF02 (DB 更新字段)")

    conn = None
    updated_count = 0
    processed_rows = 0
    skipped_count = 0
    error_count = 0
    error_flag = False # 标记是否发生错误

    try:
        print(f"  正在读取 Excel 文件: {excel_filepath}...")
        # 读取逻辑不变 (header=0, usecols, dtype, keep_default_na)
        df_plan = pd.read_excel(
            excel_filepath,
            sheet_name=sheet_name,
            header=0,
            usecols=required_excel_cols,
            dtype=str,
            keep_default_na=False
        )

        # 列存在性检查 (不变)
        missing_cols = [col for col in required_excel_cols if col not in df_plan.columns]
        if missing_cols:
             messagebox.showerror("Excel 列缺失", f"在工作表 '{sheet_name}' 中找不到以下必需列：\n{', '.join(missing_cols)}\n\n请检查Excel文件或列名映射配置。")
             print(f"  错误: 必需的Excel列缺失: {missing_cols}")
             return False

        # 数据清理 (不变)
        df_plan = df_plan.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        df_plan.replace("", None, inplace=True)

        print(f"  成功读取 {len(df_plan)} 行数据 (仅含所需列)。")

        # 数据库连接 (不变)
        print(f"  连接数据库...")
        conn = pyodbc.connect(PYODBC_CONN_STRING, autocommit=False)
        cursor = conn.cursor()
        print("  数据库连接成功。")

        print("  开始逐行处理并更新数据库...")
        # SQL 更新语句 (不变)
        sql_update = "UPDATE MOCTA SET UDF01 = ?, UDF02 = ? WHERE TA001 = ? AND TA002 = ?"

        # 遍历DataFrame行
        for index, row in df_plan.iterrows():
            excel_row_num = index + 2
            processed_rows += 1
            # <<< 修改: 变量名用 _val 后缀以示区分 >>>
            ta001_excel_val = None
            ta002_excel_val = None

            try:
                # <<< 修改: 使用修正后的 COLUMN_MAP 获取数据 >>>
                ta001_excel_val = row[COLUMN_MAP['ta001']] # 获取 "工单单别"
                ta002_excel_val = row[COLUMN_MAP['ta002']] # 获取 "工单单号"

                # 主键验证 (不变, 使用新变量名)
                if ta001_excel_val is None or ta002_excel_val is None:
                    print(f"  - 第 {excel_row_num} 行 (Excel): 跳过，缺少 '{COLUMN_MAP['ta001']}' 或 '{COLUMN_MAP['ta002']}' (值为 None)。")
                    skipped_count += 1
                    continue

                # 获取日期输入值 (不变)
                udf01_input = row[COLUMN_MAP['udf01']]
                udf02_input = row[COLUMN_MAP['udf02']]

                # 日期处理逻辑 (不变)
                udf01_db_val_str = None # 用 _db_val_str 表示最终要写入DB的值
                if udf01_input is not None:
                    try:
                        udf01_dt_obj = pd.to_datetime(udf01_input, errors='coerce')
                        if pd.notna(udf01_dt_obj):
                            udf01_db_val_str = udf01_dt_obj.date().strftime('%Y%m%d')
                        else:
                            print(f"  - 第 {excel_row_num} 行 (Excel, 工单 {ta001_excel_val}-{ta002_excel_val}): 列 '{COLUMN_MAP['udf01']}' 的值 '{udf01_input}' 格式无法解析，将设为 NULL。")
                    except Exception as date_err:
                         print(f"  - 第 {excel_row_num} 行 (Excel, 工单 {ta001_excel_val}-{ta002_excel_val}): 解析列 '{COLUMN_MAP['udf01']}' 的值 '{udf01_input}' 时出错: {date_err}，将设为 NULL。")

                udf02_db_val_str = None
                if udf02_input is not None:
                    try:
                        udf02_dt_obj = pd.to_datetime(udf02_input, errors='coerce')
                        if pd.notna(udf02_dt_obj):
                            udf02_db_val_str = udf02_dt_obj.date().strftime('%Y%m%d')
                        else:
                            print(f"  - 第 {excel_row_num} 行 (Excel, 工单 {ta001_excel_val}-{ta002_excel_val}): 列 '{COLUMN_MAP['udf02']}' 的值 '{udf02_input}' 格式无法解析，将设为 NULL。")
                    except Exception as date_err:
                        print(f"  - 第 {excel_row_num} 行 (Excel, 工单 {ta001_excel_val}-{ta002_excel_val}): 解析列 '{COLUMN_MAP['udf02']}' 的值 '{udf02_input}' 时出错: {date_err}，将设为 NULL。")

                # <<< 修改: 准备参数时，顺序必须严格对应 SQL 语句中的 ? >>>
                # SQL: SET UDF01 = ?, UDF02 = ? WHERE TA001 = ? AND TA002 = ?
                params = (
                    udf01_db_val_str,    # 第一个 ? (UDF01)
                    udf02_db_val_str,    # 第二个 ? (UDF02)
                    str(ta001_excel_val),# 第三个 ? (TA001 - 来自 "工单单别")
                    str(ta002_excel_val) # 第四个 ? (TA002 - 来自 "工单单号")
                )
                # <<<--- 修改结束 --- >>>

                # 执行更新 (不变)
                cursor.execute(sql_update, params)
                if cursor.rowcount > 0:
                     updated_count += cursor.rowcount
                # else: # 可选警告

            # 错误处理 (不变, 但引用修正后的列名)
            except Exception as row_err:
                wo_info = f"{row.get(COLUMN_MAP['ta001'], '未知')}-{row.get(COLUMN_MAP['ta002'], '未知')}"
                error_message = f"  - 处理第 {excel_row_num} 行 (Excel, 工单: {wo_info}) 时出错: {row_err}"
                print(error_message)
                traceback.print_exc()
                error_count += 1
                error_flag = True

        # 事务处理、结束逻辑、异常捕获 (不变)
        if error_flag:
            print(f"\n处理完成，但遇到 {error_count} 个错误。事务将回滚，数据库未做任何更改。")
            conn.rollback()
            messagebox.showerror("更新失败", f"处理 Excel 文件时发生 {error_count} 个错误。\n数据库未做任何更改。\n请检查 Excel 文件内容和控制台输出获取详细信息。")
            return False
        else:
            print(f"\n所有 {processed_rows} 行处理完毕，无错误。准备提交事务...")
            conn.commit()
            print(f"  事务已提交。")
            info_msg = f"成功处理文件 '{os.path.basename(excel_filepath)}'.\n\n" \
                       f"总共处理 Excel 数据行数: {processed_rows}\n" \
                       f"数据库更新记录数: {updated_count}\n" \
                       f"(更新数为0可能表示Excel中的工单在数据库中不存在或日期值无需更新)\n" \
                       f"跳过行数 (缺少主键): {skipped_count}"
            print(info_msg.replace('\n\n', '\n').replace('\n', '\n  '))
            messagebox.showinfo("更新成功", info_msg)
            return True

    except FileNotFoundError:
        # ... (其他异常处理保持不变) ...
        messagebox.showerror("文件错误", f"找不到指定的 Excel 文件：\n{excel_filepath}")
        print(f"错误: 文件未找到 {excel_filepath}")
        return False
    except ValueError as ve:
         if 'missing_cols' in locals() and missing_cols:
             pass
         else:
             messagebox.showerror("Excel 读取错误", f"读取 Excel 文件 '{os.path.basename(excel_filepath)}' 的 '{sheet_name}' 工作表时出错。\n请确保工作表存在、格式正确且文件未损坏。\n错误: {ve}")
             print(f"错误: 读取Excel时发生 ValueError: {ve}")
             traceback.print_exc()
         return False
    except pyodbc.Error as db_err:
        sqlstate = db_err.args[0]; message = str(db_err.args[1])
        print(f"数据库操作错误 SQLSTATE: {sqlstate}\n消息: {message}"); traceback.print_exc()
        messagebox.showerror("数据库错误", f"更新数据库时发生错误。\nSQLSTATE: {sqlstate}\n错误: {message}")
        if conn: conn.rollback()
        return False
    except Exception as e:
        print(f"更新总计划时发生意外错误: {e}"); traceback.print_exc()
        messagebox.showerror("执行错误", f"更新总计划时发生意外错误。\n错误类型: {type(e).__name__}\n错误详情: {e}")
        if conn and not error_flag:
             try:
                 conn.rollback()
                 print("  因意外错误，事务已回滚。")
             except Exception as rb_err:
                 print(f"  尝试回滚事务时出错: {rb_err}")
        return False
    finally:
        if conn:
            try: conn.close(); print("  数据库连接已关闭。")
            except Exception as close_err: print(f"关闭数据库连接时出错: {close_err}")
# <<<--- 函数修改结束 --- >>>

# --- “更新总计划”按钮的触发函数 ---
def run_update_master_plan():
    """
    处理“更新总计划到数据库”按钮点击事件。
    """
    master_plan_file = master_plan_file_var.get()

    if not master_plan_file:
        messagebox.showerror("错误", "请先选择包含'总计划'工作表的 Excel 文件！")
        return
    if not os.path.exists(master_plan_file):
        messagebox.showerror("错误", f"选择的文件不存在：\n{master_plan_file}")
        return
    if not (master_plan_file.lower().endswith('.xlsx') or \
            master_plan_file.lower().endswith('.xls') or \
            master_plan_file.lower().endswith('.xlsm')):
        messagebox.showerror("文件类型错误", "请选择有效的 Excel 文件 (.xlsx, .xls, .xlsm)。")
        return

    confirm = messagebox.askyesno(
        title="确认更新",
        message=f"确定要使用选定的 Excel 文件:\n'{os.path.basename(master_plan_file)}'\n中的'总计划'工作表数据更新数据库 MOCTA 表的 UDF01 和 UDF02 字段吗？\n\n此操作将修改数据库记录 (仅针对文件中存在的工单)，请确保数据准确无误。"
    )

    if confirm:
        disable_buttons()
        status_var.set(f"正在从 {os.path.basename(master_plan_file)} 更新总计划到数据库...")
        root.update_idletasks()

        # 调用核心更新函数
        success = update_mocta_from_excel(master_plan_file)

        # 更新状态栏信息
        if success:
            status_var.set("总计划数据更新成功！(详情见弹窗)")
        else:
            status_var.set("总计划数据更新失败或包含错误。请查看控制台信息。")

        enable_buttons() # 无论成功失败都重新启用按钮
        root.update_idletasks()
    else:
        status_var.set("总计划更新操作已取消。")
        # messagebox.showinfo("已取消", "更新操作已取消。") # 取消时可以不弹窗
        root.update_idletasks()

# --- 辅助函数 (disable_buttons, enable_buttons) ---
# --- 辅助函数 (disable_buttons, enable_buttons) ---
def disable_buttons():
    # 使用 try-except 避免按钮不存在或已被销毁时出错
    # <<< MODIFIED: 添加新按钮名称 >>>
    buttons_to_manage = ['btn_process', 'btn_simulate_t1', 'btn_simulate_t2', # <<< 添加 btn_simulate_t1 >>>
                         'btn_simulate_t7', 'btn_simulate_t15', 'btn_update_plan',
                         'btn_generate_all']
    for btn_name in buttons_to_manage:
         try:
             widget = globals().get(btn_name)
             if widget and widget.winfo_exists():
                 widget.config(state=tk.DISABLED)
         except (tk.TclError, NameError, KeyError):
             pass


def enable_buttons():
    # 检查库是否成功导入
    pyodbc_imported = 'pyodbc' in globals() and libraries_ok
    pandas_imported = 'pd' in globals() and libraries_ok
    openpyxl_imported = 'openpyxl' in globals() and libraries_ok

    db_ok = pyodbc_imported # 数据库操作需要 pyodbc
    excel_read_ok = pandas_imported and openpyxl_imported # Excel 读取需要 pandas 和 openpyxl
    xlsm_write_ok = openpyxl_imported # XLSM 写入需要 openpyxl
    excel_write_ok = pandas_imported and openpyxl_imported # 普通 Excel 写入

    # <<< MODIFIED: 添加新按钮及其条件 >>>
    buttons_to_manage = {
        'btn_process': excel_read_ok and xlsm_write_ok, # 维护数据: 读Excel, 写XLSM
        'btn_simulate_t1': db_ok and excel_write_ok, # T+1 齐套: 读DB, 写Excel # <<< 添加 T+1 条件 >>>
        'btn_simulate_t2': db_ok and excel_read_ok and excel_write_ok, # T+2: 读DB, 读Excel, 写Excel
        'btn_simulate_t7': db_ok and excel_write_ok, # T+7: 读DB, 写Excel
        'btn_simulate_t15': db_ok and excel_write_ok, # T+15: 读DB, 写Excel
        'btn_update_plan': db_ok and excel_read_ok, # 更新总计划: 读DB, 读Excel
        'btn_generate_all': db_ok and excel_write_ok # 全部缺料: 读DB, 写Excel
    }

    for btn_name, condition in buttons_to_manage.items():
        try:
            widget = globals().get(btn_name)
            if widget and widget.winfo_exists():
                state = tk.NORMAL if condition else tk.DISABLED
                widget.config(state=state)
        except (tk.TclError, NameError, KeyError):
            pass

# --- 新增：生成全部缺料报告的核心逻辑 ---
# <<< MODIFIED function signature and logic >>>
def generate_all_shortages_report(workshop_codes, report_name_part, use_exclusion=False):
    """
    纯 Python 实现，生成指定车间（或排除指定车间）的缺料报告。
    使用非常遥远的未来日期作为数据过滤的目标日期。
    *** 列顺序已调整: 主供应商名称 在 料件品号 之前 ***

    Args:
        workshop_codes (list): 用于筛选的工单单别 (TB001) 列表。
        report_name_part (str): 用于文件名和状态信息的报告名称部分。
        use_exclusion (bool): 如果为 True, 则使用 NOT IN 筛选 workshop_codes;
                              如果为 False, 则使用 IN 筛选 workshop_codes.
                              如果 workshop_codes 为空, 则不进行筛选 (查询所有).
    """
    # report_name_part 不再硬编码
    print(f"--- 开始生成 [{report_name_part}] 缺料报告 (基于单别筛选) ---")
    print(f"  筛选单别: {workshop_codes}")
    print(f"  筛选模式: {'排除 (NOT IN)' if use_exclusion else '包含 (IN)'}")

    disable_buttons()
    status_var.set(f"正在为 [{report_name_part}] 生成缺料报告 (Python)...")
    root.update_idletasks() # 确保 GUI 更新

    conn = None
    output_filepath_original = None
    final_filepath = None

    try:
        # --- 1. 定义目标日期 (保持不变, 使用遥远未来) ---
        target_date_yyyymmdd = '99991231'
        print(f"使用目标日期 (用于数据准备 SQL 过滤): {target_date_yyyymmdd}")
        today_date = datetime.date.today()

        # --- 2. 连接数据库 (保持不变) ---
        print(f"连接数据库...")
        conn = pyodbc.connect(PYODBC_CONN_STRING, autocommit=False)
        cursor = conn.cursor()
        print("数据库连接成功。")

        # --- 3. 执行数据准备 SQL (保持不变 - 除了 TEMPMNFL 的 SQL) ---
        print(f"执行数据准备 SQL (基于日期 {target_date_yyyymmdd} - 注意：实际日期过滤将在主查询应用或被忽略)...")
        tables_to_clear = ["TEMPMNFL", "TEMPINVMC", "TEMPMOCTA", "TEMPPURTD"]
        for table in tables_to_clear:
             # print(f" - 清空 {table}...") # 减少日志输出
             cursor.execute(f"DELETE FROM {table}")

        print(" - 插入 TEMPMNFL (全部缺料逻辑 - 允许 UDF01 为空)...") # 修改日志说明
        # <<< MODIFICATION START for 'All Shortages': Allow NULL/Empty TA.UDF01 >>>
        sql_insert_tempmfl = f"""
        INSERT INTO TEMPMNFL(TB001,TB002,TA006,TA034,TA035,TB003,MB002,MB003,MB004,TBYJYL,MOCTAUDF01,TC015,MOCTAUDF02,MA002,MD002)
        SELECT TB.TB001, TB.TB002, TA.TA006, TA.TA034, TA.TA035, TB.TB003, MB.MB002, MB.MB003, MB.MB004, (TB.TB004-TB.TB005) AS TBYJYL, TA.UDF01 AS MOCTAUDF01, TC.TC015, TA.UDF02 AS MOCTAUDF02, MA.MA002, MD.MD002
        FROM MOCTB AS TB
        INNER JOIN MOCTA AS TA ON TA.TA001 = TB.TB001 AND TA.TA002 = TB.TB002
        LEFT JOIN COPTD AS TD ON TD.TD001 = TA.TA026 AND TD.TD002 = TA.TA027 AND TD.TD003 = TA.TA028
        LEFT JOIN COPTC AS TC ON TC.TC001 = TD.TD001 AND TC.TC002 = TD.TD002
        LEFT JOIN INVMB AS MB ON MB.MB001 = TB.TB003
        LEFT JOIN PURMA AS MA ON MA.MA001 = MB.MB032
        LEFT JOIN CMSMD AS MD ON MD.MD001 = TA.TA021
        WHERE TA.TA011 IN ('1','2','3') /*工单状态*/
          AND (TB.TB009 IN ('700', '710')) /*用料库别*/
          AND TA.TA013 = 'Y' /*审核*/
          AND TB.TB004-TB.TB005 > 0 /*预计用量>已发量*/
          AND TB.TB011 IN ('1','2') /*用料确认码*/
        /* <<< 条件已移除: AND TA.UDF01 IS NOT NULL AND TA.UDF01 <> '' >>> */
        /* <<< 条件已移除: AND TA.UDF01 <= ? (因为目标日期是99991231) >>> */
        ORDER BY TB.TB003, TA.UDF01 /* 排序保留UDF01可能导致NULL在前或后，但不影响最终计算 */
        """
        # 注意：执行此SQL时不再需要传递 target_date_yyyymmdd 参数
        cursor.execute(sql_insert_tempmfl)
        # <<< MODIFICATION END >>>
        print(f"   {cursor.rowcount} 行插入 TEMPMNFL (允许UDF01为空)。") # 修改日志说明

        # --- 插入 TEMPINVMC, TEMPMOCTA, TEMPPURTD 的 SQL 保持不变 ---
        # print(" - 插入 TEMPINVMC...")
        sql_insert_tempinvmc = """
        INSERT INTO TEMPINVMC(INVMB001,SUMINVMC007) SELECT MB.MB001,SUM(ISNULL(INV.MC007,0)) AS SUMINVMC007 FROM INVMB AS MB LEFT JOIN INVMC AS INV ON INV.MC001=MB.MB001 LEFT JOIN CMSMC AS CMS ON CMS.MC001=INV.MC002 WHERE CMS.MC005='Y' AND (INV.MC002 IN ('700', '710')) AND MB.MB001 IN (SELECT TB003 FROM TEMPMNFL) GROUP BY MB.MB001
        """
        cursor.execute(sql_insert_tempinvmc)
        # print(f"   {cursor.rowcount} 行插入 TEMPINVMC。")

        # print(" - 插入 TEMPMOCTA...")
        # 注意: TEMPMOCTA 和 TEMPPURTD 仍然使用 target_date_yyyymmdd (99991231)
        #      这意味着所有未来的供应都会被考虑
        sql_insert_tempmocta = f"""
        INSERT INTO TEMPMOCTA(MOCTA006,MOCMOUNT) SELECT TA.TA006, ISNULL(SUM(TA.TA015),0) - ISNULL(SUM(TA.TA017),0) AS MOUNT FROM MOCTA AS TA LEFT JOIN CMSMC AS CMS ON TA.TA020 = CMS.MC001 WHERE CMS.MC005 = 'Y' AND TA.TA013 = 'Y' AND (TA.TA011 NOT IN ('Y', 'y')) AND (TA.TA020 IN ('700', '710')) AND TA.TA006 IN (SELECT TB003 FROM TEMPMNFL) AND TA.TA010 <= ? GROUP BY TA.TA006
        """
        cursor.execute(sql_insert_tempmocta, target_date_yyyymmdd) # 仍然传递日期参数
        # print(f"   {cursor.rowcount} 行插入 TEMPMOCTA。")

        # print(" - 插入 TEMPPURTD...")
        sql_insert_temppurtd = f"""
        INSERT INTO TEMPPURTD(PURTD004,PURMOUNT) SELECT TD.TD004, ISNULL(SUM(TD.TD008),0) - ISNULL(SUM(TD.TD015),0) AS MOUNT FROM PURTD TD INNER JOIN PURTC TC ON TC.TC001 = TD.TD001 AND TC.TC002 = TD.TD002 INNER JOIN CMSMC CMS ON TD.TD007 = CMS.MC001 LEFT JOIN INVMB MB ON TD.TD004 = MB.MB001 WHERE CMS.MC005 = 'Y' AND (TD.TD007 IN ('700', '710')) AND TD.TD018 = 'Y' AND TD.TD016 = 'N' AND TD.TD004 IN (SELECT TB003 FROM TEMPMNFL) AND TD.TD012 <= ? GROUP BY TD.TD004 HAVING (ISNULL(SUM(TD.TD008),0) - ISNULL(SUM(TD.TD015),0)) > 0
        """
        cursor.execute(sql_insert_temppurtd, target_date_yyyymmdd) # 仍然传递日期参数
        # print(f"   {cursor.rowcount} 行插入 TEMPPURTD。")

        conn.commit()
        print("数据准备 SQL 执行完毕并已提交 (全部缺料逻辑调整)。") # 修改日志说明

        # --- 4. 执行主查询 SQL ---
        # <<< MODIFIED: Dynamically add WHERE clause >>>
        print("执行主查询 SQL (带单别筛选)...")
        sql_main_query_base = """
        SELECT TMF.TB001, TMF.TB002, TMF.TA006, TMF.TA034, TMF.TA035, TMF.TB003, TMF.MB002, TMF.MB003, TMF.MB004, TMF.TBYJYL,
               TMF.MOCTAUDF01, TMF.TC015, TMF.MOCTAUDF02, TMF.MA002, TMF.MD002,
               ISNULL(TI.SUMINVMC007,0) AS SUMINV,
               ISNULL(TMOC.MOCMOUNT,0) AS SUMMOC,
               ISNULL(TPUR.PURMOUNT,0) AS SUMPUR
        FROM TEMPMNFL AS TMF
        LEFT JOIN TEMPINVMC AS TI ON TI.INVMB001 = TMF.TB003
        LEFT JOIN TEMPMOCTA AS TMOC ON TMOC.MOCTA006 = TMF.TB003
        LEFT JOIN TEMPPURTD AS TPUR ON TPUR.PURTD004 = TMF.TB003
        """

        where_clauses = []
        params = []

        # 根据 workshop_codes 和 use_exclusion 添加筛选条件
        if workshop_codes: # 仅在提供了单别列表时添加子句
            placeholders = ', '.join('?' * len(workshop_codes))
            if use_exclusion:
                where_clauses.append(f"TMF.TB001 NOT IN ({placeholders})")
                print(f"  筛选条件: TMF.TB001 NOT IN ({', '.join(workshop_codes)})")
            else:
                where_clauses.append(f"TMF.TB001 IN ({placeholders})")
                print(f"  筛选条件: TMF.TB001 IN ({', '.join(workshop_codes)})")
            params.extend(workshop_codes)
        else:
            # 理论上不应执行到这里，因为调用前应有验证
            print("  警告: 未提供筛选单别，将查询所有数据。")

        # --- 组合最终的SQL查询语句 ---
        sql_final_query = sql_main_query_base
        if where_clauses:
            sql_final_query += " WHERE " + " AND ".join(where_clauses) # 如果未来需要多个条件，可以用 AND 连接

        # 添加最终排序 (保持不变)
        sql_final_query += " ORDER BY TMF.MA002, TMF.TB003, TMF.MOCTAUDF01, TMF.TB002, TMF.TB001"

        print(f"  执行最终查询...")
        if params:
            print(f"  参数: {params}")
            df = pd.read_sql(sql_final_query, conn, params=params)
        else:
            df = pd.read_sql(sql_final_query, conn) # 无参数查询
        print(f"查询到 {len(df)} 条记录。")

        # --- 5. 在 Pandas DataFrame 中完成数据处理、筛选、调整 ---
        #    (这部分逻辑与原函数相同：列重命名、计算、格式化、筛选库存结余<0、调整日期、插入列)
        print("在 Pandas 中处理数据...")
        if not df.empty:
             # 5a. 列重命名和类型转换
             column_mapping = {
                 'TB001': '工单单别', 'TB002': '工单编号', 'TA006': '产品品号', 'TA034': '产品品名', 'TA035': '产品规格',
                 'TB003': '料件品号', 'MB002': '料件品名', 'MB003': '料件规格', 'MB004': '单位',
                 'SUMINV': '现有库存', 'TBYJYL': '预计用量', 'MOCTAUDF01': '最晚到料时间_raw', 'TC015': 'PO#',
                 'MOCTAUDF02': '预计领用日_raw', 'MA002': '主供应商名称', 'MD002': '工作中心',
                 'SUMMOC': 'SUMMOC', 'SUMPUR': 'SUMPUR'
             }
             df.rename(columns=column_mapping, inplace=True)
             numeric_cols = ['现有库存', '预计用量', 'SUMMOC', 'SUMPUR']
             for col in numeric_cols: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

             # 5b. 计算列
             df['库存结余'] = df['现有库存'] - df.groupby('料件品号')['预计用量'].cumsum()
             df['预计入库'] = df['SUMMOC'] + df['SUMPUR']

             # 5c. 格式化日期列 (定义辅助函数)
             # (假设 format_yyyymmdd_str 函数已定义或在此处定义)
             if 'format_yyyymmdd_str' not in globals():
                def format_yyyymmdd_str(date_input):
                   if pd.isna(date_input) or date_input == '': return None
                   date_str = str(date_input).strip().replace('.0', '') # 清理可能的小数和空格
                   if len(date_str) == 8 and date_str.isdigit():
                       try:
                           # 尝试从𒀗MMDD 格式解析并格式化为 YYYY-MM-DD
                           return datetime.datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
                       except ValueError:
                           # print(f"警告: 格式化日期 '{date_str}' 失败 (非YYYYMMDD 格式)。") # 减少日志
                           return date_str # 返回原始字符串
                   # 添加对 YYYY-MM-DD 或 YYYY/MM/DD 格式的兼容
                   if len(date_str) == 10 and (date_str[4] == '-' or date_str[4] == '/'):
                        try:
                            # 尝试直接解析并格式化为 YYYY-MM-DD
                            return pd.to_datetime(date_str).strftime('%Y-%m-%d')
                        except ValueError:
                            # print(f"警告: 格式化日期 '{date_str}' 失败 (无法解析为标准日期)。") # 减少日志
                            return date_str
                   # 如果不是已知格式，警告并返回原始值
                   # print(f"警告: 日期值 '{date_str}' 格式未知，将按原样返回。") # 减少日志
                   return date_str
             df['最晚到料时间'] = df['最晚到料时间_raw'].apply(format_yyyymmdd_str)
             df['预计领用日'] = df['预计领用日_raw'].apply(format_yyyymmdd_str)

             # 5d. 筛选数据 (库存结余 < 0)
             # print(" - 在 DataFrame 中筛选库存结余 < 0 的行...") # 减少日志输出
             df_filtered = df[df['库存结余'] < 0].copy()
             print(f"   筛选库存结余 < 0 后剩余 {len(df_filtered)} 行。")

             # 定义最终列顺序 (标准顺序)
             # --- 定义新的、统一的列顺序 ---
             final_columns_to_use = [
                 '工单单别', '工单编号', '产品品号', '产品品名',
                 '预计领用日',
                 '主供应商名称',
                 '料件品号', '料件品名', '料件规格', '单位', '现有库存',
                 '预计用量', '库存结余', '预计入库',
                 '回复到料时间', '最晚到料时间', 'PO#',
                 '工作中心'
             ]

             if df_filtered.empty:
                 print("信息：筛选后无缺料数据。将生成空报告。")
                 df_final = pd.DataFrame(columns=final_columns_to_use)
             else:
                 # 5e. 调整日期 (最晚到料时间 < 今天 -> 今天) - 修改版，增强对空/无效日期的规避
                 print(" - 在 DataFrame 中调整超期 '最晚到料时间' (已优化空值/无效值处理)...")
                 today_dt_str = today_date.strftime('%Y-%m-%d')
                 today_ts = pd.Timestamp(today_date)  # 预先计算 Timestamp 对象

                 # 首先，识别哪些行在格式化后有一个非 None 的 '最晚到料时间' 值
                 # 这是进行日期比较和调整的前提
                 has_date_str_mask = df_filtered['最晚到料时间'].notna()

                 # 创建临时的日期对象列，初始化为 NaT (Not a Time)
                 df_filtered['最晚到料时间_dt'] = pd.NaT

                 update_count = 0  # 用于计数实际更新了多少行

                 # 仅对那些格式化后有非 None 值的行进行处理
                 if has_date_str_mask.any():
                     # 尝试将这些非 None 的字符串转换为日期对象
                     # 只在 has_date_str_mask 为 True 的子集上操作
                     df_filtered.loc[has_date_str_mask, '最晚到料时间_dt'] = pd.to_datetime(
                         df_filtered.loc[has_date_str_mask, '最晚到料时间'],
                         errors='coerce'  # 继续使用 coerce 以防某些字符串无法转换
                     )

                     # 现在，基于转换结果创建最终的更新掩码
                     # 条件1: 转换后是一个有效的日期 (非 NaT)
                     # 条件2: 这个有效日期确实早于今天
                     # 注意：这两个条件都只在原始 has_date_str_mask 为 True 的行上评估才有意义
                     final_update_mask = has_date_str_mask & \
                                         df_filtered['最晚到料时间_dt'].notna() & \
                                         (df_filtered['最晚到料时间_dt'] < today_ts)

                     # 应用更新，仅更新 final_update_mask 为 True 的行
                     if final_update_mask.any():
                         update_count = final_update_mask.sum()
                         df_filtered.loc[final_update_mask, '最晚到料时间'] = today_dt_str
                         print(f"   {update_count} 行的超期 '最晚到料时间' 已调整为今天。")
                     else:
                         print("   没有需要调整为今天的超期 '最晚到料时间'。")
                 else:
                     print("   没有找到有效的 '最晚到料时间' 字符串进行日期调整。")

                 # 无论如何，最后都删除临时的 _dt 列
                 if '最晚到料时间_dt' in df_filtered.columns:
                     df_filtered.drop(columns=['最晚到料时间_dt'], inplace=True)

                 print(f"   日期调整逻辑完成。")
                 # 5f. 插入 '回复到料时间' 列
                 # print(" - 在 DataFrame 中插入 '回复到料时间' 列...") # 减少日志输出
                 if '回复到料时间' not in df_filtered.columns:
                     df_filtered['回复到料时间'] = None

                     # 5g. 选择最终列并按新的统一顺序排列
                     print(f" - 应用统一的列顺序 ({len(final_columns_to_use)} 列)...")
                     df_final = pd.DataFrame()
                     for col in final_columns_to_use:  # 使用新的统一列表
                         if col in df_filtered.columns:
                             df_final[col] = df_filtered[col]
                         else:
                             print(f"警告：最终列 '{col}' 在筛选结果中不存在，将添加为空列。")
                             df_final[col] = None

                 else:  # 原始查询为空
                     print("信息：未查询到符合单别条件的基础数据。将生成空报告。")
                     # --- 使用新的统一列顺序定义空表结构 ---
                     print(f"  空表结构：应用统一列顺序 ({len(final_columns_to_use)} 列)。")
                     df_final = pd.DataFrame(columns=final_columns_to_use)  # 使用新的统一列表

        # --- 6. 生成 Excel 文件 (Pandas) ---
        # <<< MODIFIED: Use passed report_name_part in filename >>>
        # --- 修改：硬编码输出路径 ---
        desktop_path = Path(r"E:\Users\Desktop")
        print(f"  信息：报告将保存到指定目录: {desktop_path}")
        # --- 修改结束 ---

        # 可选但推荐：检查目录是否存在，如果不存在则尝试创建
        try:
            if not desktop_path.exists():
                print(f"  警告：目标目录 '{desktop_path}' 不存在。正在尝试创建...")
                # parents=True 会创建所有必需的父目录
                # exist_ok=True 如果目录已存在，则不会引发错误
                desktop_path.mkdir(parents=True, exist_ok=True)
                print(f"  信息：已创建目录 '{desktop_path}'。")
            elif not os.access(desktop_path, os.W_OK):
                # 如果目录存在，检查写入权限
                print(f"  警告：没有写入权限到目录 '{desktop_path}'。文件保存可能会失败。")
                # 如果需要，可以在这里显示一个消息框警告
                # messagebox.showwarning("权限警告", f"脚本可能没有写入权限到目录:\n{desktop_path}\n\n文件保存可能会失败。")

        except OSError as os_err:
            # 处理目录创建过程中的错误（例如，上层目录权限不足）
            messagebox.showerror("目录错误",
                                 f"无法创建或访问目标目录:\n{desktop_path}\n\n错误: {os_err}\n\n请检查路径E:是否存在并且您有权限创建文件夹。")
            # 需要返回 False 或引发错误以干净地停止函数
            # 确保必要的清理（如关闭数据库连接）和启用按钮能够执行
            if conn: conn.rollback()  # 假设 'conn' 在此作用域可用
            enable_buttons()  # 假设 'enable_buttons' 可用
            status_var.set(f"[{report_name_part}] 报告生成失败 (目录错误)。")  # 假设这些变量可用
            root.update_idletasks()
            return False  # 停止函数执行
        except Exception as e:  # 捕获其他潜在错误
            messagebox.showerror("目录检查错误", f"检查或创建目录时发生意外错误:\n{desktop_path}\n\n错误: {e}")
            if conn: conn.rollback()
            enable_buttons()
            status_var.set(f"[{report_name_part}] 报告生成失败 (目录检查错误)。")
            root.update_idletasks()
            return False

        timestamp_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')
        # 使用传入的 report_name_part
        safe_report_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in report_name_part) # 清理文件名
        output_filename_original = f"TEMP_{safe_report_name}_缺料_{timestamp_str}.xlsx"
        output_filepath_original = desktop_path / output_filename_original
        print(f"将处理后的数据写入临时文件: {output_filepath_original}")
        df_final.to_excel(output_filepath_original, index=False, engine='openpyxl')
        print("临时 Excel 文件已生成。")

        # --- 7. 使用 openpyxl 进行格式化 ---
        print("应用 Excel 格式化...")
        if not output_filepath_original.exists():
            raise FileNotFoundError(f"未能找到生成的临时文件: {output_filepath_original}")

        wb_process = openpyxl.load_workbook(output_filepath_original)
        ws_process = wb_process.active
        if ws_process is None:
            raise ValueError("无法加载临时 Excel 文件的工作表。")

        # 定义样式
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # 黄色填充
        header_font = Font(name='微软雅黑', size=9, bold=True)  # 表头字体
        data_font = Font(name='微软雅黑', size=8)  # 数据区域字体
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中对齐
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 左对齐
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)  # 右对齐
        date_format = 'yyyy-mm-dd'  # 日期格式

        print(" - 格式化表头...")
        headers_final = df_final.columns.tolist()  # 获取最终的列名列表
        for col_idx, header in enumerate(headers_final, 1):  # 从1开始计数列
            cell = ws_process.cell(row=1, column=col_idx)  # 获取表头单元格
            cell.font = header_font  # 应用表头字体
            cell.alignment = center_align  # 应用居中对齐
            # 特殊列应用黄色背景 (保持不变)
            if header in ['回复到料时间', '主供应商名称']:
                cell.fill = yellow_fill

        print(" - 设置数据区域字体、格式和列宽...")
        # 定义需要右对齐和设置为日期格式的列名 (保持不变)
        right_align_headers = {'现有库存', '预计用量', '库存结余', '预计入库'}
        date_headers = {'最晚到料时间', '预计领用日', '回复到料时间'}

        # 定义固定列宽映射
        fixed_column_widths = {
            '工单单别': 5,
            '工单编号': 7.5,
            '产品品号': 10,
            '主供应商名称': 10,
            '料件品号': 7.5,
            '料件品名': 25,
            '料件规格': 15,
            '现有库存': 7.5,
            '预计用量': 7.5,
            '库存结余': 7.5,
            '预计入库': 7.5,
            '工作中心': 7
        }

        # 初始化列宽字典（对于未指定固定宽度的列）
        col_widths = {}
        for col_idx, header in enumerate(headers_final, 1):
            if header in fixed_column_widths:
                col_widths[col_idx] = fixed_column_widths[header]
            else:
                # 对于未指定固定宽度的列，计算表头字符宽度
                header_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in header)
                col_widths[col_idx] = max(header_len + 1, 6)  # 加一点边距，最小宽度为6

        # 遍历数据行进行格式化和计算非固定列的宽度
        if ws_process.max_row > 1:  # 检查是否有数据行
            for row_idx in range(2, ws_process.max_row + 1):  # 从第2行开始
                for col_idx, header in enumerate(headers_final, 1):  # 遍历列
                    cell = ws_process.cell(row=row_idx, column=col_idx)
                    cell.font = data_font  # 设置数据字体

                    # 设置默认对齐和格式
                    alignment_to_apply = left_align  # 所有单元格都设置为自动换行
                    number_format_to_apply = '@'  # 默认为文本格式

                    # 根据列名应用特定对齐和格式 (逻辑不变)
                    if header in right_align_headers:
                        alignment_to_apply = right_align
                        number_format_to_apply = '#,##0'  # 数字格式
                    elif header in date_headers:
                        number_format_to_apply = date_format  # 日期格式
                        alignment_to_apply = center_align  # 日期居中

                    cell.alignment = alignment_to_apply
                    cell.number_format = number_format_to_apply

                    # 仅对未固定宽度的列计算单元格内容宽度
                    if header not in fixed_column_widths and cell.value is not None:
                        # 将日期对象转为字符串计算宽度
                        if isinstance(cell.value, (datetime.datetime, datetime.date)):
                            cell_str = cell.value.strftime(date_format)  # 使用目标格式计算
                        else:
                            cell_str = str(cell.value)
                        # 计算字符宽度 (中文算2，英文算1)
                        cell_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_str)
                        col_widths[col_idx] = max(col_widths.get(col_idx, 0), cell_len + 1)  # 更新最大宽度

        # 应用列宽
        max_allowed_width = 60  # 非固定宽度列的最大宽度限制
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)  # 获取列字母
            if headers_final[col_idx - 1] in fixed_column_widths:
                # 使用固定宽度
                ws_process.column_dimensions[col_letter].width = width
            else:
                # 非固定宽度列，使用计算宽度但有最大限制
                adjusted_width = min(width, max_allowed_width)
                ws_process.column_dimensions[col_letter].width = adjusted_width

        print(" - 冻结首行...")
        ws_process.freeze_panes = ws_process['A2']  # 冻结第一行

        # --- 8. 重命名文件 ---
        # <<< MODIFIED: Use passed report_name_part in filename >>>
        today_mmdd = datetime.date.today().strftime('%m%d') # 获取当天月日
        # 使用清理后的传入 report_name_part
        final_filename = f"{safe_report_name}_模拟发料_{today_mmdd}.xlsx"
        final_filepath = desktop_path / final_filename # 最终文件路径
        print(f"准备将文件重命名/保存为: {final_filepath}")

        # 保存格式化后的工作簿到最终路径
        wb_process.save(final_filepath)
        print(f"文件已保存到: {final_filepath}")
        wb_process.close() # 关闭工作簿对象

        # 删除临时文件
        try:
            if output_filepath_original.exists():
                os.remove(output_filepath_original)
                print(f"已删除临时文件: {output_filepath_original}")
        except Exception as remove_err:
            print(f"警告：删除临时文件 {output_filepath_original} 时出错: {remove_err}")


        # <<< MODIFIED: Use passed report_name_part in messages >>>
        status_var.set(f"[{report_name_part}] 缺料报告已成功生成！")
        messagebox.showinfo("完成", f"缺料报告已生成并保存到桌面：\n{final_filepath}")
        return True # 表示成功

    # --- 异常处理部分 (保持不变, 但更新状态消息) ---
    except pyodbc.Error as db_err:
        sqlstate = db_err.args[0]; message = str(db_err.args[1])
        print(f"数据库错误 SQLSTATE: {sqlstate}\n消息: {message}"); traceback.print_exc()
        messagebox.showerror("数据库错误", f"处理数据时发生数据库错误。\nSQLSTATE: {sqlstate}\n错误: {message}")
        # <<< MODIFIED >>>
        status_var.set(f"[{report_name_part}] 报告生成失败 (数据库错误)。")
        if conn: conn.rollback() # 回滚事务
        return False # 表示失败
    except PermissionError as pe:
         target_path = final_filepath or output_filepath_original # 确定出错时操作的文件
         print(f"文件权限错误: {pe}"); traceback.print_exc()
         messagebox.showerror("文件错误", f"文件访问权限错误，可能是文件被占用或无权限。\n错误: {pe}\n文件: {target_path}")
         # <<< MODIFIED >>>
         status_var.set(f"[{report_name_part}] 报告生成失败 (文件权限错误)。")
         if conn: conn.rollback()
         return False
    except FileNotFoundError as fnf_err:
         print(f"文件未找到错误: {fnf_err}"); traceback.print_exc()
         messagebox.showerror("文件错误", f"未能找到所需文件。\n错误: {fnf_err}")
         # <<< MODIFIED >>>
         status_var.set(f"[{report_name_part}] 报告生成失败 (文件未找到)。")
         if conn: conn.rollback()
         return False
    except Exception as e:
        print(f"生成 [{report_name_part}] 报告时发生错误: {e}"); traceback.print_exc()
        messagebox.showerror("执行错误", f"生成报告时发生意外错误。\n错误类型: {type(e).__name__}\n错误详情: {e}")
        # <<< MODIFIED >>>
        status_var.set(f"[{report_name_part}] 报告生成失败 ({type(e).__name__})。")
        if conn: conn.rollback() # 回滚事务
        return False # 表示失败
    finally:
        # 确保数据库连接关闭
        if conn:
            try: conn.close(); print("数据库连接已关闭。")
            except Exception as close_err: print(f"关闭数据库连接时出错: {close_err}")
        # 确保按钮重新启用
        enable_buttons() # 调用全局函数启用按钮
        root.update_idletasks() # 更新 GUI
# <<<--- 函数修改结束 --- >>>

# --- 新增: T+1 齐套率报告生成核心逻辑 ---
# --- T+1 齐套率报告生成核心逻辑 (Excel计划源, 供应仅为库存, 需求为T+1净需求) ---
def generate_t1_completeness_report(
        workshop_codes_db_filter,
        workshop_name_excel_filter,
        source_excel_path,
        header_row_excel_num,
        report_options_passed_in,
        gui_base_date # <<< 新增参数：接收从GUI传入的基准日期
):
    """
    (修改版)
    生成指定车间明天 (T+1) 生产计划的物料齐套率报告。
    1. 从源Excel读取T+1的生产计划 (通过read_t1_excel_plan_data, 已按Excel的“单别”列筛选)。
    2. 对筛选后的计划，按“单别”+“工单单号”分组，取T+1父项计划产量的最大值。
    3. 计算这些计划所需的各零部件的T+1总净需求量。
    4. 供应方仅为当前仓库库存。
    5. 对比两者，判断齐套性。
    """
    report_name_part = workshop_name_excel_filter
    print(f"--- 开始生成 [{report_name_part}] T+1 齐套率报告 (Excel源, 单别筛选, 最大计划量) ---")
    # <<< 修改日志，显示传入的基准日期 >>>
    print(f"  使用GUI设定的基准日期进行T+1计算: {gui_base_date.strftime('%Y-%m-%d')}")
    print(f"  DB单别代码(参考): {workshop_codes_db_filter}, Excel筛选用显示名: '{workshop_name_excel_filter}'")
    print(f"  源Excel: {source_excel_path}, Excel表头行(1-based): {header_row_excel_num}")

    conn = None
    final_report_columns_ordered = [
        '工单单别(Excel)', '工单编号(Excel)', '父项产品(Excel)', '父项T+1代表计划产量(Excel)',
        '主供应商名称',                                          # <<< 新增：列定义 >>>
        '料件品号', '料件品名', '料件规格', '单位', 'BOM单位用量',
        '子件T+1毛需求(基于代表产量)', '子件为T+1工单已领数量(DB)', '子件T+1净需求(基于代表产量)',
        '料件当前仓库库存', '料件T+1总净需求(所有代表产量汇总)', '料件供需差额',
        '料件当日齐套状态', '工单当日齐套状态(基于代表产量)'
    ]

    def create_empty_final_df_with_cols():
        return pd.DataFrame(columns=final_report_columns_ordered)

    try:
        if 'disable_buttons' in globals(): disable_buttons()
        if 'status_var' in globals() and hasattr(status_var, 'set'):
            status_var.set(f"为[{report_name_part}]生成T+1齐套报告(新逻辑)...")
        if 'root' in globals() and hasattr(root, 'update_idletasks'):
            root.update_idletasks()

        report_t1_date = gui_base_date + datetime.timedelta(days=1) # 新代码：基于GUI基准日期计算T+1
        header_row_index_0based = header_row_excel_num - 1

        # 1. 从Excel读取T+1生产计划 (调用修改后的函数)
        # 返回的 df_t1_excel_plan_raw_filtered 应包含:
        # 'WorkOrderType_Excel_Plan' (来自Excel '单别'列),
        # 'WorkOrderNum_Excel', 'ParentProductID_Excel', 'ParentPlannedQty_T1_Excel'
        df_t1_excel_plan_raw_filtered = read_t1_excel_plan_data(
            source_excel_path,
            workshop_name_excel_filter,
            report_t1_date, # <<< 修改：传递基于GUI基准日计算的T+1日期
            header_row_index_0based,
            report_options_config=report_options_passed_in
        )

        if df_t1_excel_plan_raw_filtered.empty:
            messagebox.showinfo("无T+1 Excel计划",
                                f"源Excel [{os.path.basename(source_excel_path)}] 中，按车间 '{workshop_name_excel_filter}' (对应相关'单别') 筛选后，无明日有效生产计划或所有计划数量为0。")
            df_final_report_data = create_empty_final_df_with_cols()
            # Proceed to generate empty Excel
        else:
            print(f"  成功从Excel读取并按'单别'初步筛选得到 {len(df_t1_excel_plan_raw_filtered)} 条T+1计划行。")

            # 2. 新增数据整合步骤：按“单别”+“工单单号”取最大T+1计划量
            print(f"  整合计划：按'单别'+'工单单号'取最大T+1计划量...")

            # 确保 ParentPlannedQty_T1_Excel 是数值型，便于取max
            df_t1_excel_plan_raw_filtered['ParentPlannedQty_T1_Excel'] = pd.to_numeric(
                df_t1_excel_plan_raw_filtered['ParentPlannedQty_T1_Excel'], errors='coerce'
            ).fillna(0)

            grouping_keys_for_max = ['WorkOrderType_Excel_Plan', 'WorkOrderNum_Excel']

            df_t1_plan_consolidated = df_t1_excel_plan_raw_filtered.groupby(
                grouping_keys_for_max,
                as_index=False
            ).agg(
                ParentPlannedQty_T1_Excel_Max=('ParentPlannedQty_T1_Excel', 'max'),
                ParentProductID_Excel=('ParentProductID_Excel', 'first')  # 每个工单只有一个产品品号
                # 如果需要从 df_t1_excel_plan_raw_filtered 保留其他列到整合后的表，也在这里用 'first' 等聚合
            )

            # 重命名聚合后的数量列，以匹配后续代码的期望（如果后续代码仍用 ParentPlannedQty_T1_Excel）
            # 或者直接在后续代码中使用 ParentPlannedQty_T1_Excel_Max
            df_t1_plan_consolidated.rename(
                columns={'ParentPlannedQty_T1_Excel_Max': 'RepresentativeParentQty_T1'},
                inplace=True
            )

            # 过滤掉那些最大计划量仍然为0的
            df_t1_plan_consolidated = df_t1_plan_consolidated[
                df_t1_plan_consolidated['RepresentativeParentQty_T1'] > 0].copy()

            if df_t1_plan_consolidated.empty:
                messagebox.showinfo("无有效T+1计划量", f"根据“取最大值”规则处理后，未发现有效的T+1生产计划量大于0。")
                df_final_report_data = create_empty_final_df_with_cols()
                # Proceed to generate empty Excel
            else:
                print(
                    f"  整合后得到 {len(df_t1_plan_consolidated)} 条唯一的 '单别+工单单号' T+1生产计划 (使用最大计划量)。")

                conn = pyodbc.connect(PYODBC_CONN_STRING, autocommit=False)
                print("  数据库连接成功。")

                # 3. 计算T+1零部件需求 (基于整合后的、使用最大计划量的计划)
                print(f"  计算T+1零部件需求 (基于整合后的计划)...")
                t1_component_requirements_list = []
                skipped_for_bom_factor_count = 0

                for _, consolidated_row in df_t1_plan_consolidated.iterrows():
                    wo_type_plan = consolidated_row['WorkOrderType_Excel_Plan']  # 来自Excel的'单别'
                    wo_num_plan = consolidated_row['WorkOrderNum_Excel']
                    parent_pid_plan = consolidated_row['ParentProductID_Excel']
                    # 使用的是整合后的代表性最大计划量
                    representative_parent_qty_t1 = consolidated_row['RepresentativeParentQty_T1']

                    # DB查询：TA001是工单单别，TA002是工单编号，TA006是产品品号
                    sql_get_wo_components = f"""
                    SELECT 
                        TA.TA001 AS DB_WorkOrderType, /* 工单单别 */
                        ISNULL(TA.TA015, 0) AS DB_OriginalParentPlannedQty, /* 制令数量 MOCTA.TA015 */
                        TB.TB003 AS ComponentID, /* 料件品号 MOCTB.TB003 */
                        ISNULL(MB.MB002, '') AS MaterialName, 
                        ISNULL(MB.MB003, '') AS MaterialSpec, 
                        ISNULL(MB.MB004, '') AS Unit,
                        ISNULL(TB.TB004, 0) AS DB_CompGrossRequirement, /* 应发数量 MOCTB.TB004 */
                        ISNULL(TB.TB005, 0) AS DB_CompIssuedQuantity,   /* 已发数量 MOCTB.TB005 */
                        ISNULL(MA.MA002, '') AS SupplierName           /* <<< 新增：供应商名称 >>> */
                    FROM MOCTA AS TA
                    INNER JOIN MOCTB TB ON TA.TA001 = TB.TB001 AND TA.TA002 = TB.TB002
                    LEFT JOIN INVMB MB ON TB.TB003 = MB.MB001
                    LEFT JOIN PURMA MA ON MB.MB032 = MA.MA001        /* <<< 新增：连接供应商主文件 >>> */
                    WHERE TA.TA001 = ? AND TA.TA002 = ? AND TA.TA006 = ?
                      AND TA.TA011 IN ('1','2','3') /* 工单状态1=未生产,2=已发料,3=生产中 */
                      AND ISNULL(TB.TB004, 0) > 0; /* 只考虑BOM中应发数量大于0的子件 */
                    """
                    params_wo_components = [wo_type_plan, wo_num_plan, parent_pid_plan]
                    df_components_from_db = pd.read_sql(sql_get_wo_components, conn, params=params_wo_components)

                    if df_components_from_db.empty:
                        print(f"    注意: 计划行 (单别 {wo_type_plan}, 工单 {wo_num_plan}, 父项 {parent_pid_plan}), "
                              f"在DB中无匹配工单/BOM信息或状态不符。")
                        continue

                    for _, comp_db_row in df_components_from_db.iterrows():
                        db_original_parent_qty = comp_db_row['DB_OriginalParentPlannedQty']
                        db_comp_gross_req = comp_db_row['DB_CompGrossRequirement']

                        bom_unit_usage = 0
                        if db_original_parent_qty > 0 and db_comp_gross_req > 0:
                            bom_unit_usage = db_comp_gross_req / db_original_parent_qty
                        else:
                            print(
                                f"    警告: 工单 {comp_db_row['DB_WorkOrderType']}-{wo_num_plan}, 子件 {comp_db_row['ComponentID']}: "
                                f"无法计算BOM单位用量 (DB原父项计划 {db_original_parent_qty}, DB组件毛需求 {db_comp_gross_req})。此子件跳过。")
                            skipped_for_bom_factor_count += 1
                            continue

                        # 基于代表性的T+1父项计划量计算子件毛需求
                        comp_gross_req_t1 = representative_parent_qty_t1 * bom_unit_usage
                        comp_issued_for_wo = comp_db_row['DB_CompIssuedQuantity']  # 该工单该子项已领料量

                        # T+1子件净需求 = (T+1父项代表计划量 * BOM单位用量) - 该工单该子项已领料量
                        comp_net_demand_t1_line = comp_gross_req_t1 - comp_issued_for_wo

                        if comp_net_demand_t1_line > 0:  # 只关心净需求大于0的
                            t1_component_requirements_list.append({
                                'WorkOrderType_Excel': wo_type_plan,
                                'WorkOrderNum_Excel': wo_num_plan,
                                'ParentProductID_Excel': parent_pid_plan,
                                'RepresentativeParentQty_T1': representative_parent_qty_t1,
                                'ComponentID': comp_db_row['ComponentID'],
                                'MaterialName': comp_db_row['MaterialName'],
                                'MaterialSpec': comp_db_row['MaterialSpec'],
                                'Unit': comp_db_row['Unit'],
                                'SupplierName': comp_db_row['SupplierName'], # <<< 新增：供应商名称 >>>
                                'BOMUnitUsage': bom_unit_usage,
                                'CompGrossReq_T1_BasedOnRepQty': comp_gross_req_t1,
                                'CompIssued_ForWO_DB': comp_issued_for_wo,
                                'CompNetDemand_T1_Line_BasedOnRepQty': comp_net_demand_t1_line
                            })

                if skipped_for_bom_factor_count > 0:
                    messagebox.showwarning("T+1计算警告",
                                           f"{skipped_for_bom_factor_count} 个子件行因无法确定BOM单位用量而被跳过计算。\n详情请查看控制台日志。")

                if not t1_component_requirements_list:
                    messagebox.showinfo("无T+1净需求", f"根据T+1代表性计划产量并扣除已领料后，明日无物料净需求。")
                    df_final_report_data = create_empty_final_df_with_cols()
                else:
                    df_t1_demand_lines = pd.DataFrame(t1_component_requirements_list)

                    # 4. 计算每个料件的T+1总净需求 (汇总所有工单对该料件的净需求)
                    df_total_comp_t1_net_demand = df_t1_demand_lines.groupby('ComponentID', as_index=False)[
                        'CompNetDemand_T1_Line_BasedOnRepQty'  # 使用基于代表产量的净需求
                    ].sum()
                    df_total_comp_t1_net_demand.rename(
                        columns={'CompNetDemand_T1_Line_BasedOnRepQty': 'TotalNetDemand_T1_Comp_AllWO'}, inplace=True)
                    print(f"    计算得到 {len(df_total_comp_t1_net_demand)} 种零部件的T+1总净需求。")

                    # 5. 获取相关零部件的当前仓库库存 (供应方)
                    relevant_components_for_inv_check = df_total_comp_t1_net_demand['ComponentID'].unique().tolist()
                    df_inventory_data = pd.DataFrame(columns=['ComponentID', 'CurrentWarehouseStock'])
                    if relevant_components_for_inv_check:
                        comp_placeholders_inv = ', '.join('?' * len(relevant_components_for_inv_check))
                        sql_get_inventory = f"""
                        SELECT INV.MC001 AS ComponentID, SUM(ISNULL(INV.MC007, 0)) AS CurrentWarehouseStock
                        FROM INVMC AS INV LEFT JOIN CMSMC AS CMS ON CMS.MC001 = INV.MC002
                        WHERE CMS.MC005 = 'Y' AND (INV.MC002 IN ('700', '710')) /* 根据您的库存库位调整 */
                          AND INV.MC001 IN ({comp_placeholders_inv})
                        GROUP BY INV.MC001;
                        """
                        df_inventory_data = pd.read_sql(sql_get_inventory, conn,
                                                        params=relevant_components_for_inv_check)
                    print(f"    查询到 {len(df_inventory_data)} 种相关零部件的仓库库存。")

                    # 6. 合并库存与T+1总需求，判断零部件当日齐套
                    df_component_supply_demand = pd.merge(df_total_comp_t1_net_demand, df_inventory_data,
                                                          on='ComponentID', how='left')
                    df_component_supply_demand['CurrentWarehouseStock'].fillna(0, inplace=True)
                    df_component_supply_demand['SupplyDemandGap_Comp'] = df_component_supply_demand[
                                                                             'CurrentWarehouseStock'] - \
                                                                         df_component_supply_demand[
                                                                             'TotalNetDemand_T1_Comp_AllWO']
                    df_component_supply_demand['IsCompDailySufficient_T1'] = df_component_supply_demand[
                                                                                 'SupplyDemandGap_Comp'] >= 0
                    df_component_supply_demand['CompDailyStatus_T1'] = df_component_supply_demand[
                        'IsCompDailySufficient_T1'].apply(lambda x: '当日齐套' if x else '当日缺料')
                    print("    零部件当日齐套状态计算完成。")

                    # 7. 将零部件当日齐套状态关联回T+1需求行项目，并判断工单当日齐套
                    df_final_report_data = pd.merge(
                        df_t1_demand_lines,
                        df_component_supply_demand[[
                            'ComponentID', 'CurrentWarehouseStock', 'TotalNetDemand_T1_Comp_AllWO',
                            'SupplyDemandGap_Comp', 'IsCompDailySufficient_T1', 'CompDailyStatus_T1'
                        ]],
                        on='ComponentID',
                        how='left'
                    )
                    # 对于可能没有库存或总需求的组件（理论上不应发生，因为是left join到demand_lines），填充默认值
                    fill_na_for_supply_demand_cols = {
                        'CurrentWarehouseStock': 0, 'TotalNetDemand_T1_Comp_AllWO': 0, 'SupplyDemandGap_Comp': 0,
                        'IsCompDailySufficient_T1': False, 'CompDailyStatus_T1': '当日缺料(无汇总信息)'
                    }
                    for col, val_fill in fill_na_for_supply_demand_cols.items():
                        if col in df_final_report_data.columns:
                            df_final_report_data[col].fillna(val_fill, inplace=True)
                        else:
                            df_final_report_data[col] = val_fill  # 如果列完全缺失则添加

                    # 判断工单齐套性：基于其所有子件是否都“当日齐套”
                    # 分组键是工单的唯一标识：'WorkOrderType_Excel', 'WorkOrderNum_Excel', 'ParentProductID_Excel'
                    df_final_report_data['IsWorkOrderDailySufficient_T1'] = df_final_report_data.groupby(
                        ['WorkOrderType_Excel', 'WorkOrderNum_Excel', 'ParentProductID_Excel']
                    )['IsCompDailySufficient_T1'].transform('all')
                    df_final_report_data['WorkOrderDailyStatus_T1'] = df_final_report_data[
                        'IsWorkOrderDailySufficient_T1'].apply(
                        lambda x: '当日齐套' if x else '当日缺料')
                    print("    工单当日齐套状态计算完成。")

                    # 8. 整理最终报告列顺序和命名
                    rename_final_cols = {
                        'WorkOrderType_Excel': '工单单别(Excel)',
                        'WorkOrderNum_Excel': '工单编号(Excel)',
                        'ParentProductID_Excel': '父项产品(Excel)',
                        'RepresentativeParentQty_T1': '父项T+1代表计划产量(Excel)',
                        'SupplierName': '主供应商名称',  # <<< 新增：重命名映射 >>>

                        # 添加以下物料信息的映射：
                        'ComponentID': '料件品号',
                        'MaterialName': '料件品名',
                        'MaterialSpec': '料件规格',
                        'Unit': '单位',

                        # 保留其他已有的映射：
                        'BOMUnitUsage': 'BOM单位用量',
                        'CompGrossReq_T1_BasedOnRepQty': '子件T+1毛需求(基于代表产量)',
                        'CompIssued_ForWO_DB': '子件为T+1工单已领数量(DB)',
                        'CompNetDemand_T1_Line_BasedOnRepQty': '子件T+1净需求(基于代表产量)',
                        'CurrentWarehouseStock': '料件当前仓库库存',
                        'TotalNetDemand_T1_Comp_AllWO': '料件T+1总净需求(所有代表产量汇总)',
                        'SupplyDemandGap_Comp': '料件供需差额',
                        'CompDailyStatus_T1': '料件当日齐套状态',
                        'WorkOrderDailyStatus_T1': '工单当日齐套状态(基于代表产量)'
                    }
                    df_final_report_data.rename(columns=rename_final_cols, inplace=True)

                    # 确保所有定义的 final_report_columns_ordered 列都存在，不够的补空
                    for col_name_ordered in final_report_columns_ordered:
                        if col_name_ordered not in df_final_report_data.columns:
                            df_final_report_data[col_name_ordered] = None
                    df_final_report_data = df_final_report_data[final_report_columns_ordered]

        # --- 生成 Excel 文件 (无论是否有数据) ---
        # desktop_path = Path("E:/Users/Desktop") # 或者从全局获取，或者硬编码
        # <<< 保持您原来的桌面路径获取逻辑，这里假设 desktop_path 已定义 >>>
        # --- 修改：硬编码输出路径 ---
        desktop_path = Path(r"E:\Users\Desktop")
        try:
            if not desktop_path.exists():
                desktop_path.mkdir(parents=True, exist_ok=True)
            elif not os.access(desktop_path, os.W_OK):
                print(f"警告：无写入权限到目录 {desktop_path}")
        except Exception as dir_err:
            messagebox.showerror("目录错误(T+1齐套)", f"无法创建或访问目标目录: {desktop_path} ({dir_err})")
            # Ensure buttons are enabled in case of early exit
            if 'enable_buttons' in globals(): enable_buttons()
            return False

        timestamp_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')
        safe_report_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in report_name_part)
        # temp_filename = f"TEMP_{safe_report_name}_T+1齐套(新逻辑)_{timestamp_str}.xlsx"
        final_filename = f"{safe_report_name}_T+1齐套报告(新逻辑)_{datetime.date.today().strftime('%m%d')}.xlsx"
        # output_filepath_temp = desktop_path / temp_filename
        final_filepath = desktop_path / final_filename

        print(f"  将齐套率数据(新逻辑)写入最终文件: {final_filepath}")
        # 直接写入最终文件，不再使用临时文件然后重命名，除非格式化过程非常耗时或易错
        df_final_report_data.to_excel(final_filepath, index=False, engine='openpyxl')

        # --- Excel格式化 (与您原代码类似) ---
        print("  应用 Excel 格式化 (新逻辑)...")
        wb_process = openpyxl.load_workbook(final_filepath)  # 打开刚保存的文件
        ws_process = wb_process.active
        if ws_process is None: raise ValueError("无法加载 Excel 文件的工作表进行格式化。")

        header_font = Font(name='微软雅黑', size=9, bold=True)
        data_font = Font(name='微软雅黑', size=8)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')  # 淡红色
        green_fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')  # 淡绿色

        current_excel_headers = [cell.value for cell in ws_process[1]]  # 获取实际写入的表头

        for col_idx, header_text in enumerate(current_excel_headers, 1):
            cell = ws_process.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = center_align

        # 定义哪些列需要特殊格式化
        numeric_cols_to_format = {  # 这些列名是 final_report_columns_ordered 中的名字
            '父项T+1代表计划产量(Excel)', 'BOM单位用量',
            '子件T+1毛需求(基于代表产量)', '子件为T+1工单已领数量(DB)',
            '子件T+1净需求(基于代表产量)', '料件当前仓库库存',
            '料件T+1总净需求(所有代表产量汇总)', '料件供需差额'
        }
        status_comp_col_name = '料件当日齐套状态'
        status_wo_col_name = '工单当日齐套状态(基于代表产量)'

        # 初始化列宽 (基于表头)
        col_widths_calc = {
            idx + 1: max(sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in str(h_text)) + 2, 10)
            for idx, h_text in enumerate(current_excel_headers)
        }

        if ws_process.max_row > 1:  # 如果有数据行
            for row_idx in range(2, ws_process.max_row + 1):
                comp_status_val_for_format = None
                wo_status_val_for_format = None
                try:  # 安全地获取状态列的单元格值，以防列名不匹配或列不存在
                    if status_comp_col_name in current_excel_headers:
                        comp_status_val_for_format = ws_process.cell(row=row_idx, column=current_excel_headers.index(
                            status_comp_col_name) + 1).value
                    if status_wo_col_name in current_excel_headers:
                        wo_status_val_for_format = ws_process.cell(row=row_idx, column=current_excel_headers.index(
                            status_wo_col_name) + 1).value
                except ValueError:
                    pass  # 列名未在 current_excel_headers 中找到

                for col_idx, header_text_iter in enumerate(current_excel_headers, 1):
                    cell = ws_process.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.alignment = left_align  # 默认左对齐
                    cell.number_format = '@'  # 默认文本

                    if header_text_iter in numeric_cols_to_format:
                        cell.alignment = right_align
                        if header_text_iter == 'BOM单位用量':
                            cell.number_format = '#,##0.0000'  # BOM用量多几位小数
                        else:
                            cell.number_format = '#,##0'
                    elif header_text_iter == status_comp_col_name or header_text_iter == status_wo_col_name:
                        cell.alignment = center_align

                    # 条件格式填充颜色
                    if header_text_iter == status_comp_col_name and comp_status_val_for_format:
                        if '缺料' in str(comp_status_val_for_format):
                            cell.fill = red_fill
                        elif '齐套' in str(comp_status_val_for_format):
                            cell.fill = green_fill
                    if header_text_iter == status_wo_col_name and wo_status_val_for_format:
                        if '缺料' in str(wo_status_val_for_format):
                            cell.fill = red_fill
                        elif '齐套' in str(wo_status_val_for_format):
                            cell.fill = green_fill

                    # 更新列宽计算
                    if cell.value is not None:
                        cell_content_str = str(cell.value)
                        current_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_content_str)
                        col_widths_calc[col_idx] = max(col_widths_calc.get(col_idx, 0), current_len + 2)

        # 应用计算后的列宽
        # (与您原代码的列宽设置逻辑类似，可根据需要调整固定宽度和最大宽度限制)
        fixed_col_widths_map = {  # 列名: 宽度 (示例)
            '工单单别(Excel)': 10, '工单编号(Excel)': 12, '父项产品(Excel)': 15,
            '料件品名': 30, '料件规格': 25,
        }
        max_auto_width = 50  # 自动调整列宽的最大值

        for c_idx_apply, width_val in col_widths_calc.items():
            col_letter_apply = get_column_letter(c_idx_apply)
            header_for_fixed_width = current_excel_headers[c_idx_apply - 1]
            if header_for_fixed_width in fixed_col_widths_map:
                ws_process.column_dimensions[col_letter_apply].width = fixed_col_widths_map[header_for_fixed_width]
            else:
                ws_process.column_dimensions[col_letter_apply].width = min(width_val, max_auto_width)

        ws_process.freeze_panes = 'A2'  # 冻结首行

        # --- 单元格合并逻辑 (与您原代码类似，基于最终报告的列名) ---
        print("    尝试合并工单识别相关的单元格...")
        if ws_process.max_row > 2:  # 至少需要表头+2行数据才可能合并
            try:
                # 这些是最终报告中代表工单唯一性的列 (这里假设我们只合并前3列)
                cols_to_merge_names = ['工单单别(Excel)', '工单编号(Excel)', '父项产品(Excel)']
                cols_to_merge_indices = []
                for col_name_mg in cols_to_merge_names:
                    if col_name_mg in current_excel_headers:
                        cols_to_merge_indices.append(current_excel_headers.index(col_name_mg) + 1)
                    else:
                        print(f"      警告: 合并列 '{col_name_mg}' 在Excel表头中未找到，跳过此列合并。")

                if len(cols_to_merge_indices) > 0:  # 确保至少有一列可以合并
                    start_merge_row = 2
                    for current_eval_row in range(2, ws_process.max_row + 2):  # +2 处理最后一块
                        # 以第一列作为主要判断是否新块的依据 (可以根据实际情况选择更稳定的判断列)
                        # 例如，工单编号(Excel) 通常是好的判断依据
                        key_col_for_block_check_idx = -1
                        if '工单编号(Excel)' in current_excel_headers:
                            key_col_for_block_check_idx = current_excel_headers.index('工单编号(Excel)') + 1
                        elif cols_to_merge_indices:  # Fallback to the first mergeable column
                            key_col_for_block_check_idx = cols_to_merge_indices[0]

                        if key_col_for_block_check_idx != -1:  # Make sure we have a key column
                            val_at_start_block = ws_process.cell(row=start_merge_row,
                                                                 column=key_col_for_block_check_idx).value
                            val_at_current_eval_row = None
                            if current_eval_row <= ws_process.max_row:
                                val_at_current_eval_row = ws_process.cell(row=current_eval_row,
                                                                          column=key_col_for_block_check_idx).value

                            if current_eval_row > ws_process.max_row or val_at_current_eval_row != val_at_start_block:
                                end_merge_row = current_eval_row - 1
                                if end_merge_row > start_merge_row:  # 只有块大于1行才合并
                                    for col_idx_to_merge in cols_to_merge_indices:
                                        try:
                                            ws_process.merge_cells(start_row=start_merge_row,
                                                                   start_column=col_idx_to_merge,
                                                                   end_row=end_merge_row, end_column=col_idx_to_merge)
                                            merged_cell_fmt = ws_process.cell(row=start_merge_row,
                                                                              column=col_idx_to_merge)
                                            merged_cell_fmt.alignment = Alignment(horizontal='left', vertical='top',
                                                                                  wrap_text=True)
                                        except Exception as e_merge_specific:
                                            print(
                                                f"      警告: 合并单元格(行{start_merge_row}-{end_merge_row}, 列{col_idx_to_merge})失败: {e_merge_specific}")
                                if current_eval_row <= ws_process.max_row:
                                    start_merge_row = current_eval_row
                        else:  # No key column found for block check, skip merging
                            print("      警告: 未找到合适的判断列来执行合并，跳过合并。")
                            break
                print("    单元格合并操作完成。")
            except Exception as e_merge_outer:
                print(f"      执行单元格合并时发生错误: {e_merge_outer}")
        else:
            print("      数据行不足，跳过单元格合并。")
        # --- 结束合并单元格逻辑 ---

        wb_process.save(final_filepath)  # 保存带格式的Excel
        print(f"  格式化完成，文件已保存到: {final_filepath}")
        wb_process.close()

        # 清理临时文件 (如果之前是先写临时文件再重命名的话)
        # if os.path.exists(output_filepath_temp): os.remove(output_filepath_temp)

        if 'status_var' in globals() and hasattr(status_var, 'set'):
            status_var.set(f"[{report_name_part}] T+1 齐套报告(新逻辑)已成功生成！")
        messagebox.showinfo("完成 (T+1 新逻辑)", f"T+1 齐套报告(新逻辑)已生成并保存到：\n{final_filepath}")
        return True

    except pyodbc.Error as db_err_main:
        sqlstate = db_err_main.args[0];
        message_text = str(db_err_main.args[1])
        print(f"DB错误(T+1齐套): {message_text}");
        traceback.print_exc()
        messagebox.showerror("数据库错误 (T+1齐套)", f"SQLSTATE: {sqlstate}\n错误: {message_text}")
        if 'status_var' in globals() and hasattr(status_var, 'set'): status_var.set(
            f"[{report_name_part}] T+1 报告失败 (数据库错误)。")
    except PermissionError as pe_main:
        target_path_pe = final_filepath if 'final_filepath' in locals() and final_filepath else source_excel_path
        print(f"文件权限错误(T+1齐套): {pe_main}");
        traceback.print_exc()
        messagebox.showerror("文件错误 (T+1齐套)", f"文件权限错误: {pe_main}\n文件: {target_path_pe}")
        if 'status_var' in globals() and hasattr(status_var, 'set'): status_var.set(
            f"[{report_name_part}] T+1 报告失败 (文件权限错误)。")
    except Exception as e_main:
        print(f"生成 T+1 报告(新逻辑)时发生错误: {e_main}");
        traceback.print_exc()
        messagebox.showerror("执行错误 (T+1 新逻辑)", f"生成T+1报告时发生意外错误: {type(e_main).__name__}\n{e_main}")
        if 'status_var' in globals() and hasattr(status_var, 'set'): status_var.set(
            f"[{report_name_part}] T+1 报告失败 ({type(e_main).__name__})。")
    finally:
        if conn:
            try:
                conn.close(); print("  数据库连接已关闭。")
            except Exception:
                pass
        if 'enable_buttons' in globals(): enable_buttons()
        if 'root' in globals() and hasattr(root, 'update_idletasks'): root.update_idletasks()

    return False  # 如果try块中没有成功返回True，则这里返回False


# <<< MODIFIED: Logic to read dropdown and call the new function >>>
def run_generate_all_shortages():
    """
    处理“生成全部缺料报告”按钮点击事件。
    现在根据“选择报告车间”下拉框的选项来筛选数据。
    """
    selected_display_name = report_scope_var.get() # 获取下拉框当前选中的文本

    if not selected_display_name:
        messagebox.showerror("选择错误", "请先在下拉框中选择要生成报告的车间范围！")
        return

    # 从全局字典 report_options 获取选项信息
    if selected_display_name not in report_options:
        messagebox.showerror("内部错误", f"未知的报告车间选项: {selected_display_name}")
        return

    # 获取映射的代码列表和报告名称部分
    codes_from_options, name_part = report_options[selected_display_name]

    # --- 确定筛选逻辑和参数 ---
    codes_to_pass = []      # 最终传递给核心函数的代码列表
    use_exclusion = False   # 是否使用 NOT IN 模式

    if selected_display_name == "深加工车间 (多)":
        # 深加工逻辑: 排除组装一和组装二的代码
        # 需要排除的单别列表
        codes_to_pass = ['511', '521', '516','526']
        use_exclusion = True # 设置为排除模式
        # name_part 已经从 report_options 中获取为 "深加工车间"
    elif selected_display_name in ["组装一 (511, 521)", "组装二 (516, 526)"]:
        # 组装一或组装二: 使用它们自己的代码进行包含筛选
        codes_to_pass = codes_from_options # 直接使用选项中的代码列表
        use_exclusion = False # 设置为包含模式
        # name_part 已经从 report_options 中获取
    else:
        # 处理其他可能未预料到的情况，例如 report_options 被扩展时
        messagebox.showerror("逻辑错误", f"无法确定 '{selected_display_name}' 的筛选逻辑。请检查 report_options 配置。")
        return

    # --- 确认操作，提示用户将基于所选车间生成 ---
    confirm = messagebox.askyesno(
        title="确认操作",
        # 更清晰地告知用户筛选条件
        message=f"确定要为选定的车间范围生成缺料报告吗？\n\n范围: [{name_part}]\n筛选单别: {codes_to_pass} ({'排除这些' if use_exclusion else '仅包含这些'})\n\n此操作将查询数据库并生成一个 Excel 文件到桌面。"
    )

    if confirm:
        disable_buttons()
        # 更新状态信息，反映实际操作
        status_var.set(f"正在为 [{name_part}] 生成缺料报告...")
        root.update_idletasks() # 确保 GUI 更新状态

        # 调用修改后的核心函数，传递筛选参数
        success = generate_all_shortages_report(codes_to_pass, name_part, use_exclusion=use_exclusion)

        # 状态栏信息由核心函数内部在完成后设置
        # enable_buttons() # 核心函数内部的 finally 块会重新启用按钮
        root.update_idletasks() # 再次确保 GUI 更新
    else:
        status_var.set("生成缺料报告操作已取消。")
        root.update_idletasks() # 更新状态

# --- 全局: 定义车间选项和映射关系 (用于模拟发料报告) ---
report_options = {
    # 显示名称: ([数据库车间代码列表], 文件名/源Excel筛选名部分)
    "组装一 (511, 521)": (['511', '521'], "组装一"),
    # --- 修改下面这行 ---
    "组装二 (516, 526)": (['516', '526'], "组装二"), # 添加 526，并更新显示名称
    # --- 修改结束 ---
    "深加工车间 (多)": (['514', '517', '513', '515'], "深加工车间") # 深加工本身的定义不变，但排除逻辑会变
}
# 确保更新 report_option_names 列表（如果它在 report_options 定义之后）
report_option_names = list(report_options.keys()) # 获取更新后的显示名称列表

# --- T+1 齐套率检查的触发函数 (修改版，以适配新的Excel计划源) ---
# --- T+1 齐套率检查的触发函数 (修改版，以适配新的Excel计划源) ---
def trigger_t1_completeness_check():
    """
    处理 "生成 (T+1)" 按钮点击事件。
    新逻辑:
    1. 从源Excel获取明天T+1的【父项生产计划】。
    2. 调用核心逻辑函数 generate_t1_completeness_report 进行齐套分析。
    """
    selected_display_name = report_scope_var.get()
    source_excel_file = source_file_var.get()
    header_row_num_excel = header_row_var.get()
    base_date_str_from_gui = base_date_var.get()  # <<< 新增：获取基准日期字符串

    if not selected_display_name:
        messagebox.showerror("错误", "请先选择报告车间！")
        return
    if not source_excel_file:
        messagebox.showerror("错误 (T+1)", "请先选择“源数据文件(维护用)”以获取T+1计划！")
        return
    if not os.path.exists(source_excel_file):
        messagebox.showerror("错误 (T+1)", f"选择的“源数据文件(维护用)”不存在：\n{source_excel_file}")
        return
    if header_row_num_excel <= 0:
        messagebox.showerror("错误 (T+1)", "源文件表头行号必须大于0。")
        return

    # <<< 新增：校验和转换基准日期 >>>
    if not base_date_str_from_gui:
        messagebox.showerror("错误 (T+1)", "请输入模拟发料的基准日期！")
        return
    try:
        base_date_for_t1_calc = datetime.datetime.strptime(base_date_str_from_gui, '%Y-%m-%d').date()
    except ValueError:
        messagebox.showerror("错误 (T+1)", "基准日期格式不正确，请输入YYYY-MM-DD 格式！")
        return
    # <<< 新增结束 >>>

    if selected_display_name in report_options:  # report_options 是全局定义的字典
        codes_for_db, name_part_for_excel_filter_and_report = report_options[selected_display_name]

        try:
            print(f"--- 触发 T+1 齐套率报告 (Excel计划源) ---")
            print(f"  GUI设定的基准日期: {base_date_for_t1_calc.strftime('%Y-%m-%d')}")  # <<< 新增日志
            print(f"  源 Excel 文件: {os.path.basename(source_excel_file)}")
            print(f"  Excel表头行号(1-based): {header_row_num_excel}")
            print(f"  筛选车间标识(Excel用): '{name_part_for_excel_filter_and_report}'")
            print(f"  筛选车间代码(DB用): {codes_for_db}")

            # 调用修改后的核心函数，传递所需参数
            generate_t1_completeness_report(
                codes_for_db,
                name_part_for_excel_filter_and_report,
                source_excel_file,
                header_row_num_excel,
                report_options,
                base_date_for_t1_calc  # <<< 新增：传递基准日期
            )
        except Exception as e:
            messagebox.showerror("执行错误 (T+1 Trigger)",
                                 f"执行 T+1 流程时发生意外错误。\n错误: {type(e).__name__}: {e}")
            traceback.print_exc()
            status_var.set("T+1 报告生成失败 (触发器内部错误)。")
            if 'enable_buttons' in globals(): enable_buttons()
            if 'root' in globals() and hasattr(root, 'update_idletasks'): root.update_idletasks()
    else:
        messagebox.showerror("错误", f"无效的报告范围选项: {selected_display_name}")

# --- 用于触发报告生成的辅助函数 (修改版，处理 T+2 特殊逻辑 + 周日顺延) ---
def trigger_report_generation(offset_days):
    """
    根据选择的报告范围和偏移天数触发相应的缺料报告生成流程。
    对于 T+2 报告，会先从源Excel文件筛选 T+0 到 T+3/T+4 (如果含周日则顺延)
    范围内的工单，再进行数据库查询和报告生成。
    对于 T+7 和 T+15 报告，直接基于数据库车间代码进行查询和生成。

    Args:
        offset_days (int): 报告的偏移天数 (2, 7, 或 15)。
    """
    selected_display_name = report_scope_var.get()
    if not selected_display_name:
        messagebox.showerror("错误", "请先选择报告车间！")
        return

    # --- T+2 特殊逻辑: 从源 Excel 筛选工单 ---
    if offset_days == 2:
        source_excel_file = source_file_var.get()
        base_date_str = base_date_var.get() # T+2 也需要基准日期

        # 输入验证
        if not source_excel_file:
            messagebox.showerror("错误 (T+2)", "请先选择“源数据文件(维护用)”以生成 T+2 报告！")
            return
        if not os.path.exists(source_excel_file):
             messagebox.showerror("错误 (T+2)", f"选择的“源数据文件(维护用)”不存在：\n{source_excel_file}")
             return
        if not base_date_str:
            messagebox.showerror("错误 (T+2)", "请输入模拟发料的基准日期！")
            return
        try:
            base_date_for_calc = datetime.datetime.strptime(base_date_str, '%Y-%m-%d').date()
        except ValueError:
             messagebox.showerror("错误 (T+2)", "基准日期格式不正确，请输入YYYY-MM-DD 格式！")
             return

        # 获取选择的车间信息
        if selected_display_name in report_options:
            # codes 是数据库筛选用的单别列表, name_part 是Excel筛选用的名称标识和报告文件名部分
            codes, name_part = report_options[selected_display_name]

            # --- 计算 Excel 筛选日期范围 (T+0 到 T+3/T+4，根据是否含周日调整) ---
            # --- 计算 Excel 筛选日期范围 (T-2 到 T+3/T+4，根据是否含周日调整) ---
            start_date_for_plan = base_date_for_calc - datetime.timedelta(days=2)  # T-2 (基准日往前2天)
            initial_end_date = base_date_for_calc + datetime.timedelta(days=3)  # 默认 T+3
            final_end_date_for_plan = initial_end_date  # 先假定最终结束日期为 T+3

            print(f"--- 触发 T+2 报告生成流程 (周日顺延逻辑) ---")
            print(f"  基准日期: {base_date_for_calc.strftime('%Y-%m-%d')}")
            print(
                f"  初始 T-2 到 T+3 日期范围: {start_date_for_plan.strftime('%Y-%m-%d')} 到 {initial_end_date.strftime('%Y-%m-%d')}")

            # 检查 T-2 到 T+3 是否包含周日
            contains_sunday = False
            for i in range(-2, 4):  # 检查 T-2, T-1, T+0, T+1, T+2, T+3
                check_date = base_date_for_calc + datetime.timedelta(days=i)
                # datetime.date.isoweekday(): Monday is 1 and Sunday is 7
                if check_date.isoweekday() == 7:
                    contains_sunday = True
                    # Using %A to get the full name of the weekday (locale-dependent)
                    try:
                        # Attempt to format with weekday name, handle potential locale issues
                        weekday_name = check_date.strftime('%A')
                    except Exception: # Catch broader exceptions just in case
                        weekday_name = "Sunday" # Fallback
                    print(f"  检测到周日: {check_date.strftime('%Y-%m-%d')} ({weekday_name})")
                    break # 找到一个周日就够了, 无需继续检查

            # 如果包含周日，则将结束日期顺延一天 (到 T+4)
            if contains_sunday:
                final_end_date_for_plan = base_date_for_calc + datetime.timedelta(days=4) # 调整为 T+4
                print(f"  包含周日，结束日期顺延至 T+4: {final_end_date_for_plan.strftime('%Y-%m-%d')}")
            else:
                 print(f"  T+0 到 T+3 不包含周日，结束日期保持为 T+3。")
            # --- 日期计算逻辑结束 ---

            print(f"  源 Excel 文件: {os.path.basename(source_excel_file)}")
            print(f"  筛选车间标识 (用于 Excel): '{name_part}'")
            print(f"  最终 Excel 筛选日期范围: {start_date_for_plan.strftime('%Y-%m-%d')} 到 {final_end_date_for_plan.strftime('%Y-%m-%d')}") # 打印最终使用的范围

            # 调用辅助函数从 Excel 获取特定工单列表
            # 注意: get_work_orders_from_plan_excel 内部会根据 name_part 处理 深加工 的特殊逻辑
            # 它接收开始和结束日期，并筛选该日期范围内有数据的工单
            specific_work_orders = get_work_orders_from_plan_excel(
                source_excel_file,
                name_part,
                start_date_for_plan,      # 开始日期不变 (T+0)
                final_end_date_for_plan   # 使用计算得到的最终结束日期 (T+3 或 T+4)
            )

            # --- 处理 get_work_orders_from_plan_excel 的返回结果 ---
            if specific_work_orders is None:
                # 如果返回 None，表示在 get_work_orders_from_plan_excel 内部发生了错误，且已弹窗提示
                status_var.set("T+2 报告生成失败 (源数据文件处理错误)。")
                root.update_idletasks() # 更新GUI状态
                return # 中断执行
            elif not specific_work_orders:
                 # 如果返回空列表，表示源文件筛选后无符合条件的工单
                 messagebox.showinfo("无匹配工单 (T-2~T+3/T+4)",
                                     f"在源数据文件 '{os.path.basename(source_excel_file)}' 中，\n"
                                     f"对于车间 '{name_part}'，未找到在日期范围 "
                                     f"{start_date_for_plan.strftime('%Y-%m-%d')} 到 {final_end_date_for_plan.strftime('%Y-%m-%d')} 内\n"
                                     f"有非零数据的工单。\n\n将不生成缺料报告。")
                 status_var.set(f"T+2: 对于 '{name_part}' 未找到符合 T-2~T+3/T+4 条件的工单。") # 更新状态
                 root.update_idletasks() # 更新GUI状态
                 return # 中断执行
            else:
                # 如果返回包含工单号的列表，则继续生成报告
                print(f"  从源文件筛选到 {len(specific_work_orders)} 个特定工单，准备生成报告...")
                # 调用主报告生成函数 generate_simulation_report
                # 参数:
                #   offset=2:  用于文件名和内部数据库日期过滤逻辑(T+3)
                #   codes:     数据库查询时可能用到的车间代码列表 (来自 report_options)
                #   name_part: 用于最终报告文件名部分 (来自 report_options)
                #   specific_work_orders: 从 Excel 筛选出的工单列表，用于精确过滤数据库查询结果
                generate_simulation_report(2, codes, name_part, specific_work_orders=specific_work_orders)
        else:
            # 如果下拉框选中的值不在 report_options 字典中
            messagebox.showerror("错误", f"无效的报告范围选项: {selected_display_name}")

    # --- T+7 / T+15 逻辑 (保持不变) ---
    else: # offset_days is 7 or 15
        base_date_str = base_date_var.get() # T+7/T+15 也需要基准日期
        if not base_date_str:
            messagebox.showerror(f"错误 (T+{offset_days})", "请输入模拟发料的基准日期！")
            return
        try:
            # 仅验证日期格式是否正确，不需要保存日期对象在此处
            datetime.datetime.strptime(base_date_str, '%Y-%m-%d').date()
        except ValueError:
             messagebox.showerror(f"错误 (T+{offset_days})", "基准日期格式不正确，请输入YYYY-MM-DD 格式！")
             return

        # 获取选择的车间信息
        if selected_display_name in report_options:
            codes, name_part = report_options[selected_display_name]
            print(f"--- 触发 T+{offset_days} 报告生成流程 ---")
            print(f"  筛选车间代码 (用于 DB): {codes}")
            print(f"  报告名称部分: '{name_part}'")
            # T+7/T+15 不从 Excel 筛选工单，因此 specific_work_orders 传 None
            generate_simulation_report(offset_days, codes, name_part, specific_work_orders=None)
        else:
            # 如果下拉框选中的值不在 report_options 字典中
            messagebox.showerror("错误", f"无效的报告范围选项: {selected_display_name}")


# --- GUI 界面设置 ---
root = tk.Tk()
# <<< MODIFIED >>> 更新版本号和功能描述
root.title("工单数据处理工具 v2.7 (添加全部缺料报告)")

# Tkinter 变量
source_file_var = tk.StringVar()
target_file_var = tk.StringVar()
workshop_var = tk.StringVar()
master_plan_file_var = tk.StringVar()
status_var = tk.StringVar()
status_var.set("请选择文件并操作。")
header_row_var = tk.IntVar(value=2) # <<< 注意: 已按之前请求修改默认值为 2 >>>
base_date_var = tk.StringVar(value=datetime.date.today().strftime('%Y-%m-%d'))
report_scope_var = tk.StringVar()

# --- 样式配置 ---
style = ttk.Style()
try:
     available_themes = style.theme_names()
     if 'vista' in available_themes: style.theme_use('vista')
     elif 'clam' in available_themes: style.theme_use('clam')
except tk.TclError:
     print("Info: Could not set 'vista' or 'clam' theme. Using default.")

style.configure('.', font=('微软雅黑', 9))
style.configure("TButton", padding=6, relief="flat")
style.map("TButton", background=[('active', '#e0e0e0'), ('disabled', '#f0f0f0')])
style.configure("TLabel", padding=5)
style.configure("Bold.TLabel", font=('微软雅黑', 9, "bold"))
style.configure("TEntry", padding=(5, 3), relief="solid", borderwidth=1)
style.configure("TCombobox", padding=(5, 3))
style.map("TCombobox", fieldbackground=[('readonly', 'white')])
style.configure("TSpinbox", padding=(5, 3), relief="solid", borderwidth=1)
style.configure("Status.TLabel", foreground="#00008B", background="#f0f0f0", padding=5)

# --- 主框架 ---
main_frame = ttk.Frame(root, padding="10 10 10 10")
main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
root.columnconfigure(0, weight=1); root.rowconfigure(0, weight=1)

# --- 控件布局 ---
row_num = 0 # 行计数器

# --- 维护数据部分 ---
ttk.Label(main_frame, text=" 数据维护 (源Excel -> 目标XLSM) ", style="Bold.TLabel").grid(row=row_num, column=0, columnspan=3, sticky=tk.W, pady=(0,5))
row_num += 1
ttk.Label(main_frame, text="源数据文件 (维护/T+2用):").grid(row=row_num, column=0, sticky=tk.W, padx=5, pady=3)
source_entry = ttk.Entry(main_frame, textvariable=source_file_var, width=60)
source_entry.grid(row=row_num, column=1, sticky=(tk.W, tk.E), padx=5, pady=3)
source_button = ttk.Button(main_frame, text="浏览...", command=lambda: select_file(source_file_var), width=10)
source_button.grid(row=row_num, column=2, sticky=tk.E, padx=5, pady=3)
row_num += 1
ttk.Label(main_frame, text="源文件表头行号:").grid(row=row_num, column=0, sticky=tk.W, padx=5, pady=3)
# 使用已设置默认值为 2 的 header_row_var
header_spinbox = ttk.Spinbox(main_frame, from_=1, to=50, textvariable=header_row_var, width=5, font=('微软雅黑', 9), justify=tk.CENTER)
header_spinbox.grid(row=row_num, column=1, sticky=tk.W, padx=5, pady=3)
row_num += 1
ttk.Label(main_frame, text="选择处理车间 (维护用):").grid(row=row_num, column=0, sticky=tk.W, padx=5, pady=3)
workshop_options_maintain = ["组装一", "组装二", "深加工车间"]
workshop_combobox_maintain = ttk.Combobox(main_frame, textvariable=workshop_var, values=workshop_options_maintain, state="readonly", width=18, font=('微软雅黑', 9))
workshop_combobox_maintain.grid(row=row_num, column=1, sticky=tk.W, padx=5, pady=3)
if workshop_options_maintain: workshop_var.set(workshop_options_maintain[0])
row_num += 1
ttk.Label(main_frame, text="目标 XLSM 文件 (维护用):").grid(row=row_num, column=0, sticky=tk.W, padx=5, pady=3)
target_entry = ttk.Entry(main_frame, textvariable=target_file_var, width=60)
target_entry.grid(row=row_num, column=1, sticky=(tk.W, tk.E), padx=5, pady=3)
target_button = ttk.Button(main_frame, text="浏览...", command=lambda: select_xlsm_file(target_file_var), width=10)
target_button.grid(row=row_num, column=2, sticky=tk.E, padx=5, pady=3)
row_num += 1
# 维护数据按钮 (赋值给全局变量)
btn_process = ttk.Button(main_frame, text="执行数据维护", command=run_single_process)
btn_process.grid(row=row_num, column=1, columnspan=2, sticky=tk.E, padx=5, pady=(5, 10))
row_num += 1
# 分隔线
ttk.Separator(main_frame, orient='horizontal').grid(row=row_num, column=0, columnspan=3, sticky='ew', pady=10)
row_num += 1

# --- 模拟发料报告部分 ---
ttk.Label(main_frame, text=" 模拟发料缺料报告 (数据库 -> 新Excel) ", style="Bold.TLabel").grid(row=row_num, column=0, columnspan=3, sticky=tk.W, pady=(5,5))
row_num += 1
ttk.Label(main_frame, text="模拟发料基准日期:").grid(row=row_num, column=0, sticky=tk.W, padx=5, pady=3)
base_date_entry = ttk.Entry(main_frame, textvariable=base_date_var, width=15, justify=tk.CENTER)
base_date_entry.grid(row=row_num, column=1, sticky=tk.W, padx=5, pady=3)
ttk.Label(main_frame, text="(T+2报告也使用上方“源数据文件”筛选)", font=('微软雅黑', 8, 'italic')).grid(row=row_num, column=1, sticky=tk.W, padx=(140, 5), pady=3)
row_num += 1
ttk.Label(main_frame, text="选择报告车间:").grid(row=row_num, column=0, sticky=tk.W, padx=5, pady=3)
report_scope_combobox = ttk.Combobox(main_frame, textvariable=report_scope_var, values=report_option_names, state="readonly", width=25, font=('微软雅黑', 9))
report_scope_combobox.grid(row=row_num, column=1, sticky=tk.W, padx=5, pady=3)
if report_option_names: report_scope_var.set(report_option_names[0])
row_num += 1
# 模拟发料按钮框架 (T+x)
sim_button_frame = ttk.Frame(main_frame)
sim_button_frame.grid(row=row_num, column=1, columnspan=2, sticky=tk.W, padx=0, pady=(5, 10))
# T+x 按钮 (赋值给全局变量)
btn_simulate_t1 = ttk.Button(sim_button_frame, text="生成 (T+1)", width=10, command=lambda: trigger_t1_completeness_check()) # 修改按钮文字为 "T+1", 调整宽度 (可选)
btn_simulate_t1.pack(side=tk.LEFT, padx=2)
btn_simulate_t2 = ttk.Button(sim_button_frame, text="生成 (T+2)", width=10, command=lambda: trigger_report_generation(2))
btn_simulate_t2.pack(side=tk.LEFT, padx=2)
btn_simulate_t7 = ttk.Button(sim_button_frame, text="生成 (T+7)", width=10, command=lambda: trigger_report_generation(7))
btn_simulate_t7.pack(side=tk.LEFT, padx=2)
btn_simulate_t15 = ttk.Button(sim_button_frame, text="生成 (T+15)", width=10, command=lambda: trigger_report_generation(15))
btn_simulate_t15.pack(side=tk.LEFT, padx=2)

# <<< 新增：全部缺料报告按钮框架和按钮 >>>
row_num += 1 # 移到下一行
all_shortage_button_frame = ttk.Frame(main_frame)
# 让它和上面的 T+x 按钮左对齐
all_shortage_button_frame.grid(row=row_num, column=1, columnspan=2, sticky=tk.W, padx=0, pady=(0, 10)) # pady top=0
# 全部缺料按钮 (赋值给全局变量)
btn_generate_all = ttk.Button(all_shortage_button_frame, text="生成全部工单缺料报告", command=run_generate_all_shortages)
btn_generate_all.pack(side=tk.LEFT, padx=2)
# <<< 新增结束 >>>

row_num += 1
# 分隔线
ttk.Separator(main_frame, orient='horizontal').grid(row=row_num, column=0, columnspan=3, sticky='ew', pady=10)
row_num += 1

# --- 更新总计划部分 ---
ttk.Label(main_frame, text=" 更新总计划 (Excel -> 数据库) ", style="Bold.TLabel").grid(row=row_num, column=0, columnspan=3, sticky=tk.W, pady=(5,5))
row_num += 1
ttk.Label(main_frame, text="选择'总计划'文件:").grid(row=row_num, column=0, sticky=tk.W, padx=5, pady=3)
master_plan_entry = ttk.Entry(main_frame, textvariable=master_plan_file_var, width=60)
master_plan_entry.grid(row=row_num, column=1, sticky=(tk.W, tk.E), padx=5, pady=3)
master_plan_button = ttk.Button(main_frame, text="浏览...", command=lambda: select_master_plan_file(master_plan_file_var), width=10)
master_plan_button.grid(row=row_num, column=2, sticky=tk.E, padx=5, pady=3)
row_num += 1
# 更新总计划按钮 (赋值给全局变量)
btn_update_plan = ttk.Button(main_frame, text="更新总计划到数据库", command=run_update_master_plan)
btn_update_plan.grid(row=row_num, column=1, columnspan=2, sticky=tk.E, padx=5, pady=(5, 10))
row_num += 1
# 分隔线
ttk.Separator(main_frame, orient='horizontal').grid(row=row_num, column=0, columnspan=3, sticky='ew', pady=10)
row_num += 1

# --- 状态栏标签 ---
status_label = ttk.Label(main_frame, textvariable=status_var, style="Status.TLabel", anchor=tk.W)
status_label.grid(row=row_num, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0), ipady=3)

# 配置主框架列权重，让第1列（输入框和控件列）可以拉伸
main_frame.columnconfigure(1, weight=1)

# --- 初始检查和启动 ---
def initial_button_state_check():
    # 确保调用修改后的 enable_buttons
    enable_buttons()

    if not libraries_ok:
        status_var.set("错误: 启动时检测到缺少必要的库。请根据提示安装后重启。")
        try:
            if status_label.winfo_exists():
                status_label.config(foreground="red", background="#ffe0e0")
        except tk.TclError:
            print("警告: 初始化检查时无法配置状态栏样式。")
            pass
    else:
        status_var.set("就绪。请选择文件并执行操作。")
        try:
             if status_label.winfo_exists():
                status_label.config(foreground="#00008B", background="#f0f0f0")
        except tk.TclError:
            pass

# 延迟执行初始检查，确保窗口已完全加载
root.after(100, initial_button_state_check)

# 启动主循环
root.mainloop()
